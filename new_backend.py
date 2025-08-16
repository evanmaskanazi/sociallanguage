"""
Enhanced Therapeutic Companion Backend
With PostgreSQL, Authentication, Role-Based Access, Client Reports, and Password Reset
"""
import os
import re
import random
import string
import secrets
import json
import time
import uuid
import html
import bleach
import redis
import smtplib
import openpyxl
import jwt
import logging
import traceback
import signal
from pathlib import Path
from datetime import datetime, date, timedelta
from functools import wraps, lru_cache
from threading import Thread
from io import BytesIO

# Flask imports
from flask import (Flask, request, jsonify, send_file, session,
                   render_template, redirect, url_for, make_response,
                   flash, Response, g, abort)
from flask_cors import CORS
from flask_bcrypt import Bcrypt
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from flask_wtf.csrf import CSRFProtect, generate_csrf
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from markupsafe import escape, Markup

# Database imports
from sqlalchemy import text, and_, or_, func

# Email imports
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Excel imports
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF imports
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Cryptography import (FIXED - was missing)
from cryptography.fernet import Fernet
import base64


# Image processing (optional - you can remove this block if not needed)
# from PIL import Image
# import io

# === LOGGING CONFIGURATION (MOVED UP - before it's used) ===
class StructuredFormatter(logging.Formatter):
    def format(self, record):
        log_obj = {
            'timestamp': self.formatTime(record),
            'level': record.levelname,
            'logger': record.name,
            'message': record.getMessage(),
            'module': record.module,
            'function': record.funcName,
            'line': record.lineno
        }

        # Add extra fields if present
        if hasattr(record, 'request_id'):
            log_obj['request_id'] = record.request_id
        if hasattr(record, 'user_id'):
            log_obj['user_id'] = record.user_id
        if hasattr(record, 'extra_data'):
            log_obj.update(record.extra_data)

        return json.dumps(log_obj)


# Set up logger
logger = logging.getLogger('therapy_companion')
handler = logging.StreamHandler()
handler.setFormatter(StructuredFormatter())
logger.addHandler(handler)
logger.setLevel(logging.INFO)

# Disable werkzeug logging in production
if os.environ.get('PRODUCTION'):
    logging.getLogger('werkzeug').setLevel(logging.WARNING)

# === CONSTANTS AND CONFIGURATION ===
JWT_SECRET = os.environ.get('SECRET_KEY', 'your-secret-key')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Define allowed origins
ALLOWED_ORIGINS = [
    'https://therapy-companion.onrender.com',
    'http://localhost:5000',  # For development
    'http://127.0.0.1:5000'  # For development
]

# Dangerous patterns for XSS protection
DANGEROUS_PATTERNS = [
    r'<script',
    r'javascript:',
    r'onerror=',
    r'onclick=',
    r'onload=',
    r'<iframe',
    r'<object',
    r'<embed',
    r'vbscript:',
    r'data:text/html'
]

# === ENCRYPTION KEY SETUP (FIXED - after logger is defined) ===
ENCRYPTION_KEY = os.environ.get('FIELD_ENCRYPTION_KEY')
if not ENCRYPTION_KEY:
    if os.environ.get('PRODUCTION'):
        raise ValueError("FIELD_ENCRYPTION_KEY must be set in production")
    else:
        # Generate a development key
        ENCRYPTION_KEY = Fernet.generate_key()
        logger.warning("Using generated encryption key - DO NOT use in production!")
else:
    # Ensure it's bytes
    ENCRYPTION_KEY = ENCRYPTION_KEY if isinstance(ENCRYPTION_KEY, bytes) else ENCRYPTION_KEY.encode()

fernet = Fernet(ENCRYPTION_KEY)

# === REDIS CLIENT SETUP ===
redis_client = redis.from_url(os.environ.get('REDIS_URL', 'redis://localhost:6379'))


# === UTILITY FUNCTIONS ===
def encrypt_field(data):
    """Encrypt sensitive field data"""
    if not data:
        return None
    if isinstance(data, str):
        data = data.encode()
    return fernet.encrypt(data).decode()


def decrypt_field(encrypted_data):
    """Decrypt sensitive field data"""
    if not encrypted_data:
        return None
    return fernet.decrypt(encrypted_data.encode()).decode()


def safe_render(template, **kwargs):
    """Automatically escape all template variables"""
    safe_kwargs = {}
    for key, value in kwargs.items():
        if isinstance(value, str):
            safe_kwargs[key] = escape(value)
        else:
            safe_kwargs[key] = value
    return render_template(template, **safe_kwargs)


def sanitize_html(text):
    """Remove all HTML tags from text"""
    if not text:
        return text
    clean = re.compile('<.*?>')
    return re.sub(clean, '', str(text))


def sanitize_input(text, allow_html=False):
    """Sanitize user input to prevent XSS"""
    if not text:
        return text

    if allow_html:
        allowed_tags = ['b', 'i', 'u', 'em', 'strong', 'p', 'br']
        allowed_attributes = {}
        text = bleach.clean(text, tags=allowed_tags, attributes=allowed_attributes, strip=True)
    else:
        text = html.escape(text)

    text = text.replace('\x00', '')  # Remove null bytes
    return text


def validate_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_date(date_string):
    """Validate date format and range"""
    try:
        date_obj = datetime.strptime(date_string, '%Y-%m-%d').date()
        if date_obj < date(2020, 1, 1) or date_obj > date.today() + timedelta(days=30):
            return False
        return True
    except:
        return False


def validate_cors_origin():
    """Manually validate CORS origins to prevent case-sensitivity issues"""
    origin = request.headers.get('Origin')

    if not origin:
        return True  # No origin header, allow

    # Normalize origin to lowercase for comparison
    origin_lower = origin.lower()

    # Check against allowed origins (case-insensitive)
    for allowed in ALLOWED_ORIGINS:
        if origin_lower == allowed.lower():
            return True

    return False


def timeout(seconds):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            def timeout_handler(signum, frame):
                raise TimeoutError(f"Function {func.__name__} timed out after {seconds} seconds")

            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(seconds)
            try:
                result = func(*args, **kwargs)
            finally:
                signal.alarm(0)
            return result

        return wrapper

    return decorator


def cache_response(timeout=300):
    """Cache response for specified timeout (default 5 minutes)"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if request.method != 'GET':
                return f(*args, **kwargs)

            cache_key = f"{request.path}?{request.query_string.decode()}"
            response = f(*args, **kwargs)

            if isinstance(response, tuple):
                resp, status = response[0], response[1]
            else:
                resp = response

            if hasattr(resp, 'headers'):
                resp.headers['Cache-Control'] = f'private, max-age={timeout}'

            return response

        return decorated_function

    return decorator


# === CACHE MANAGER CLASS ===
class CacheManager:
    """Centralized cache management with namespace support"""

    def __init__(self, redis_client, default_ttl=3600):
        self.redis = redis_client
        self.default_ttl = default_ttl

    def _make_key(self, namespace, key):
        """Create namespaced cache key"""
        return f"{namespace}:{key}"

    def get(self, namespace, key):
        """Get value from cache"""
        if not self.redis:
            return None
        try:
            data = self.redis.get(self._make_key(namespace, key))
            if data:
                return json.loads(data)
        except Exception as e:
            logger.error(f"Cache get error: {e}")
        return None

    def set(self, namespace, key, value, ttl=None):
        """Set value in cache"""
        if not self.redis:
            return False
        try:
            self.redis.setex(
                self._make_key(namespace, key),
                ttl or self.default_ttl,
                json.dumps(value)
            )
            return True
        except Exception as e:
            logger.error(f"Cache set error: {e}")
        return False

    def delete(self, namespace, key):
        """Delete from cache"""
        if not self.redis:
            return
        try:
            self.redis.delete(self._make_key(namespace, key))
        except Exception as e:
            logger.error(f"Cache delete error: {e}")

    def invalidate_pattern(self, pattern):
        """Invalidate all keys matching pattern"""
        if not self.redis:
            return
        try:
            keys = self.redis.keys(pattern)
            if keys:
                self.redis.delete(*keys)
        except Exception as e:
            logger.error(f"Cache invalidate error: {e}")


# Initialize cache manager
cache = CacheManager(redis_client)


def cached_endpoint(namespace, ttl=3600, key_func=None):
    """Decorator for caching endpoint responses"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if key_func:
                cache_key = key_func(*args, **kwargs)
            else:
                cache_key = f"{f.__name__}:{str(args)}:{str(kwargs)}"

            cached_data = cache.get(namespace, cache_key)
            if cached_data:
                return jsonify(cached_data)

            result = f(*args, **kwargs)

            if isinstance(result, tuple) and result[1] == 200:
                response_data = result[0].get_json()
                cache.set(namespace, cache_key, response_data, ttl)
            elif hasattr(result, 'status_code') and result.status_code == 200:
                response_data = result.get_json()
                cache.set(namespace, cache_key, response_data, ttl)

            return result

        return decorated_function

    return decorator


# === CREATE FLASK APP ===
app = Flask(__name__)

# Apply proxy fix for Render
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Set directories
BASE_DIR = Path(__file__).resolve().parent
app.static_folder = BASE_DIR
app.template_folder = BASE_DIR

# === FLASK CONFIGURATION ===
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=24),
    SESSION_COOKIE_NAME='__Host-session'
)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://localhost/therapy_companion'
).replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# In your Flask app configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Extend session
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Refresh on each request




# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('SMTP_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('SYSTEM_EMAIL')
app.config['MAIL_PASSWORD'] = os.environ.get('SYSTEM_EMAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('SYSTEM_EMAIL')

# Database connection pooling
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 20,
    'pool_recycle': 3600,
    'pool_pre_ping': True,
    'max_overflow': 40,
    'pool_timeout': 30,
    'connect_args': {
        'connect_timeout': 10,
        'options': '-c statement_timeout=30000'
    }
}

# CSRF configuration
app.config['WTF_CSRF_EXEMPT_LIST'] = ['health_check', 'index', 'login_page',
                                      'therapist_dashboard_page', 'client_dashboard_page',
                                      'serve_i18n', 'favicon', 'debug_server_time',
                                      'static', 'login', 'unsubscribe']
app.config['WTF_CSRF_CHECK_DEFAULT'] = False

# === INITIALIZE EXTENSIONS ===
db = SQLAlchemy(app)
migrate = Migrate(app, db)
bcrypt = Bcrypt(app)
CORS(app, supports_credentials=True)

# Initialize CSRF protection
csrf = CSRFProtect()
csrf.init_app(app)

# Initialize Talisman with security settings
Talisman(app,
         force_https=False if app.debug else True,
         strict_transport_security={'max_age': 31536000, 'include_subdomains': True},
         content_security_policy=False,
         frame_options='SAMEORIGIN',
         content_security_policy_nonce_in=[]
         )

# Initialize rate limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["50000 per day", "10000 per hour"],
    storage_uri="memory://"
)


# === REQUEST HANDLERS ===
@app.before_request
def before_request():
    try:
        g.request_id = str(uuid.uuid4())
        g.request_start_time = time.time()

        extra_data = {
            'method': request.method,
            'path': request.path,
            'remote_addr': request.remote_addr,
            'user_agent': request.headers.get('User-Agent', ''),
            'request_id': g.request_id
        }

        logger.info('request_started', extra={'extra_data': extra_data, 'request_id': g.request_id})
    except Exception as e:
        g.request_id = str(uuid.uuid4())
        g.request_start_time = time.time()
        logger.error(f'Error in before_request: {e}')


@app.before_request
def csrf_protect():
    if request.endpoint in app.config['WTF_CSRF_EXEMPT_LIST']:
        return
    if request.method in ["POST", "PUT", "DELETE", "PATCH"]:
        token = None
        if request.is_json:
            token = request.json.get('csrf_token')
        token = token or request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')

        if not token:
            return jsonify({'error': 'CSRF token missing'}), 403


@app.before_request
def check_cors():
    if request.method == 'OPTIONS':
        return

    if not validate_cors_origin():
        logger.warning(f"CORS validation failed for origin: {request.headers.get('Origin')}")
        abort(403, description="CORS validation failed")


@app.before_request
def validate_inputs():
    """Check all inputs for XSS attempts"""
    for key, value in request.values.items():
        if value and isinstance(value, str):
            for pattern in DANGEROUS_PATTERNS:
                if re.search(pattern, value, re.IGNORECASE):
                    logger.warning(f"Potential XSS attempt blocked: {key}={value[:50]}...")
                    return jsonify({'error': 'Invalid input detected'}), 400

    if request.is_json and request.json:
        def check_dict(d):
            for key, value in d.items():
                if isinstance(value, str):
                    for pattern in DANGEROUS_PATTERNS:
                        if re.search(pattern, value, re.IGNORECASE):
                            return False
                elif isinstance(value, dict):
                    if not check_dict(value):
                        return False
            return True

        if not check_dict(request.json):
            return jsonify({'error': 'Invalid input detected'}), 400



@app.before_request
def log_progress_requests():
    if request.path == '/api/client/progress':
        logger.info('progress_request', extra={
            'extra_data': {
                'user_id': request.current_user.id if hasattr(request, 'current_user') and request.current_user.is_authenticated else 'anonymous',
                'path': request.path,
                'method': request.method
            }
        })




@app.after_request
def after_request(response):
    if hasattr(g, 'request_start_time'):
        duration = time.time() - g.request_start_time
    else:
        duration = 0

    extra_data = {
        'method': request.method,
        'path': request.path,
        'status_code': response.status_code,
        'duration_ms': round(duration * 1000, 2) if duration > 0 else 'unknown',
        'request_id': getattr(g, 'request_id', 'unknown')
    }

    if hasattr(request, 'current_user') and request.current_user:
        extra_data['user_id'] = request.current_user.id
        extra_data['user_role'] = request.current_user.role

    logger.info('request_completed',
                extra={'extra_data': extra_data, 'request_id': getattr(g, 'request_id', 'unknown')})

    if hasattr(g, 'request_id'):
        response.headers['X-Request-ID'] = g.request_id

    if request.method == 'GET':
        if request.path.startswith('/api/categories') or request.path.startswith('/api/tracking-categories'):
            response.headers['Cache-Control'] = 'public, max-age=3600'
        elif request.path.endswith('.js') or request.path.endswith('.css'):
            response.headers['Cache-Control'] = 'public, max-age=86400'

    return response


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'

    origin = request.headers.get('Origin')
    if origin and validate_cors_origin():
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRF-Token'

    return response


@app.after_request
def set_csrf_cookie(response):
    if 'csrf_token' not in session:
        session['csrf_token'] = generate_csrf()
    response.set_cookie('csrf_token', session['csrf_token'],
                        secure=app.config['SESSION_COOKIE_SECURE'],
                        httponly=False,
                        samesite='Strict',
                        max_age=86400)
    return response


# === CSRF TOKEN ENDPOINT ===
@app.route('/api/csrf-token', methods=['GET'])
def get_csrf_token():
    return jsonify({'csrf_token': generate_csrf()})



# Elasticsearch configuration
try:
    from elasticsearch import Elasticsearch
    import elasticsearch.helpers
    ELASTICSEARCH_AVAILABLE = True
except ImportError:
    ELASTICSEARCH_AVAILABLE = False
    Elasticsearch = None


class SearchManager:
    """Manage Elasticsearch operations"""

    def __init__(self):
        self.es = None
        if not ELASTICSEARCH_AVAILABLE:
            logger.warning("Elasticsearch not installed - search will use database fallback")
            return
        es_url = os.environ.get('ELASTICSEARCH_URL', 'http://localhost:9200')
        if es_url:
            try:
                self.es = Elasticsearch([es_url], timeout=30, max_retries=3)
                if self.es.ping():
                    logger.info("Elasticsearch connected successfully")
                    self._create_indices()
                else:
                    logger.warning("Elasticsearch ping failed")
                    self.es = None
            except Exception as e:
                logger.error(f"Elasticsearch connection failed: {e}")
                self.es = None

    def _create_indices(self):
        """Create indices if they don't exist"""
        indices = {
            'clients': {
                'mappings': {
                    'properties': {
                        'client_id': {'type': 'integer'},
                        'client_serial': {'type': 'keyword'},
                        'client_name': {'type': 'text'},
                        'therapist_id': {'type': 'integer'},
                        'start_date': {'type': 'date'},
                        'is_active': {'type': 'boolean'},
                        'last_checkin': {'type': 'date'},
                        'checkin_count': {'type': 'integer'},
                        'avg_emotional': {'type': 'float'},
                        'avg_medication': {'type': 'float'},
                        'tags': {'type': 'keyword'},
                        'suggest': {
                            'type': 'completion',
                            'contexts': [{'name': 'therapist_id', 'type': 'category'}]
                        }
                    }
                }
            },
            'checkins': {
                'mappings': {
                    'properties': {
                        'client_id': {'type': 'integer'},
                        'checkin_date': {'type': 'date'},
                        'emotional_value': {'type': 'integer'},
                        'medication_value': {'type': 'integer'},
                        'activity_value': {'type': 'integer'},
                        'categories': {'type': 'nested'},
                        'notes_sentiment': {'type': 'float'}
                    }
                }
            }
        }

        for index_name, index_config in indices.items():
            if not self.es.indices.exists(index=index_name):
                self.es.indices.create(index=index_name, body=index_config)
                logger.info(f"Created Elasticsearch index: {index_name}")

    def index_client(self, client, checkin_stats=None):
        """Index or update client in Elasticsearch"""
        if not self.es:
            return

        try:
            doc = {
                'client_id': client.id,
                'client_serial': client.client_serial,
                'client_name': client.client_name or client.client_serial,
                'therapist_id': client.therapist_id,
                'start_date': client.start_date.isoformat(),
                'is_active': client.is_active,
                'suggest': {
                    'input': [client.client_serial, client.client_name] if client.client_name else [
                        client.client_serial],
                    'contexts': {'therapist_id': str(client.therapist_id)}
                }
            }

            if checkin_stats:
                doc.update(checkin_stats)

            self.es.index(
                index='clients',
                id=client.id,
                body=doc,
                refresh=True
            )
        except Exception as e:
            logger.error(f"Error indexing client {client.id}: {e}")

    def search_clients(self, therapist_id, query, filters=None):
        """Search clients with autocomplete"""
        if not self.es:
            return []

        try:
            # Build search query
            search_body = {
                'query': {
                    'bool': {
                        'must': [
                            {'term': {'therapist_id': therapist_id}}
                        ],
                        'should': [
                            {'match': {'client_name': {'query': query, 'boost': 2}}},
                            {'prefix': {'client_serial': {'value': query.upper()}}},
                            {'match': {'tags': query}}
                        ],
                        'minimum_should_match': 1
                    }
                },
                'suggest': {
                    'client_suggest': {
                        'prefix': query,
                        'completion': {
                            'field': 'suggest',
                            'contexts': {'therapist_id': str(therapist_id)},
                            'size': 10
                        }
                    }
                },
                'size': 20
            }

            # Add filters if provided
            if filters:
                if 'is_active' in filters:
                    search_body['query']['bool']['must'].append(
                        {'term': {'is_active': filters['is_active']}}
                    )
                if 'date_range' in filters:
                    search_body['query']['bool']['must'].append({
                        'range': {
                            'last_checkin': {
                                'gte': filters['date_range']['from'],
                                'lte': filters['date_range']['to']
                            }
                        }
                    })

            result = self.es.search(index='clients', body=search_body)

            # Combine search hits and suggestions
            clients = []
            for hit in result['hits']['hits']:
                clients.append(hit['_source'])

            return clients

        except Exception as e:
            logger.error(f"Error searching clients: {e}")
            return []

    def get_analytics(self, therapist_id, date_range):
        """Get aggregated analytics"""
        if not self.es:
            return {}

        try:
            aggs_body = {
                'query': {
                    'bool': {
                        'must': [
                            {'term': {'therapist_id': therapist_id}},
                            {'range': {'checkin_date': {
                                'gte': date_range['from'],
                                'lte': date_range['to']
                            }}}
                        ]
                    }
                },
                'aggs': {
                    'avg_emotional': {'avg': {'field': 'emotional_value'}},
                    'avg_medication': {'avg': {'field': 'medication_value'}},
                    'checkins_over_time': {
                        'date_histogram': {
                            'field': 'checkin_date',
                            'calendar_interval': 'day'
                        }
                    },
                    'client_activity': {
                        'terms': {
                            'field': 'client_id',
                            'size': 100
                        }
                    }
                }
            }

            result = self.es.search(index='checkins', body=aggs_body, size=0)
            return result['aggregations']

        except Exception as e:
            logger.error(f"Error getting analytics: {e}")
            return {}


# Initialize search manager
search_manager = SearchManager()













# Translation mappings for categories - UPDATED WITH DESCRIPTIONS
CATEGORY_TRANSLATIONS = {
    'en': {
        'Emotion Level': 'Emotion Level',
        'Emotion Level_desc': 'Overall emotional state',
        'Energy': 'Energy',
        'Energy_desc': 'Physical and mental energy levels',
        'Social Activity': 'Social Activity',
        'Social Activity_desc': 'Engagement in social interactions',
        'Sleep Quality': 'Sleep Quality',
        'Sleep Quality_desc': 'Quality of sleep',
        'Anxiety Level': 'Anxiety Level',
        'Anxiety Level_desc': 'Level of anxiety experienced',
        'Motivation': 'Motivation',
        'Motivation_desc': 'Level of motivation and drive',
        'Medication': 'Medication',
        'Medication_desc': 'Medication adherence',
        'Physical Activity': 'Physical Activity',
        'Physical Activity_desc': 'Physical activity level'
    },
    'he': {
        'Emotion Level': 'רמה רגשית',
        'Emotion Level_desc': 'מצב רגשי כללי',
        'Energy': 'אנרגיה',
        'Energy_desc': 'רמות אנרגיה פיזית ומנטלית',
        'Social Activity': 'פעילות חברתית',
        'Social Activity_desc': 'מעורבות באינטראקציות חברתיות',
        'Sleep Quality': 'איכות שינה',
        'Sleep Quality_desc': 'איכות השינה',
        'Anxiety Level': 'רמת חרדה',
        'Anxiety Level_desc': 'רמת החרדה שחווית',
        'Motivation': 'מוטיבציה',
        'Motivation_desc': 'רמת המוטיבציה והדחף',
        'Medication': 'תרופות',
        'Medication_desc': 'היענות לנטילת תרופות',
        'Physical Activity': 'פעילות גופנית',
        'Physical Activity_desc': 'רמת הפעילות הגופנית'
    },
    'ru': {
        'Emotion Level': 'Эмоциональный уровень',
        'Emotion Level_desc': 'Общее эмоциональное состояние',
        'Energy': 'Энергия',
        'Energy_desc': 'Уровни физической и умственной энергии',
        'Social Activity': 'Социальная активность',
        'Social Activity_desc': 'Участие в социальных взаимодействиях',
        'Sleep Quality': 'Качество сна',
        'Sleep Quality_desc': 'Качество сна',
        'Anxiety Level': 'Уровень тревожности',
        'Anxiety Level_desc': 'Испытываемый уровень тревожности',
        'Motivation': 'Мотивация',
        'Motivation_desc': 'Уровень мотивации и стремления',
        'Medication': 'Лекарства',
        'Medication_desc': 'Приверженность лечению',
        'Physical Activity': 'Физическая активность',
        'Physical Activity_desc': 'Уровень физической активности'
    },
    'ar': {
        'Emotion Level': 'المستوى العاطفي',
        'Emotion Level_desc': 'الحالة العاطفية العامة',
        'Energy': 'الطاقة',
        'Energy_desc': 'مستويات الطاقة الجسدية والعقلية',
        'Social Activity': 'النشاط الاجتماعي',
        'Social Activity_desc': 'المشاركة في التفاعلات الاجتماعية',
        'Sleep Quality': 'جودة النوم',
        'Sleep Quality_desc': 'جودة النوم',
        'Anxiety Level': 'مستوى القلق',
        'Anxiety Level_desc': 'مستوى القلق الذي تم اختباره',
        'Motivation': 'الدافع',
        'Motivation_desc': 'مستوى الدافع والحافز',
        'Medication': 'الأدوية',
        'Medication_desc': 'الالتزام بالأدوية',
        'Physical Activity': 'النشاط البدني',
        'Physical Activity_desc': 'مستوى النشاط البدني'
    }
}

# Add other translations for reports
REPORT_TRANSLATIONS = {
    'en': {
        'weekly_report_title': 'Weekly Progress Report',
        'client': 'Client',
        'week': 'Week',
        'daily_checkins': 'Daily Check-ins',
        'date': 'Date',
        'day': 'Day',
        'checkin_time': 'Check-in Time',
        'emotional': 'Emotional',
        'notes': 'Notes',
        'medication': 'Medication',
        'activity': 'Activity',
        'completion': 'Completion',
        'no_checkin': 'No check-in',
        'weekly_summary': 'Weekly Summary',
        'checkin_completion': 'Check-in Completion',
        'days': 'days',
        'average_rating': 'Average Rating',
        'adherence': 'Adherence',
        'weekly_goals': 'Weekly Goals',
        'completion_rate': 'Completion Rate',
        'therapist_notes': 'Therapist Notes',
        'type': 'Type',
        'content': 'Content',
        'status': 'Status',
        'mission': 'MISSION',
        'completed': 'Completed',
        'pending': 'Pending',
        'excellent': 'Excellent',
        'good': 'Good',
        'needs_improvement': 'Needs Improvement',
        'needs_support': 'Needs Support',
        'metric': 'Metric',
        'value': 'Value',
        'percentage': 'Percentage',
        'rating': 'Rating',
        'goal': 'Goal',
        'months': ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December'],
        'needs_encouragement': 'Needs Encouragement'
    },
    'he': {
        'weekly_report_title': 'דוח התקדמות שבועי',
        'client': 'מטופל',
        'week': 'שבוע',
        'daily_checkins': 'צ׳ק-אין יומי',
        'date': 'תאריך',
        'day': 'יום',
        'checkin_time': 'זמן צ׳ק-אין',
        'emotional': 'רגשי',
        'notes': 'הערות',
        'medication': 'תרופות',
        'activity': 'פעילות',
        'completion': 'השלמה',
        'no_checkin': 'אין צ׳ק-אין',
        'weekly_summary': 'סיכום שבועי',
        'checkin_completion': 'השלמת צ׳ק-אין',
        'days': 'ימים',
        'average_rating': 'דירוג ממוצע',
        'adherence': 'היענות',
        'weekly_goals': 'יעדים שבועיים',
        'completion_rate': 'אחוז השלמה',
        'therapist_notes': 'הערות מטפל',
        'type': 'סוג',
        'content': 'תוכן',
        'status': 'סטטוס',
        'mission': 'משימה',
        'completed': 'הושלם',
        'pending': 'ממתין',
        'excellent': 'מצוין',
        'good': 'טוב',
        'needs_improvement': 'דורש שיפור',
        'needs_support': 'זקוק לתמיכה',
        'metric': 'מדד',
        'value': 'ערך',
        'percentage': 'אחוז',
        'rating': 'דירוג',
        'goal': 'יעד',
        'months': ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                   'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'],
        'needs_encouragement': 'זקוק לעידוד'
    },
    'ru': {
        'weekly_report_title': 'Еженедельный отчет о прогрессе',
        'client': 'Клиент',
        'week': 'Неделя',
        'daily_checkins': 'Ежедневные отметки',
        'date': 'Дата',
        'day': 'День',
        'checkin_time': 'Время отметки',
        'emotional': 'Эмоциональное',
        'notes': 'Заметки',
        'medication': 'Лекарства',
        'activity': 'Активность',
        'completion': 'Завершение',
        'no_checkin': 'Нет отметки',
        'weekly_summary': 'Недельная сводка',
        'checkin_completion': 'Завершение отметок',
        'days': 'дней',
        'average_rating': 'Средний рейтинг',
        'adherence': 'Приверженность',
        'weekly_goals': 'Недельные цели',
        'completion_rate': 'Процент выполнения',
        'therapist_notes': 'Заметки терапевта',
        'type': 'Тип',
        'content': 'Содержание',
        'status': 'Статус',
        'mission': 'ЗАДАНИЕ',
        'completed': 'Выполнено',
        'pending': 'В ожидании',
        'excellent': 'Отлично',
        'good': 'Хорошо',
        'needs_improvement': 'Требует улучшения',
        'needs_support': 'Нуждается в поддержке',
        'metric': 'Метрика',
        'value': 'Значение',
        'percentage': 'Процент',
        'rating': 'Рейтинг',
        'goal': 'Цель',
        'months': ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'],
        'needs_encouragement': 'Нуждается в поощрении'
    },
    'ar': {
        'weekly_report_title': 'تقرير التقدم الأسبوعي',
        'client': 'العميل',
        'week': 'الأسبوع',
        'daily_checkins': 'تسجيلات الحضور اليومية',
        'date': 'التاريخ',
        'day': 'اليوم',
        'checkin_time': 'وقت تسجيل الحضور',
        'emotional': 'عاطفي',
        'notes': 'ملاحظات',
        'medication': 'الأدوية',
        'activity': 'النشاط',
        'completion': 'الإكمال',
        'no_checkin': 'لا يوجد تسجيل حضور',
        'weekly_summary': 'الملخص الأسبوعي',
        'checkin_completion': 'إكمال تسجيل الحضور',
        'days': 'أيام',
        'average_rating': 'التقييم المتوسط',
        'adherence': 'الالتزام',
        'weekly_goals': 'الأهداف الأسبوعية',
        'completion_rate': 'معدل الإنجاز',
        'therapist_notes': 'ملاحظات المعالج',
        'type': 'النوع',
        'content': 'المحتوى',
        'status': 'الحالة',
        'mission': 'مهمة',
        'completed': 'مكتمل',
        'pending': 'معلق',
        'excellent': 'ممتاز',
        'good': 'جيد',
        'needs_improvement': 'يحتاج إلى تحسين',
        'needs_support': 'يحتاج إلى دعم',
        'metric': 'المقياس',
        'value': 'القيمة',
        'percentage': 'النسبة المئوية',
        'rating': 'التقييم',
        'goal': 'الهدف',
        'months': ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                   'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'],
        'needs_encouragement': 'يحتاج إلى تشجيع'
    }
}

# Days of week translations
DAYS_TRANSLATIONS = {
    'en': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
    'he': ['יום שני', 'יום שלישי', 'יום רביעי', 'יום חמישי', 'יום שישי', 'שבת', 'יום ראשון'],
    'ru': ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'],
    'ar': ['الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']
}


# ============= DATABASE MODELS =============

class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)

    # Relationships
    therapist = db.relationship('Therapist', backref='user', uselist=False, cascade='all, delete-orphan')
    client = db.relationship('Client', backref='user', uselist=False, cascade='all, delete-orphan')
    session_tokens = db.relationship('SessionToken', backref='user', cascade='all, delete-orphan')
    password_resets = db.relationship('PasswordReset', backref='user', cascade='all, delete-orphan')


class PasswordReset(db.Model):
    __tablename__ = 'password_resets'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    reset_token = db.Column(db.String(255), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Therapist(db.Model):
    __tablename__ = 'therapists'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True)
    license_number = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    organization = db.Column(db.String(255))
    specializations = db.Column(db.JSON)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    clients = db.relationship('Client', backref='therapist', lazy='dynamic')
    notes = db.relationship('TherapistNote', backref='therapist', lazy='dynamic')


class Client(db.Model):
    __tablename__ = 'clients'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), unique=True)
    client_serial = db.Column(db.String(50), unique=True, nullable=False)
    client_name = db.Column(db.String(255), nullable=True)
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    start_date = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    checkins = db.relationship('DailyCheckin', backref='client', lazy='dynamic', cascade='all, delete-orphan')
    tracking_plans = db.relationship('ClientTrackingPlan', backref='client', lazy='dynamic',
                                     cascade='all, delete-orphan')
    goals = db.relationship('WeeklyGoal', backref='client', lazy='dynamic', cascade='all, delete-orphan')
    reminders = db.relationship('Reminder', backref='client', lazy='dynamic', cascade='all, delete-orphan')


class TrackingCategory(db.Model):
    __tablename__ = 'tracking_categories'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    scale_min = db.Column(db.Integer, default=1)
    scale_max = db.Column(db.Integer, default=5)
    is_default = db.Column(db.Boolean, default=False)


class ClientTrackingPlan(db.Model):
    __tablename__ = 'client_tracking_plans'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    category_id = db.Column(db.Integer, db.ForeignKey('tracking_categories.id'))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    category = db.relationship('TrackingCategory', backref='client_plans')


class CustomCategory(db.Model):
    __tablename__ = 'custom_categories'

    id = db.Column(db.Integer, primary_key=True)
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'), nullable=False)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    scale_min = db.Column(db.Integer, default=1)
    scale_max = db.Column(db.Integer, default=5)
    reverse_scoring = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

    # Relationships
    therapist = db.relationship('Therapist', backref='custom_categories')
    client = db.relationship('Client', backref='custom_categories')







class WeeklyGoal(db.Model):
    __tablename__ = 'weekly_goals'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    goal_text = db.Column(db.Text, nullable=False)
    week_start = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    completions = db.relationship('GoalCompletion', backref='goal', lazy='dynamic', cascade='all, delete-orphan')


class DailyCheckin(db.Model):
    __tablename__ = 'daily_checkins'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    checkin_date = db.Column(db.Date, nullable=False)
    checkin_time = db.Column(db.Time, nullable=False)
    emotional_value = db.Column(db.Integer)
    emotional_notes_encrypted = db.Column(db.Text)  # Encrypted
    medication_value = db.Column(db.Integer)
    medication_notes_encrypted = db.Column(db.Text)  # Encrypted
    activity_value = db.Column(db.Integer)
    activity_notes_encrypted = db.Column(db.Text)  # Encrypted
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint('client_id', 'checkin_date'),)

    @property
    def emotional_notes(self):
        return decrypt_field(self.emotional_notes_encrypted)

    @emotional_notes.setter
    def emotional_notes(self, value):
        self.emotional_notes_encrypted = encrypt_field(value)

    @property
    def medication_notes(self):
        return decrypt_field(self.medication_notes_encrypted)

    @medication_notes.setter
    def medication_notes(self, value):
        self.medication_notes_encrypted = encrypt_field(value)

    @property
    def activity_notes(self):
        return decrypt_field(self.activity_notes_encrypted)

    @activity_notes.setter
    def activity_notes(self, value):
        self.activity_notes_encrypted = encrypt_field(value)


class CategoryResponse(db.Model):
    __tablename__ = 'category_responses'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    category_id = db.Column(db.Integer, db.ForeignKey('tracking_categories.id'), nullable=True)
    custom_category_id = db.Column(db.Integer, db.ForeignKey('custom_categories.id'), nullable=True)
    response_date = db.Column(db.Date, nullable=False)
    value = db.Column(db.Integer, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    category = db.relationship('TrackingCategory', backref='responses')
    custom_category = db.relationship('CustomCategory', backref='responses')

    def validate(self):
        """Ensure exactly one category type is set"""
        if (self.category_id is None) == (self.custom_category_id is None):
            raise ValueError("Exactly one of category_id or custom_category_id must be set")

    def save(self, commit=True):
        """Save with validation"""
        self.validate()
        db.session.add(self)
        if commit:
            db.session.commit()
        return self

    @property
    def category_name(self):
        """Get the category name regardless of type"""
        if self.category_id:
            return self.category.name
        elif self.custom_category_id:
            return self.custom_category.name
        return None

    @property
    def is_custom(self):
        """Check if this is a custom category response"""
        return self.custom_category_id is not None

    @classmethod
    def create_for_category(cls, client_id, category_id, response_date, value, notes=''):
        """Create a response for a standard category"""
        return cls(
            client_id=client_id,
            category_id=category_id,
            custom_category_id=None,
            response_date=response_date,
            value=value,
            notes=notes
        )

    @classmethod
    def create_for_custom_category(cls, client_id, custom_category_id, response_date, value, notes=''):
        """Create a response for a custom category"""
        return cls(
            client_id=client_id,
            category_id=None,
            custom_category_id=custom_category_id,
            response_date=response_date,
            value=value,
            notes=notes
        )



class GoalCompletion(db.Model):
    __tablename__ = 'goal_completions'

    id = db.Column(db.Integer, primary_key=True)
    goal_id = db.Column(db.Integer, db.ForeignKey('weekly_goals.id'))
    completion_date = db.Column(db.Date, nullable=False)
    completed = db.Column(db.Boolean, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (db.UniqueConstraint('goal_id', 'completion_date'),)


class Reminder(db.Model):
    __tablename__ = 'reminders'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    reminder_type = db.Column(db.String(50), nullable=False)
    reminder_time = db.Column(db.Time, nullable=False)
    local_reminder_time = db.Column(db.String(5))
    reminder_email = db.Column(db.String(255))
    reminder_language = db.Column(db.String(2), default='en')
    is_active = db.Column(db.Boolean, default=True)
    last_sent = db.Column(db.DateTime)
    day_of_week = db.Column(db.Integer, default=1)

class EmailQueue(db.Model):
    """Queue for email sending with retry logic"""
    __tablename__ = 'email_queue'
    
    id = db.Column(db.Integer, primary_key=True)
    to_email = db.Column(db.String(255), nullable=False)
    subject = db.Column(db.String(500), nullable=False)
    body = db.Column(db.Text, nullable=False)
    html_body = db.Column(db.Text)
    status = db.Column(db.String(50), default='pending')  # pending, processing, sent, failed
    attempts = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    sent_at = db.Column(db.DateTime)
    last_attempt_at = db.Column(db.DateTime)
    error_message = db.Column(db.Text)
    
    def __repr__(self):
        return f'<EmailQueue {self.id}: {self.to_email} - {self.status}>'


class Report(db.Model):
    __tablename__ = 'reports'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    report_type = db.Column(db.String(50), nullable=False)
    week_start = db.Column(db.Date, nullable=False)
    generated_at = db.Column(db.DateTime, default=datetime.utcnow)
    file_path = db.Column(db.Text)
    data = db.Column(db.JSON)


class TherapistNote(db.Model):
    __tablename__ = 'therapist_notes'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'))
    therapist_id = db.Column(db.Integer, db.ForeignKey('therapists.id'))
    note_type = db.Column(db.String(50), default='general')
    content = db.Column(db.Text, nullable=False)
    is_mission = db.Column(db.Boolean, default=False)
    mission_completed = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime)


class SessionToken(db.Model):
    __tablename__ = 'session_tokens'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    token = db.Column(db.String(255), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def cleanup_expired_sessions():
    """Clean up expired session tokens"""
    try:
        expired_count = SessionToken.query.filter(
            SessionToken.expires_at < datetime.utcnow()
        ).delete()
        db.session.commit()
        if expired_count > 0:
            logger.info(f"Cleaned up {expired_count} expired sessions")
    except Exception as e:
        logger.error(f"Error cleaning up sessions: {e}")
        db.session.rollback()



class AuditLog(db.Model):
    __tablename__ = 'audit_logs'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    user_email = db.Column(db.String(255))  # Store email separately for deleted users
    action = db.Column(db.String(100), nullable=False)
    resource_type = db.Column(db.String(50))  # 'client', 'checkin', 'report', etc.
    resource_id = db.Column(db.Integer)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(500))
    details = db.Column(db.JSON)
    phi_accessed = db.Column(db.Boolean, default=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Indexes for compliance queries
    __table_args__ = (
        db.Index('idx_audit_user_timestamp', 'user_id', 'timestamp'),
        db.Index('idx_audit_phi_timestamp', 'phi_accessed', 'timestamp'),
    )


class ConsentRecord(db.Model):
    __tablename__ = 'consent_records'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=False)
    consent_type = db.Column(db.String(50), nullable=False)  # 'data_sharing', 'treatment', 'communication'
    consent_version = db.Column(db.String(20), nullable=False)
    consented = db.Column(db.Boolean, nullable=False)
    consent_date = db.Column(db.DateTime, default=datetime.utcnow)
    ip_address = db.Column(db.String(45))
    withdrawal_date = db.Column(db.DateTime)

    client = db.relationship('Client', backref='consents')










def log_audit(action, resource_type=None, resource_id=None, details=None, phi_accessed=False):
    """Log an audit trail entry for HIPAA compliance"""
    try:
        audit = AuditLog(
            user_id=request.current_user.id if hasattr(request, 'current_user') else None,
            user_email=request.current_user.email if hasattr(request, 'current_user') else 'system',
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')[:500],
            details=details,
            phi_accessed=phi_accessed
        )
        db.session.add(audit)
        db.session.commit()
    except Exception as e:
        logger.error(f"Failed to log audit: {e}")


# ============= AUTHENTICATION HELPERS =============

def generate_token(user_id, role):
    """Generate JWT token"""
    payload = {
        'user_id': user_id,
        'role': role,
        'exp': datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def verify_token(token):
    """Verify JWT token"""
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def require_auth(allowed_roles=None):
    """Authentication decorator that supports both JWT and cookie sessions"""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # First try cookie-based session
            session_token = request.cookies.get('session_token')

            if session_token:
                # Validate session token
                session = SessionToken.query.filter_by(
                    token=session_token
                ).first()

                if session and session.expires_at > datetime.utcnow():
                    user = User.query.get(session.user_id)
                    if user and user.is_active:
                        # Check role permissions
                        if allowed_roles and user.role not in allowed_roles:
                            logger.warning('auth_failed', extra={
                                'extra_data': {
                                    'reason': 'insufficient_permissions',
                                    'user_role': user.role,
                                    'required_roles': allowed_roles,
                                    'request_id': getattr(g, 'request_id', 'unknown')
                                },
                                'request_id': getattr(g, 'request_id', 'unknown')
                            })
                            return jsonify({'error': 'Insufficient permissions'}), 403

                        # Log successful auth
                        logger.info('auth_success', extra={
                            'extra_data': {
                                'user_id': user.id,
                                'role': user.role,
                                'auth_method': 'cookie',
                                'request_id': getattr(g, 'request_id', 'unknown')
                            },
                            'request_id': getattr(g, 'request_id', 'unknown'),
                            'user_id': user.id
                        })

                        # Add user info to request
                        request.current_user = user
                        request.user_id = user.id
                        request.user_role = user.role

                        return f(*args, **kwargs)

            # Fall back to JWT token in Authorization header
            auth_header = request.headers.get('Authorization', '')
            if not auth_header.startswith('Bearer '):
                logger.warning('auth_failed', extra={
                    'extra_data': {
                        'reason': 'missing_bearer_token',
                        'has_cookie': bool(session_token),
                        'request_id': getattr(g, 'request_id', 'unknown')
                    },
                    'request_id': getattr(g, 'request_id', 'unknown')
                })
                return jsonify({'error': 'Invalid authorization header'}), 401

            token = auth_header.replace('Bearer ', '')
            payload = verify_token(token)

            if not payload:
                logger.warning('auth_failed', extra={
                    'extra_data': {
                        'reason': 'invalid_token',
                        'request_id': getattr(g, 'request_id', 'unknown')
                    },
                    'request_id': getattr(g, 'request_id', 'unknown')
                })
                return jsonify({'error': 'Invalid or expired token'}), 401

            # Check if user exists and is active
            user = User.query.get(payload['user_id'])
            if not user or not user.is_active:
                logger.warning('auth_failed', extra={
                    'extra_data': {
                        'reason': 'user_not_found_or_inactive',
                        'user_id': payload['user_id'],
                        'request_id': getattr(g, 'request_id', 'unknown')
                    },
                    'request_id': getattr(g, 'request_id', 'unknown')
                })
                return jsonify({'error': 'User not found or inactive'}), 401

            # Check role permissions
            if allowed_roles and user.role not in allowed_roles:
                logger.warning('auth_failed', extra={
                    'extra_data': {
                        'reason': 'insufficient_permissions',
                        'user_role': user.role,
                        'required_roles': allowed_roles,
                        'request_id': getattr(g, 'request_id', 'unknown')
                    },
                    'request_id': getattr(g, 'request_id', 'unknown')
                })
                return jsonify({'error': 'Insufficient permissions'}), 403

            # Log successful auth
            logger.info('auth_success', extra={
                'extra_data': {
                    'user_id': user.id,
                    'role': user.role,
                    'auth_method': 'jwt',
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown'),
                'user_id': user.id
            })

            # Add user info to request
            request.current_user = user
            request.user_id = user.id
            request.user_role = user.role

            return f(*args, **kwargs)

        return decorated_function

    return decorator


def generate_client_serial():
    """Generate unique client serial number"""
    import uuid

    max_attempts = 10

    for attempt in range(max_attempts):
        # Generate UUID-based serial
        unique_id = str(uuid.uuid4()).replace('-', '').upper()[:12]
        serial = f'C{unique_id}'

        # Check if exists
        existing = Client.query.filter_by(client_serial=serial).first()
        if not existing:
            return serial

    # Fallback with timestamp
    import time
    timestamp = str(int(time.time() * 1000000))[-8:]
    unique_id = str(uuid.uuid4()).replace('-', '').upper()[:4]
    return f'C{unique_id}{timestamp}'


class EmailCircuitBreaker:
    """Circuit breaker for email service with persistent state"""

    def __init__(self, failure_threshold=5, recovery_timeout=60):
        self.failure_threshold = failure_threshold
        self.recovery_timeout = recovery_timeout
        self._load_state()

    def _load_state(self):
        """Load state from database"""
        try:
            with app.app_context():
                result = db.session.execute(
                    text(
                        "SELECT failure_count, last_failure_time, is_open FROM circuit_breaker_state WHERE service = 'email'")
                ).first()

                if result:
                    self.failure_count = result[0]
                    self.last_failure_time = result[1]
                    self.is_open = result[2]
                else:
                    self.failure_count = 0
                    self.last_failure_time = None
                    self.is_open = False
                    self._save_state()
        except:
            # Table might not exist yet
            self.failure_count = 0
            self.last_failure_time = None
            self.is_open = False

    def _save_state(self):
        """Save state to database"""
        try:
            with app.app_context():
                db.session.execute(
                    text("""
                        INSERT INTO circuit_breaker_state (service, failure_count, last_failure_time, is_open)
                        VALUES ('email', :failure_count, :last_failure, :is_open)
                        ON CONFLICT (service) DO UPDATE SET
                            failure_count = :failure_count,
                            last_failure_time = :last_failure,
                            is_open = :is_open
                    """),
                    {
                        'failure_count': self.failure_count,
                        'last_failure': self.last_failure_time,
                        'is_open': self.is_open
                    }
                )
                db.session.commit()
        except:
            pass

    def call_succeeded(self):
        """Reset the circuit breaker on success"""
        self.failure_count = 0
        self.is_open = False
        self.last_failure_time = None
        self._save_state()

    def call_failed(self):
        """Record a failure and potentially open the circuit"""
        self.failure_count += 1
        self.last_failure_time = datetime.utcnow()

        if self.failure_count >= self.failure_threshold:
            self.is_open = True
            app.logger.error(f"Email circuit breaker opened after {self.failure_count} failures")

        self._save_state()

    def can_attempt_call(self):
        """Check if we can attempt to send email"""
        if not self.is_open:
            return True

        # Check if we should attempt recovery
        if self.last_failure_time:
            time_since_failure = (datetime.utcnow() - self.last_failure_time).total_seconds()
            if time_since_failure >= self.recovery_timeout:
                # Attempt recovery
                self.is_open = False
                self.failure_count = 0
                self._save_state()
                app.logger.info("Email circuit breaker attempting recovery")
                return True

        return False


# Create global circuit breaker instance
email_circuit_breaker = EmailCircuitBreaker()


class EmailBounceHandler:
    """Handle email bounces and invalid addresses"""

    def __init__(self):
        self.bounce_threshold = 3
        self.bounced_emails = {}  # In production, use Redis or database

    def record_bounce(self, email):
        """Record an email bounce"""
        if email not in self.bounced_emails:
            self.bounced_emails[email] = 0
        self.bounced_emails[email] += 1

        # If exceeded threshold, mark email as invalid
        if self.bounced_emails[email] >= self.bounce_threshold:
            try:
                # Mark user email as invalid in database
                user = User.query.filter_by(email=email).first()
                if user:
                    # Add a field to track email validity
                    db.session.execute(
                        text("UPDATE users SET email_valid = FALSE WHERE id = :user_id"),
                        {'user_id': user.id}
                    )
                    db.session.commit()
            except Exception as e:
                logger.error(f"Error marking email as invalid: {e}")

    def is_valid_email(self, email):
        """Check if email is valid for sending"""
        if email in self.bounced_emails and self.bounced_emails[email] >= self.bounce_threshold:
            return False

        # Check database
        try:
            result = db.session.execute(
                text("SELECT email_valid FROM users WHERE email = :email"),
                {'email': email}
            ).first()
            if result and result[0] is False:
                return False
        except:
            pass

        return True


# Create global instance
email_bounce_handler = EmailBounceHandler()


def send_email_async(app, to_email, subject, body, html_body=None):
    """Send email asynchronously in app context with circuit breaker"""
    with app.app_context():
        # Check if email is valid
        if not email_bounce_handler.is_valid_email(to_email):
            app.logger.warning(f"Skipping email to {to_email} - marked as invalid")
            return

        # Check circuit breaker
        if not email_circuit_breaker.can_attempt_call():
            app.logger.warning(f"Email circuit breaker is open, not sending email to {to_email}")
            return

        try:
            msg = MIMEMultipart('alternative')
            msg['From'] = app.config['MAIL_USERNAME']
            msg['To'] = to_email
            msg['Subject'] = subject

            # Add unsubscribe link to body if we have user context
            unsubscribe_url = None
            try:
                if hasattr(request, 'current_user') and request.current_user:
                    unsubscribe_token = generate_unsubscribe_token(request.current_user.id)
                    base_url = os.environ.get('APP_BASE_URL', 'https://therapy-companion.onrender.com')
                    unsubscribe_url = f"{base_url}/api/unsubscribe/{unsubscribe_token}"
                    body += f"\n\n---\nTo unsubscribe from these emails, visit: {unsubscribe_url}"

                    if html_body:
                        html_body += f"""
                                           <hr style="margin-top: 40px; border: none; border-top: 1px solid #ccc;">
                                           <p style="text-align: center; color: #999; font-size: 12px;">
                                               <a href="{unsubscribe_url}" style="color: #999;">Unsubscribe from these emails</a>
                                           </p>
                                           """
            except RuntimeError:
                # We're not in a request context (e.g., running from Celery)
                # Skip unsubscribe link for system-generated emails
                pass

            # Add plain text part
            msg.attach(MIMEText(body, 'plain'))

            # Add HTML part if provided
            if html_body:
                msg.attach(MIMEText(html_body, 'html'))

            # Send email
            server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])

            # Enable TLS encryption
            try:
                server.starttls()
                server.ehlo()  # Re-identify ourselves over TLS connection
                # Log encryption status for compliance
                app.logger.info(f"Sending encrypted email to {to_email} via TLS")
            except Exception as e:
                app.logger.error(f"Failed to enable TLS encryption: {e}")
                email_circuit_breaker.call_failed()
                return False # Re-identify ourselves over TLS connection

            # Login and send
            server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])

            # Log encryption status for compliance
            app.logger.info(f"Sending encrypted email to {to_email} via TLS")

            server.send_message(msg)
            server.quit()

            app.logger.info(f"Email sent successfully to {to_email}")
            email_circuit_breaker.call_succeeded()
            return True  # ADD THIS LINE - Return True on success

        except Exception as e:
            app.logger.error(f"Error sending email: {e}")
            email_circuit_breaker.call_failed()
            return False  # ADD THIS LINE - Return False on failure


def process_email_queue_batch():
    """Process email queue in batches for better performance"""
    from datetime import datetime, timedelta

    with app.app_context():
        # Get batch of pending emails
        batch_size = 50  # Process 50 at a time

        pending_emails = EmailQueue.query.filter(
            EmailQueue.status == 'pending',
            EmailQueue.attempts < 3
        ).order_by(EmailQueue.created_at).limit(batch_size).all()

        for email_record in pending_emails:
            try:
                # Mark as processing
                email_record.status = 'processing'
                email_record.last_attempt_at = datetime.utcnow()
                email_record.attempts += 1
                db.session.commit()

                # Send email
                send_email_async(
                    app,
                    email_record.to_email,
                    email_record.subject,
                    email_record.body,
                    email_record.html_body
                )

                # Mark as sent
                email_record.status = 'sent'
                email_record.sent_at = datetime.utcnow()
                db.session.commit()

            except Exception as e:
                # Mark as failed
                email_record.status = 'failed'
                email_record.error_message = str(e)
                db.session.commit()

                logger.error(f"Failed to send email {email_record.id}: {e}")


def send_email(to_email, subject, body, html_body=None):
    """Send email using Celery if available, otherwise use thread"""
    if not app.config.get('MAIL_USERNAME'):
        # Email not configured, return silently
        return False

    # Check if Celery is available and configured
    try:
        from celery_app import send_email_task

        # Create EmailQueue entry for Celery
        email_queue = EmailQueue(
            to_email=to_email,
            subject=subject,
            body=body,
            html_body=html_body,
            status='pending'
        )
        db.session.add(email_queue)
        db.session.commit()

        # Use Celery to send email asynchronously - pass only the ID
        task = send_email_task.delay(email_queue.id)
        logger.info(f"Email queued via Celery with task id: {task.id} for queue id: {email_queue.id}")
        return True

    except ImportError:
        # Celery not available, fall back to thread-based sending
        logger.info("Celery not available, using thread-based email sending")

        # Check circuit breaker before creating thread
        if not email_circuit_breaker.can_attempt_call():
            app.logger.warning(f"Email circuit breaker is open, skipping email to {to_email}")
            return False

        # Send email in background thread
        thread = Thread(
            target=send_email_async,
            args=(app, to_email, subject, body, html_body)
        )
        thread.daemon = True
        thread.start()

        return True

    except Exception as e:
        logger.error(f"Error queueing email with Celery: {e}")
        db.session.rollback()  # Rollback the failed EmailQueue entry

        # Check circuit breaker before falling back
        if not email_circuit_breaker.can_attempt_call():
            app.logger.warning(f"Email circuit breaker is open, skipping email to {to_email}")
            return False

        # Fall back to thread-based sending
        thread = Thread(
            target=send_email_async,
            args=(app, to_email, subject, body, html_body)
        )
        thread.daemon = True
        thread.start()
        return True


def check_client_inactivity():
    """Check for inactive clients and notify therapists"""
    with app.app_context():
        try:
            # Get all active clients who haven't checked in for 7 days
            seven_days_ago = date.today() - timedelta(days=7)

            inactive_clients = db.session.query(Client).join(User).filter(
                Client.is_active == True,
                User.is_active == True
            ).all()

            notifications_sent = 0

            for client in inactive_clients:
                # Get last checkin
                last_checkin = client.checkins.order_by(
                    DailyCheckin.checkin_date.desc()
                ).first()

                if not last_checkin or last_checkin.checkin_date < seven_days_ago:
                    # Notify therapist
                    if client.therapist and client.therapist.user:
                        therapist_email = client.therapist.user.email

                        # Get therapist's language preference
                        # You might want to store this in the Therapist model
                        # For now, default to English
                        lang = 'en'  # TODO: Get from therapist preferences

                        # Translated subjects
                        subjects = {
                            'en': f"Client Inactivity Alert - {sanitize_input(client.client_serial)}",
                            'he': f"התראת חוסר פעילות - מטופל {sanitize_input(client.client_serial)}",
                            'ru': f"Предупреждение о неактивности клиента - {sanitize_input(client.client_serial)}",
                            'ar': f"تنبيه عدم نشاط العميل - {sanitize_input(client.client_serial)}"
                        }

                        # Translated bodies
                        if lang == 'he':
                            body = f"""שלום {client.therapist.name},

המטופל שלך {sanitize_input(client.client_serial)} לא השלים צ'ק-אין במשך יותר מ-7 ימים.

צ'ק-אין אחרון: {last_checkin.checkin_date.strftime('%d/%m/%Y') if last_checkin else 'אף פעם'}
תאריך התחלת טיפול: {client.start_date.strftime('%d/%m/%Y')}

אנא שקול ליצור קשר כדי לבדוק את ההתקדמות שלהם.

בברכה,
מערכת הליווי הטיפולי"""
                        elif lang == 'ru':
                            body = f"""Уважаемый {client.therapist.name},

Ваш клиент {sanitize_input(client.client_serial)} не выполнял отметку более 7 дней.

Последняя отметка: {last_checkin.checkin_date.strftime('%d.%m.%Y') if last_checkin else 'Никогда'}
Дата начала терапии: {client.start_date.strftime('%d.%m.%Y')}

Пожалуйста, подумайте о том, чтобы связаться с ним для проверки прогресса.

С уважением,
Система терапевтического сопровождения"""
                        elif lang == 'ar':
                            body = f"""عزيزي {client.therapist.name}،

لم يكمل عميلك {sanitize_input(client.client_serial)} تسجيل الحضور لأكثر من 7 أيام.

آخر تسجيل حضور: {last_checkin.checkin_date.strftime('%d/%m/%Y') if last_checkin else 'أبداً'}
تاريخ بدء العلاج: {client.start_date.strftime('%d/%m/%Y')}

يرجى التفكير في التواصل معهم للتحقق من تقدمهم.

مع أطيب التحيات،
نظام المرافقة العلاجية"""
                        else:  # English
                            body = f"""Dear {client.therapist.name},

Your client {sanitize_input(client.client_serial)} has not completed a check-in for over 7 days.

Last check-in: {last_checkin.checkin_date.strftime('%Y-%m-%d') if last_checkin else 'Never'}
Client start date: {client.start_date.strftime('%Y-%m-%d')}

Please consider reaching out to check on their progress.

Best regards,
Therapeutic Companion System"""

                        subject = subjects.get(lang, subjects['en'])

                        if send_email(therapist_email, subject, body):
                            notifications_sent += 1

                            # Log notification
                            logger.info('inactivity_notification_sent', extra={
                                'extra_data': {
                                    'client_id': client.id,
                                    'client_serial': client.client_serial,
                                    'therapist_id': client.therapist.id,
                                    'last_checkin': last_checkin.checkin_date.isoformat() if last_checkin else None
                                }
                            })

            return notifications_sent

        except Exception as e:
            logger.error(f"Error checking client inactivity: {e}")
            return 0


# Add scheduled task endpoint
@app.route('/api/admin/check-inactivity', methods=['POST'])
@require_auth(['therapist'])  # Could be restricted to admin
def trigger_inactivity_check():
    """Manually trigger inactivity check"""
    try:
        count = check_client_inactivity()
        return jsonify({
            'success': True,
            'notifications_sent': count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# Add this endpoint to check session status
@app.route('/api/session/check')
@require_auth(['client', 'therapist'])  # Allow both client and therapist
def check_session():
    """Check if session is still valid"""
    try:
        # Get user info
        user = request.current_user

        # Log the check
        logger.info('session_check', extra={
            'extra_data': {
                'user_id': user.id,
                'user_type': user.role,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': user.id
        })

        return jsonify({
            'valid': True,
            'expires_in': 3600,  # You could calculate actual time remaining
            'user_id': user.id,
            'role': user.role
        })

    except Exception as e:
        logger.error('session_check_error', extra={
            'extra_data': {
                'error': str(e),
                'request_id': g.request_id if hasattr(g, 'request_id') else None
            }
        })
        return jsonify({'valid': False}), 401




def get_language_from_header():
    """Get preferred language from Accept-Language header"""
    # First check cookie
    lang_cookie = request.cookies.get('preferred_language')
    if lang_cookie and lang_cookie in ['en', 'he', 'ru', 'ar']:
        return lang_cookie

    # Then check header
    accept_language = request.headers.get('Accept-Language', 'en')
    print(f"Accept-Language header: {accept_language}")
    # Simple parsing - just get the first language code
    lang = accept_language.split(',')[0].split('-')[0].lower()

    # Map to supported languages
    if lang in ['en', 'he', 'ru', 'ar']:
        return lang
    return 'en'


def translate_category_name(name, lang='en'):
    """Translate category name to specified language"""
    result = CATEGORY_TRANSLATIONS.get(lang, {}).get(name, name)
    print(f"Translating '{name}' to '{lang}': '{result}'")
    return result


def translate_report_term(term, lang='en'):
    """Translate report term to specified language"""
    return REPORT_TRANSLATIONS.get(lang, {}).get(term, term)


def get_translated_month(date_obj, lang='en'):
    """Get translated month name"""
    months = REPORT_TRANSLATIONS.get(lang, {}).get('months', None)
    if months:
        return months[date_obj.month - 1]
    return date_obj.strftime('%B')



def translate_day_name(day_index, lang='en'):
    """Translate day name to specified language"""
    days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])
    return days[day_index] if 0 <= day_index < 7 else ''


_categories_cache = None
_categories_cache_time = None


def ensure_default_categories():
    """Ensure all default tracking categories exist in the database"""
    global _categories_cache, _categories_cache_time

    # Return cached categories if fresh (less than 5 minutes old)
    if _categories_cache and _categories_cache_time:
        if (datetime.utcnow() - _categories_cache_time).seconds < 300:
            return _categories_cache

    default_categories = [
        ('Emotion Level', 'Overall emotional state', True),
        ('Energy', 'Physical and mental energy levels', True),
        ('Social Activity', 'Engagement in social interactions', True),
        ('Sleep Quality', 'Quality of sleep', False),
        ('Anxiety Level', 'Level of anxiety experienced', False),
        ('Motivation', 'Level of motivation and drive', False),
        ('Medication', 'Medication adherence', True),
        ('Physical Activity', 'Physical activity level', True)
    ]

    categories_added = 0
    for name, description, is_default in default_categories:
        # Check if category exists
        category = TrackingCategory.query.filter_by(name=name).first()
        if not category:
            category = TrackingCategory(
                name=name,
                description=description,
                is_default=is_default
            )
            db.session.add(category)
            categories_added += 1
            print(f"Added missing category: {name}")

    if categories_added > 0:
        db.session.commit()
        print(f"Added {categories_added} missing categories")

        # Cache the results
    _categories_cache = TrackingCategory.query.all()
    _categories_cache_time = datetime.utcnow()
    return _categories_cache


def fix_existing_clients():
    """Automatically fix all existing clients to have all categories"""
    try:
        all_categories = ensure_default_categories()
        clients = Client.query.all()
        fixed_count = 0

        for client in clients:
            # Get existing category IDs for this client
            existing_category_ids = set()
            for plan in client.tracking_plans.all():
                existing_category_ids.add(plan.category_id)

            # Add missing categories
            for category in all_categories:
                if category.id not in existing_category_ids:
                    plan = ClientTrackingPlan(
                        client_id=client.id,
                        category_id=category.id,
                        is_active=True
                    )
                    db.session.add(plan)
                    fixed_count += 1
                    print(f"Added {category.name} to client {sanitize_input(client.client_serial)}")

        if fixed_count > 0:
            db.session.commit()
            print(f"Fixed {fixed_count} missing category assignments")

    except Exception as e:
        print(f"Error fixing client categories: {e}")
        db.session.rollback()


# ============= HTML PAGE ROUTES =============

@app.route('/')
def index():
    """Serve the main HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'index.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            app.logger.error(f"index.html not found at {file_path}")
            return "index.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving index.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/login.html')
def login_page():
    """Serve the login HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'login.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            app.logger.error(f"login.html not found at {file_path}")
            return "login.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving login.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/therapist-dashboard.html')
def therapist_dashboard_page():
    """Serve the therapist dashboard HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'therapist_dashboard.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            # Try alternative filename
            alt_path = os.path.join(BASE_DIR, 'therapist-dashboard.html')
            if os.path.exists(alt_path):
                return send_file(alt_path)
            app.logger.error(f"therapist_dashboard.html not found at {file_path}")
            return "therapist_dashboard.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving therapist_dashboard.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/client-dashboard.html')
def client_dashboard_page():
    """Serve the client dashboard HTML file"""
    try:
        file_path = os.path.join(BASE_DIR, 'client_dashboard.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            # Try alternative filename
            alt_path = os.path.join(BASE_DIR, 'client-dashboard.html')
            if os.path.exists(alt_path):
                return send_file(alt_path)
            app.logger.error(f"client_dashboard.html not found at {file_path}")
            return "client_dashboard.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving client_dashboard.html: {e}")
        return f"Error: {str(e)}", 500


@app.route('/i18n.js')
def serve_i18n():
    """Serve the i18n.js file"""
    try:
        file_path = os.path.join(BASE_DIR, 'i18n.js')
        if os.path.exists(file_path):
            return send_file(file_path, mimetype='application/javascript')
        else:
            app.logger.error(f"i18n.js not found at {file_path}")
            return "i18n.js not found", 404
    except Exception as e:
        app.logger.error(f"Error serving i18n.js: {e}")
        return f"Error: {str(e)}", 500


# ============= API ENDPOINTS =============

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register new user (therapist or client)"""
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')
        role = data.get('role')

        # Log registration attempt
        logger.info('registration_attempt', extra={
            'extra_data': {
                'email': email,
                'role': role,
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown')
        })

        # Validate input
        if not all([email, password, role]):
            logger.warning('registration_failed', extra={
                'extra_data': {
                    'reason': 'missing_required_fields',
                    'email': email,
                    'role': role,
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown')
            })
            return jsonify({'error': 'Missing required fields'}), 400

        if role not in ['therapist', 'client']:
            logger.warning('registration_failed', extra={
                'extra_data': {
                    'reason': 'invalid_role',
                    'role': role,
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown')
            })
            return jsonify({'error': 'Invalid role'}), 400

        # Check if email exists
        if User.query.filter_by(email=email).first():
            logger.warning('registration_failed', extra={
                'extra_data': {
                    'reason': 'email_already_exists',
                    'email': email,
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown')
            })
            return jsonify({'error': 'Email already registered'}), 400

        if len(password) < 12:  # Increased from 8
            return jsonify({'error': 'Password must be at least 12 characters long'}), 400

            # Check password complexity
        import re
        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]', password):
            return jsonify({'error': 'Password must contain uppercase, lowercase, number and special character'}), 400

        # Check against common passwords
        common_passwords = ['Password123!', 'Therapy123!', 'Welcome123!']  # Add more
        if password in common_passwords:
            return jsonify({'error': 'Password is too common. Please choose a stronger password'}), 400

        # Create user
        user = User(
            email=email,
            password_hash=bcrypt.generate_password_hash(password).decode('utf-8'),
            role=role
        )
        db.session.add(user)
        db.session.flush()

        # Create role-specific profile
        if role == 'therapist':
            # Check if license number already exists
            if Therapist.query.filter_by(license_number=data.get('license_number', '')).first():
                db.session.rollback()
                logger.warning('registration_failed', extra={
                    'extra_data': {
                        'reason': 'license_already_exists',
                        'license_number': data.get('license_number', ''),
                        'request_id': getattr(g, 'request_id', 'unknown')
                    },
                    'request_id': getattr(g, 'request_id', 'unknown')
                })
                return jsonify({'error': 'License number already registered'}), 400

            therapist = Therapist(
                user_id=user.id,
                license_number=data.get('license_number', ''),
                name=data.get('name', ''),
                organization=data.get('organization', ''),
                specializations=data.get('specializations', [])
            )
            db.session.add(therapist)
        else:  # client
            client = Client(
                user_id=user.id,
                client_serial=generate_client_serial(),
                client_name=email,  # Use the email they registered with as their name
                therapist_id=data.get('therapist_id'),
                start_date=date.today()
            )
            db.session.add(client)

            # Add default tracking categories
            default_categories = TrackingCategory.query.filter_by(is_default=True).all()
            for category in default_categories:
                plan = ClientTrackingPlan(
                    client_id=client.id,
                    category_id=category.id
                )
                db.session.add(plan)

        db.session.commit()

        # Generate token
        token = generate_token(user.id, role)

        # Log successful registration
        logger.info('registration_success', extra={
            'extra_data': {
                'user_id': user.id,
                'email': email,
                'role': role,
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown'),
            'user_id': user.id
        })

        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'id': user.id,
                'email': user.email,
                'role': user.role
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error('registration_error', extra={
            'extra_data': {
                'error': str(e),
                'email': email,
                'role': role,
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown')
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("20 per minute")
def login():
    """Login user with secure session management"""
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')

        if not all([email, password]):
            logger.warning('login_failed', extra={
                'extra_data': {
                    'reason': 'missing_credentials',
                    'email': email,
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })
            return jsonify({'error': 'Missing email or password'}), 400

        # Check for account lockout
        lockout_key = f"lockout:{email}"
        attempts_key = f"attempts:{email}"

        if redis_client:  # Add Redis client to your app
            if redis_client.get(lockout_key):
                return jsonify({'error': 'Account locked. Please try again later.'}), 429

            attempts = int(redis_client.get(attempts_key) or 0)
            if attempts >= 5:
                redis_client.setex(lockout_key, 432000, '1')  # 5 day lockout
                return jsonify({'error': 'Too many failed attempts. Account locked for 15 minutes.'}), 429

        # Find user
        user = User.query.filter_by(email=email).first()
        if not user or not bcrypt.check_password_hash(user.password_hash, password):
            # Increment failed attempts
            if redis_client:
                redis_client.incr(attempts_key)
                redis_client.expire(attempts_key, 900)

            logger.warning('login_failed', extra={
                'extra_data': {
                    'reason': 'invalid_credentials',
                    'email': email,
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })
            return jsonify({'error': 'Invalid credentials'}), 401

        if not user.is_active:
            logger.warning('login_failed', extra={
                'extra_data': {
                    'reason': 'account_deactivated',
                    'email': email,
                    'user_id': user.id,
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })
            return jsonify({'error': 'Account deactivated'}), 401

        # Check password strength for existing users
        def meets_password_requirements(password):
            """Check if password meets current security requirements"""
            import re
            if len(password) < 12:
                return False
            if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]', password):
                return False
            return True

        if not meets_password_requirements(password):
            # Create a one-time token for password reset
            reset_token = str(uuid.uuid4())

            # Store reset token in database
            password_reset = PasswordReset(
                user_id=user.id,
                reset_token=reset_token,
                expires_at=datetime.utcnow() + timedelta(hours=1)
            )
            db.session.add(password_reset)
            db.session.commit()

            # Log security event
            logger.warning('weak_password_detected', extra={
                'extra_data': {
                    'user_id': user.id,
                    'email': email,
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })

            # Return special response
            return jsonify({
                'success': False,
                'requires_password_reset': True,
                'reset_token': reset_token,
                'message': 'Your password no longer meets security requirements. Please reset it.',
                'reset_url': f'/reset-password.html?token={reset_token}'
            }), 403

        # Clear failed attempts on successful login
        if redis_client:
            redis_client.delete(attempts_key)

        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Generate token but store it securely
        token = generate_token(user.id, user.role)

        # Create secure session
        session_id = str(uuid.uuid4())
        session_token = SessionToken(
            user_id=user.id,
            token=session_id,  # Store session ID, not JWT
            expires_at=datetime.utcnow() + timedelta(hours=JWT_EXPIRATION_HOURS)
        )
        db.session.add(session_token)
        db.session.commit()

        # Log successful login
        logger.info('login_success', extra={
            'extra_data': {
                'user_id': user.id,
                'role': user.role,
                'email': email,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': user.id
        })

        # Audit log
        log_audit('USER_LOGIN', 'user', user.id, {'role': user.role})

        # Get role-specific data
        response_data = {
            'success': True,
            'user': {
                'id': user.id,
                'email': user.email,
                'role': user.role
            }
        }

        if user.role == 'therapist' and user.therapist:
            response_data['therapist'] = {
                'id': user.therapist.id,
                'name': user.therapist.name,
                'license_number': user.therapist.license_number,
                'organization': user.therapist.organization
            }
        elif user.role == 'client' and user.client:
            response_data['client'] = {
                'id': user.client.id,
                'serial': user.client.client_serial,
                'start_date': user.client.start_date.isoformat()
            }

        # Set secure cookie instead of returning token
        resp = jsonify(response_data)
        resp.set_cookie(
             'session_token',
            session_id,
            secure=True,  # HTTPS only
            httponly=True,  # Not accessible via JavaScript
            samesite='Strict',  # Changed from 'Lax' to 'Strict'
            max_age=86400  # 24 hours
        )

        return resp

    except Exception as e:
        logger.error('login_error', extra={
            'extra_data': {
                'error': str(e),
                'request_id': g.request_id
            },
            'request_id': g.request_id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/request-reset', methods=['POST'])
@limiter.limit("10 per hour, 30 per day")  # More reasonable limits
def request_password_reset():
    """Request password reset"""
    try:
        data = request.json
        email = data.get('email')

        logger.info('password_reset_request', extra={
            'extra_data': {
                'email': email,
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown')
        })

        if not email:
            logger.warning('password_reset_failed', extra={
                'extra_data': {
                    'reason': 'missing_email',
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown')
            })
            return jsonify({'error': 'Email is required'}), 400

        # Find user
        user = User.query.filter_by(email=email).first()

        # Always return success to prevent email enumeration
        if not user:
            logger.info('password_reset_user_not_found', extra={
                'extra_data': {
                    'email': email,
                    'request_id': getattr(g, 'request_id', 'unknown')
                },
                'request_id': getattr(g, 'request_id', 'unknown')
            })
            return jsonify({
                'success': True,
                'message': 'If an account exists with this email, a password reset link has been sent.'
            })

        # Create reset token
        reset_token = str(uuid.uuid4())
        expires_at = datetime.utcnow() + timedelta(hours=1)

        # Save to database
        password_reset = PasswordReset(
            user_id=user.id,
            reset_token=reset_token,
            expires_at=expires_at
        )
        db.session.add(password_reset)
        db.session.commit()

        # Create reset link
        base_url = request.host_url.rstrip('/')
        reset_link = f"{base_url}/reset-password.html?token={reset_token}"

        # Email body
        subject = "Password Reset Request - Therapeutic Companion"
        body = f"""Hello,

You requested a password reset for your Therapeutic Companion account.

Click the link below to reset your password:
{reset_link}

This link will expire in 1 hour.

If you did not request this reset, please ignore this email.

Best regards,
Therapeutic Companion Team"""

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2>Password Reset Request</h2>
            <p>Hello,</p>
            <p>You requested a password reset for your Therapeutic Companion account.</p>
            <p>Click the button below to reset your password:</p>
            <p style="margin: 30px 0;">
                <a href="{reset_link}" style="background-color: #4CAF50; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px;">
                    Reset Password
                </a>
            </p>
            <p>Or copy this link: {reset_link}</p>
            <p><strong>This link will expire in 1 hour.</strong></p>
            <p>If you did not request this reset, please ignore this email.</p>
            <p>Best regards,<br>Therapeutic Companion Team</p>
        </body>
        </html>
        """

        # Send email
        email_sent = send_email(email, subject, body, html_body)

        logger.info('password_reset_success', extra={
            'extra_data': {
                'user_id': user.id,
                'email': email,
                'email_sent': email_sent,
                'token_expires': expires_at.isoformat(),
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown'),
            'user_id': user.id
        })

        return jsonify({
            'success': True,
            'message': 'If an account exists with this email, a password reset link has been sent.',
            'email_configured': email_sent
        })

    except Exception as e:
        db.session.rollback()
        logger.error('password_reset_error', extra={
            'extra_data': {
                'error': str(e),
                'email': email,
                'request_id': getattr(g, 'request_id', 'unknown')
            },
            'request_id': getattr(g, 'request_id', 'unknown')
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """Reset password with token"""
    try:
        data = request.json
        reset_token = data.get('token')
        new_password = data.get('password')

        if not all([reset_token, new_password]):
            return jsonify({'error': 'Token and password are required'}), 400

        if len(new_password) < 12:
            return jsonify({'error': 'Password must be at least 8 characters long'}), 400

        import re
        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]', new_password):
            return jsonify({'error': 'Password must contain uppercase, lowercase, number and special character'}), 400

        # Check against common passwords
        common_passwords = ['Password123!', 'Therapy123!', 'Welcome123!']  # Add more
        if new_password in common_passwords:
            return jsonify({'error': 'Password is too common. Please choose a stronger password'}), 400

        # Find valid reset token
        password_reset = PasswordReset.query.filter_by(
            reset_token=reset_token,
            used=False
        ).first()

        if not password_reset:
            return jsonify({'error': 'Invalid or expired reset token'}), 400

        # Check if expired
        if password_reset.expires_at < datetime.utcnow():
            return jsonify({'error': 'Reset token has expired'}), 400

        # Update password
        user = password_reset.user
        user.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')

        # Mark token as used
        password_reset.used = True

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Password has been reset successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============= THERAPIST ENDPOINTS =============

@app.route('/api/therapist/dashboard', methods=['GET'])
@require_auth(['therapist'])
def therapist_dashboard():
    """Get therapist dashboard data"""
    try:
        therapist = request.current_user.therapist

        # Get client statistics
        total_clients = therapist.clients.count()
        active_clients = therapist.clients.filter_by(is_active=True).count()

        # Get recent activity
        recent_checkins = db.session.query(DailyCheckin).join(Client).filter(
            Client.therapist_id == therapist.id,
            DailyCheckin.checkin_date >= date.today() - timedelta(days=7)
        ).count()

        # Get pending missions
        pending_missions = TherapistNote.query.filter_by(
            therapist_id=therapist.id,
            is_mission=True,
            mission_completed=False
        ).count()

        return jsonify({
            'success': True,
            'therapist': {
                'name': therapist.name,
                'license_number': therapist.license_number,
                'organization': therapist.organization
            },
            'statistics': {
                'total_clients': total_clients,
                'active_clients': active_clients,
                'recent_checkins': recent_checkins,
                'pending_missions': pending_missions
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/clients', methods=['GET'])
@require_auth(['therapist'])
def get_therapist_clients():
    """Get list of therapist's clients with pagination"""
    try:
        therapist = request.current_user.therapist
        lang = get_language_from_header()

        # Log the request
        logger.info('get_clients_request', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'language': lang,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Get filter and pagination parameters
        status = request.args.get('status', 'all')
        sort_by = request.args.get('sort_by', 'start_date')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        # Validate pagination parameters
        page = max(1, page)
        per_page = min(max(1, per_page), 50)  # Max 100 items per page

        # Build query
        query = therapist.clients

        if status == 'active':
            query = query.filter_by(is_active=True)
        elif status == 'inactive':
            query = query.filter_by(is_active=False)

        # Sort
        if sort_by == 'start_date':
            query = query.order_by(Client.start_date.desc())
        elif sort_by == 'serial':
            query = query.order_by(Client.client_serial)

        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        clients = pagination.items

        # Build response with optimized queries
        # Get all client IDs
        client_ids = [c.id for c in clients]

        # Batch load all tracking plans with categories in one query
        from sqlalchemy.orm import joinedload
        all_plans = ClientTrackingPlan.query.filter(
            ClientTrackingPlan.client_id.in_(client_ids),
            ClientTrackingPlan.is_active == True
        ).options(joinedload(ClientTrackingPlan.category)).all()

        # Group by client_id
        tracking_plans_data = {}
        for plan in all_plans:
            if plan.client_id not in tracking_plans_data:
                tracking_plans_data[plan.client_id] = []
            tracking_plans_data[plan.client_id].append(plan)

        # Get last check-in dates for all clients
        week_start = date.today() - timedelta(days=date.today().weekday())

        # Get last checkin dates and week counts with optimized queries
        from sqlalchemy import select, and_

        # Get all last checkins in one query
        last_checkin_subquery = db.session.query(
            DailyCheckin.client_id,
            func.max(DailyCheckin.checkin_date).label('last_date')
        ).filter(
            DailyCheckin.client_id.in_(client_ids)
        ).group_by(DailyCheckin.client_id).subquery()

        last_checkin_results = db.session.query(
            last_checkin_subquery.c.client_id,
            last_checkin_subquery.c.last_date
        ).all()

        last_checkins = {r[0]: r[1] for r in last_checkin_results}

        # Get all week counts in one query
        week_count_results = db.session.query(
            DailyCheckin.client_id,
            func.count(DailyCheckin.id).label('count')
        ).filter(
            DailyCheckin.client_id.in_(client_ids),
            DailyCheckin.checkin_date >= week_start
        ).group_by(DailyCheckin.client_id).all()

        week_counts = {r[0]: r[1] for r in week_count_results}
        # Set default 0 for clients with no checkins
        for client_id in client_ids:
            if client_id not in week_counts:
                week_counts[client_id] = 0

        # Build response
        client_data = []
        for client in clients:
            # Get tracking categories
            tracking_categories = []
            plans = tracking_plans_data.get(client.id, [])
            for plan in plans:
                if plan.is_active and plan.category:
                    translated_name = translate_category_name(plan.category.name, lang)
                    tracking_categories.append(translated_name)

            # Get stats
            last_checkin_date = last_checkins.get(client.id)
            week_checkin_count = week_counts.get(client.id, 0)

            client_data.append({
                'id': client.id,
                'serial': client.client_serial,
                'client_name': client.client_name if client.client_name else client.client_serial,
                'start_date': client.start_date.isoformat(),
                'is_active': client.is_active,
                'last_checkin': last_checkin_date.isoformat() if last_checkin_date else None,
                'week_completion': f"{week_checkin_count}/7",
                'tracking_categories': tracking_categories
            })

        # Log successful response
        logger.info('get_clients_success', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_count': len(client_data),
                'page': page,
                'total_pages': pagination.pages,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        return jsonify({
            'success': True,
            'clients': client_data,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_prev': pagination.has_prev,
                'has_next': pagination.has_next
            }
        })

    except Exception as e:
        logger.error('get_clients_error', extra={
            'extra_data': {
                'error': str(e),
                'therapist_id': therapist.id if therapist else None,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/search-clients', methods=['GET'])
@require_auth(['therapist'])
@cached_endpoint('client_search', ttl=60,
                 key_func=lambda: f"{request.current_user.therapist.id}:{request.args.get('q', '')}")
def search_therapist_clients():
    """Search clients with Elasticsearch"""
    try:
        therapist = request.current_user.therapist
        query = request.args.get('q', '').strip()

        if len(query) < 2:
            return jsonify({'error': 'Query must be at least 2 characters'}), 400

        # Search filters
        filters = {}
        if request.args.get('active_only') == 'true':
            filters['is_active'] = True

        # Use Elasticsearch if available
        if search_manager.es:
            results = search_manager.search_clients(therapist.id, query, filters)
            return jsonify({
                'success': True,
                'results': results,
                'source': 'elasticsearch'
            })

        # Fallback to database search
        clients_query = therapist.clients

        # Apply search
        search_filter = or_(
            Client.client_serial.ilike(f'%{query}%'),
            Client.client_name.ilike(f'%{query}%')
        )
        clients_query = clients_query.filter(search_filter)

        # Apply filters
        if request.args.get('active_only') == 'true':
            clients_query = clients_query.filter_by(is_active=True)

        # Get results
        clients = clients_query.limit(20).all()

        results = []
        for client in clients:
            results.append({
                'client_id': client.id,
                'client_serial': client.client_serial,
                'client_name': client.client_name or client.client_serial,
                'is_active': client.is_active,
                'start_date': client.start_date.isoformat()
            })

        return jsonify({
            'success': True,
            'results': results,
            'source': 'database'
        })

    except Exception as e:
        logger.error(f"Search error: {e}")
        return jsonify({'error': str(e)}), 500





@app.route('/api/therapist/client/<int:client_id>', methods=['GET'])
@require_auth(['therapist'])
def get_client_details(client_id):
    """Get detailed client information with strict access control"""
    try:
        therapist = request.current_user.therapist

        # Strict verification with explicit query
        client = db.session.query(Client).filter(
            Client.id == client_id,
            Client.therapist_id == therapist.id,
            Client.is_active == True
        ).first()

        if not client:
            # Log potential unauthorized access attempt
            logger.warning('unauthorized_client_access_attempt', extra={
                'extra_data': {
                    'therapist_id': therapist.id,
                    'requested_client_id': client_id,
                    'ip_address': request.remote_addr,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })

            # Audit log for security
            log_audit(
                action='UNAUTHORIZED_ACCESS_ATTEMPT',
                resource_type='client',
                resource_id=client_id,
                details={'therapist_id': therapist.id},
                phi_accessed=False
            )

            # Return generic error to prevent enumeration
            return jsonify({'error': 'Resource not found'}), 404

        # Log authorized access
        log_audit(
            action='VIEW_CLIENT_DETAILS',
            resource_type='client',
            resource_id=client_id,
            details={'client_serial': client.client_serial},
            phi_accessed=True
        )

        # Get tracking plans with additional verification
        tracking_plans = []
        for plan in client.tracking_plans.filter_by(is_active=True):
            # Verify plan belongs to client
            if plan.client_id != client.id:
                logger.error(f"Data integrity issue: plan {plan.id} has mismatched client_id")
                continue

            tracking_plans.append({
                'id': plan.id,
                'category': plan.category.name,
                'description': plan.category.description
            })

        # Get active goals with date validation
        active_goals = []
        week_start = date.today() - timedelta(days=date.today().weekday())

        goals_query = client.goals.filter_by(
            week_start=week_start,
            is_active=True
        ).filter(
            WeeklyGoal.therapist_id == therapist.id  # Extra verification
        )

        for goal in goals_query:
            # Get completions for each day
            completions = {}
            for i in range(7):
                day = week_start + timedelta(days=i)
                completion = goal.completions.filter_by(completion_date=day).first()
                completions[day.isoformat()] = completion.completed if completion else None

            active_goals.append({
                'id': goal.id,
                'text': goal.goal_text,
                'completions': completions
            })

        # Get recent check-ins with limit
        recent_checkins = []
        for checkin in client.checkins.order_by(
                DailyCheckin.checkin_date.desc()
        ).limit(7):  # Limit to prevent data exposure
            recent_checkins.append({
                'date': checkin.checkin_date.isoformat(),
                'emotional': checkin.emotional_value,
                'medication': checkin.medication_value,
                'activity': checkin.activity_value
                # Note: Not including notes in list view for privacy
            })

        # Get therapist notes/missions with verification
        notes = []
        notes_query = TherapistNote.query.filter_by(
            client_id=client_id,
            therapist_id=therapist.id  # Ensure notes are from this therapist
        ).order_by(TherapistNote.created_at.desc()).limit(10)

        for note in notes_query:
            notes.append({
                'id': note.id,
                'type': note.note_type,
                'content': note.content,  # Consider encrypting this
                'is_mission': note.is_mission,
                'completed': note.mission_completed,
                'created_at': note.created_at.isoformat()
            })

        # Get custom categories for this client
        custom_categories = CustomCategory.query.filter_by(
            client_id=client_id,
            is_active=True
        ).all()

        custom_categories_data = []
        for cat in custom_categories:
            custom_categories_data.append({
                'id': cat.id,
                'name': cat.name,
                'description': cat.description,
                'reverse_scoring': cat.reverse_scoring
            })

        return jsonify({
            'success': True,
            'client': {
                'id': client.id,
        'serial': client.client_serial,
        'client_name': client.client_name if client.client_name else client.client_serial,
        'start_date': client.start_date.isoformat(),
        'is_active': client.is_active,
        'tracking_plans': tracking_plans,
        'active_goals': active_goals,
        'recent_checkins': recent_checkins,
        'notes': notes,
        'custom_categories': custom_categories_data
            }
        })

    except Exception as e:
        logger.error('get_client_details_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': client_id,
                'therapist_id': therapist.id if therapist else None,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': 'An error occurred'}), 500


# ============= TRACKING CATEGORY MANAGEMENT =============


# NEW ENDPOINT - Added after /api/categories


# ============= REMINDER ENDPOINTS =============


# ============= PROFILE MANAGEMENT =============


# ============= ANALYTICS ENDPOINTS =============

@app.route('/api/therapist/analytics/<int:client_id>', methods=['GET'])
@require_auth(['therapist'])
def get_client_analytics(client_id):
    """Get detailed analytics for a client"""
    try:
        therapist = request.current_user.therapist
        lang = get_language_from_header()

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Get date range and pagination from query params
        days = int(request.args.get('days', 30))
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        # Validate pagination
        page = max(1, page)
        per_page = min(max(1, per_page), 100)
        days = min(max(1, days), 365)

        end_date = date.today()
        start_date = end_date - timedelta(days=days)

        # Get check-in completion stats
        total_days = (end_date - start_date).days + 1
        checkins_count = client.checkins.filter(
            DailyCheckin.checkin_date.between(start_date, end_date)
        ).count()
        completion_rate = (checkins_count / total_days) * 100

        # Get category trends
        # Get category trends with pagination
        category_trends = {}
        for plan in client.tracking_plans.filter_by(is_active=True):
            responses_query = CategoryResponse.query.filter(
                CategoryResponse.client_id == client_id,
                CategoryResponse.category_id == plan.category_id,
                CategoryResponse.response_date.between(start_date, end_date)
            ).order_by(CategoryResponse.response_date)

            # Paginate responses
            paginated = responses_query.paginate(page=page, per_page=per_page, error_out=False)
            responses = paginated.items

            if responses:
                translated_name = translate_category_name(plan.category.name, lang)
                category_trends[translated_name] = {
                    'data': [
                        {
                            'date': resp.response_date.isoformat(),
                            'value': resp.value
                        } for resp in responses
                    ],
                    'average': sum(r.value for r in responses) / len(responses),
                    'total_records': paginated.total,
                    'has_more': paginated.has_next
                }

        # Get goal completion stats
        goals = client.goals.filter(
            WeeklyGoal.week_start >= start_date
        ).all()

        goal_stats = {
            'total_goals': len(goals),
            'completion_data': []
        }

        for goal in goals:
            completions = goal.completions.all()
            completed_count = sum(1 for c in completions if c.completed)
            total_days = len(completions)
            goal_stats['completion_data'].append({
                'goal': goal.goal_text,
                'week_start': goal.week_start.isoformat(),
                'completion_rate': (completed_count / total_days * 100) if total_days > 0 else 0
            })

        # Get mission completion stats
        missions = TherapistNote.query.filter_by(
            client_id=client_id,
            therapist_id=therapist.id,
            is_mission=True
        ).all()
        completed_missions = sum(1 for m in missions if m.mission_completed)

        return jsonify({
            'success': True,
            'analytics': {
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat(),
                    'days': days
                },
                'completion': {
                    'rate': completion_rate,
                    'checkins': checkins_count,
                    'total_days': total_days
                },
                'category_trends': category_trends,
                'goals': goal_stats,
                'missions': {
                    'total': len(missions),
                    'completed': completed_missions,
                    'completion_rate': (completed_missions / len(missions) * 100) if missions else 0
                }
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/ai-insights/<int:client_id>', methods=['GET'])
@require_auth(['therapist'])
def get_client_ai_insights(client_id):
    """Get AI-powered insights for a specific client"""
    try:
        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id,
            is_active=True
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        insights = []

        # Analyze last 30 days for patterns
        thirty_days_ago = date.today() - timedelta(days=30)
        all_checkins = client.checkins.filter(
            DailyCheckin.checkin_date >= thirty_days_ago
        ).order_by(DailyCheckin.checkin_date).all()

        # Need minimum data
        if len(all_checkins) < 5:
            return jsonify({
                'success': True,
                'insights': [],
                'message': 'Not enough data for analysis'
            })

        # Split into periods for comparison
        midpoint = len(all_checkins) // 2
        first_half = all_checkins[:midpoint]
        second_half = all_checkins[midpoint:]

        # 1. EMOTIONAL DECLINE DETECTION
        emotional_first = [c.emotional_value for c in first_half if c.emotional_value]
        emotional_second = [c.emotional_value for c in second_half if c.emotional_value]

        if emotional_first and emotional_second:
            avg_first = sum(emotional_first) / len(emotional_first)
            avg_second = sum(emotional_second) / len(emotional_second)
            decline_percent = ((avg_first - avg_second) / avg_first * 100) if avg_first > 0 else 0

            if decline_percent > 20:  # 20% decline
                insights.append({
                    'id': f'{client.id}_emotional_{int(time.time())}',
                    'client_id': client.id,
                    'priority': 'critical' if decline_percent > 30 else 'warning',
                    'title': 'therapist.declining_emotional_state',
                    'description': 'therapist.emotional_decline_desc',
                    'description_params': {'percent': int(decline_percent)},
                    'recommendation': 'therapist.schedule_session_rec',
                    'trend': 'declining',
                    'trend_description': 'therapist.from_to_trend',
                    'trend_params': {'from': f'{avg_first:.1f}', 'to': f'{avg_second:.1f}'}
                })
        # 2. INCONSISTENT CHECK-INS
        days_between_checkins = []
        for i in range(1, len(all_checkins)):
            days_diff = (all_checkins[i].checkin_date - all_checkins[i - 1].checkin_date).days
            days_between_checkins.append(days_diff)

        if days_between_checkins:
            avg_gap = sum(days_between_checkins) / len(days_between_checkins)
            if avg_gap > 3:  # More than 3 days average between check-ins
                insights.append({
                    'id': f'{client.id}_engagement_{int(time.time())}',
                    'client_id': client.id,
                    'priority': 'warning',
                    'title': 'therapist.inconsistent_engagement',
                    'description': 'therapist.inconsistent_desc',
                    'description_params': {'days': int(avg_gap)},
                    'recommendation': 'therapist.discuss_barriers_rec',
                    'trend': 'stable',
                    'trend_description': 'therapist.checkins_in_days',
                    'trend_params': {'count': len(all_checkins), 'days': 30}  # Using 30 since we're analyzing 30 days
                })

        # 3. MEDICATION ADHERENCE
        med_values = [c.medication_value for c in all_checkins if c.medication_value is not None]
        if med_values and sum(1 for m in med_values if m < 3) > len(med_values) * 0.3:
            non_adherent_days = sum(1 for m in med_values if m < 3)
            insights.append({
                'id': f'{client.id}_medication_{int(time.time())}',
                'client_id': client.id,
                'priority': 'warning' if non_adherent_days < 10 else 'critical',
                'title': 'therapist.medication_concerns',
                'description': 'therapist.medication_desc',
                'description_params': {'days': non_adherent_days},
                'recommendation': 'therapist.discuss_medication_rec',
                'trend': 'declining',
                'trend_description': 'therapist.non_adherent_percent',
                'trend_params': {'percent': int((non_adherent_days / 30) * 100)}
            })

        # 4. POSITIVE PROGRESS
        if emotional_first and emotional_second:
            improvement = avg_second - avg_first
            improvement_percent=(avg_second - avg_first)/(avg_first *100)
            insights.append({
                'id': f'{client.id}_progress_{int(time.time())}',
                'client_id': client.id,
                'priority': 'info',
                'title': 'therapist.positive_progress',
                'description': 'therapist.improvement_desc',
                'description_params': {'percent': int(improvement_percent)},
                'recommendation': 'therapist.acknowledge_progress_rec',
                'trend': 'improving',
                'trend_description': 'therapist.averaging_score',
                'trend_params': {'score': f'{avg_second:.1f}'}
            })
        # 5. CRISIS DETECTION
        if len(emotional_second) >= 3:
            last_three = emotional_second[-3:]
            if all(e <= 2 for e in last_three) and len(emotional_first) > 0:
                avg_first = sum(emotional_first) / len(emotional_first)
                if avg_first > 3:
                    insights.append({
                        'id': f'{client.id}_crisis_{int(time.time())}',
                        'client_id': client.id,
                        'priority': 'critical',
                        'title': 'therapist.potential_crisis',
                        'description': 'therapist.crisis_desc',
                        'description_params': {},
                        'recommendation': 'therapist.immediate_outreach_rec',
                        'trend': 'declining',
                        'trend_description': 'therapist.urgent_attention',
                        'trend_params': {}
                    })
        # 6. SLEEP QUALITY ANALYSIS (if tracking)
        sleep_responses = []
        for checkin in all_checkins:
            sleep_cat = CategoryResponse.query.join(TrackingCategory).filter(
                CategoryResponse.client_id == client.id,
                CategoryResponse.response_date == checkin.checkin_date,
                TrackingCategory.name.ilike('%sleep%')
            ).first()
            if sleep_cat:
                sleep_responses.append(sleep_cat.value)

        if len(sleep_responses) >= 7:
            poor_sleep_days = sum(1 for s in sleep_responses[-7:] if s <= 2)
            if poor_sleep_days >= 4:  # Poor sleep 4+ of last 7 days
                insights.append({
                    'id': f'{client.id}_sleep_{int(time.time())}',
                    'client_id': client.id,
                    'priority': 'warning',
                    'title': 'therapist.sleep_concerns',
                    'description': 'therapist.sleep_desc',
                    'description_params': {'days': poor_sleep_days},
                    'recommendation': 'therapist.sleep_hygiene_rec',
                    'trend': 'stable',
                    'trend_description': 'therapist.persistent_sleep',
                    'trend_params': {}
                })
        # Sort by priority
        priority_order = {'critical': 0, 'warning': 1, 'info': 2}
        insights.sort(key=lambda x: priority_order.get(x['priority'], 3))

        return jsonify({
            'success': True,
            'insights': insights[:10]  # Limit to top 10
        })

    except Exception as e:
        logger.error(f"AI insights error for client {client_id}: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/dismiss-insight/<string:insight_id>', methods=['POST'])
@require_auth(['therapist'])
def dismiss_insight(insight_id):
    """Dismiss an AI insight"""
    try:
        # In a real implementation, you'd track dismissed insights in the database
        # For now, just return success
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500






# ============= EXPORT ENDPOINTS =============


# ============= ADMIN ENDPOINTS (if needed) =============

@app.route('/api/admin/stats', methods=['GET'])
@require_auth(['therapist'])  # Could restrict to admin role
def get_system_stats():
    """Get system-wide statistics"""
    try:
        # Only allow specific therapists or add admin role check
        stats = {
            'total_users': User.query.count(),
            'total_therapists': Therapist.query.count(),
            'total_clients': Client.query.count(),
            'active_clients': Client.query.filter_by(is_active=True).count(),
            'total_checkins': DailyCheckin.query.count(),
            'checkins_today': DailyCheckin.query.filter_by(checkin_date=date.today()).count()
        }

        return jsonify({
            'success': True,
            'stats': stats
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500





# ============= CLIENT REPORT ENDPOINTS =============

@app.route('/api/client/generate-report/<week>', methods=['GET'])
@require_auth(['client'])
def client_generate_report(week):
    """Generate client's own weekly report with translations"""
    try:
        client = request.current_user.client
        lang = get_language_from_header()

        # Parse week
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Use shared function to create workbook with language support
        wb = create_weekly_report_excel(client, None, week_start, week_end, week_num, year, lang)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        filename = f"my_therapy_report_{sanitize_input(client.client_serial)}_week_{week_num}_{year}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def create_weekly_report_excel(client, therapist, week_start, week_end, week_num, year, lang='en'):
    """Create Excel workbook for weekly report with language support"""
    from datetime import datetime, date

    if isinstance(week_start, datetime):
        week_start = week_start.date()
    if isinstance(week_end, datetime):
        week_end = week_end.date()
    # Create Excel workbook
    wb = openpyxl.Workbook()

    # Get translations
    trans = lambda key: translate_report_term(key, lang)
    days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])

    # 1. Daily Check-ins Sheet
    ws_checkins = wb.active
    ws_checkins.title = trans('daily_checkins')

    # Header styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cell styles
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws_checkins.merge_cells('A1:J1')
    title_cell = ws_checkins['A1']
    title_cell.value = f"{trans('weekly_report_title')} - {trans('client')} {sanitize_input(client.client_name if client.client_name else client.client_serial)}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = header_alignment

    # Week info
    ws_checkins.merge_cells('A2:J2')
    week_cell = ws_checkins['A2']
    start_month = get_translated_month(week_start, lang)
    end_month = get_translated_month(week_end, lang)
    week_cell.value = f"{trans('week')} {week_num}, {year} ({start_month} {week_start.day} - {end_month} {week_end.day}, {year})"
    week_cell.font = Font(size=14)
    week_cell.alignment = header_alignment

    # Headers - dynamically based on categories
    all_categories = TrackingCategory.query.all()
    custom_categories = CustomCategory.query.filter_by(
        client_id=client.id,
        is_active=True
    ).all()

    headers = [trans('date'), trans('day'), trans('checkin_time')]

    # Add default category headers
    for category in all_categories:
        cat_name = translate_category_name(category.name, lang)
        headers.append(f"{cat_name} (1-5)")
        headers.append(f"{cat_name} {trans('notes')}")

        # Add custom category headers
    for custom_cat in custom_categories:
            headers.append(f"{sanitize_input(custom_cat.name)} (1-5)")
            headers.append(f"{sanitize_input(custom_cat.name)} {trans('notes')}")





    headers.append(trans('completion'))

    for col, header in enumerate(headers, 1):
        cell = ws_checkins.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Get check-ins for the week
    checkins = client.checkins.filter(
        DailyCheckin.checkin_date.between(week_start, week_end)
    ).order_by(DailyCheckin.checkin_date).all()

    # Color fills for ratings
    excellent_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    good_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    poor_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    # Populate daily check-ins
    row = 5
    checkins_completed = 0
    category_values = {cat.id: [] for cat in all_categories}

    for i in range(7):
        current_date = week_start + timedelta(days=i)
        checkin = next((c for c in checkins if c.checkin_date == current_date), None)

        ws_checkins.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
        ws_checkins.cell(row=row, column=2).value = days[i]

        if checkin:
            checkins_completed += 1
            ws_checkins.cell(row=row, column=3).value = checkin.checkin_time.strftime('%H:%M')

            # Get category responses
            col_idx = 4
            for category in all_categories:
                response = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    category_id=category.id,
                    response_date=current_date
                ).first()

                if response:
                    # Value cell
                    value_cell = ws_checkins.cell(row=row, column=col_idx)
                    value_cell.value = response.value
                    category_values[category.id].append(response.value)

                    # Apply color coding
                    if 'anxiety' in category.name.lower():
                        if response.value <= 2:
                            value_cell.fill = excellent_fill  # Green for low anxiety
                        elif response.value == 3:
                            value_cell.fill = good_fill  # Yellow for medium
                        else:
                            value_cell.fill = poor_fill  # Red for high anxiety
                    else:
                        if response.value >= 4:
                            value_cell.fill = excellent_fill
                        elif response.value == 3:
                            value_cell.fill = good_fill
                        else:
                            value_cell.fill = poor_fill

                    # Notes cell
                    ws_checkins.cell(row=row, column=col_idx + 1).value = response.notes or ''

                col_idx += 2

            ws_checkins.cell(row=row, column=len(headers)).value = "✓"
            ws_checkins.cell(row=row, column=len(headers)).fill = excellent_fill
        else:
            ws_checkins.cell(row=row, column=3).value = trans('no_checkin')
            ws_checkins.cell(row=row, column=3).font = Font(italic=True, color="999999")
            ws_checkins.cell(row=row, column=len(headers)).value = "✗"
            ws_checkins.cell(row=row, column=len(headers)).fill = poor_fill

        # Apply borders to all cells
        for col in range(1, len(headers) + 1):
            ws_checkins.cell(row=row, column=col).border = cell_border

        row += 1

    # 2. Weekly Summary Sheet
    ws_summary = wb.create_sheet(trans('weekly_summary'))

    # Summary title
    ws_summary.merge_cells('A1:E1')
    summary_title = ws_summary['A1']
    summary_title.value = trans('weekly_summary')
    summary_title.font = Font(bold=True, size=16)
    summary_title.alignment = header_alignment

    # Summary headers
    summary_headers = ['Metric', 'Value', 'Percentage', 'Rating', trans('notes')]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Calculate statistics
    summary_data = []

    # Check-in completion
    completion_rate = (checkins_completed / 7) * 100
    summary_data.append({
        'metric': trans('checkin_completion'),
        'value': f"{checkins_completed}/7 {trans('days')}",
        'percentage': f"{completion_rate:.1f}%",
        'rating': trans('excellent') if completion_rate >= 80 else trans('good') if completion_rate >= 60 else trans(
            'needs_improvement'),
        'notes': ''
    })

    # Category averages
    for category in all_categories:
        if category_values[category.id]:
            avg_value = sum(category_values[category.id]) / len(category_values[category.id])
            cat_name = translate_category_name(category.name, lang)

            # Check if this is anxiety category
            if 'anxiety' in category.name.lower():
                # Reverse logic for anxiety - low values are good, high values are bad
                rating = trans('excellent') if avg_value <= 2 else trans('good') if avg_value <= 3 else trans(
                    'needs_support')
            else:
                # Normal logic for other categories - high values are good
                rating = trans('excellent') if avg_value >= 4 else trans('good') if avg_value >= 3 else trans(
                    'needs_support')

            summary_data.append({
                'metric': f"{trans('average_rating')} - {cat_name}",
                'value': f"{avg_value:.2f}/5",
                'percentage': f"{(avg_value / 5) * 100:.1f}%",
                'rating': rating,
                'notes': ''
            })

    # Write summary data
    row = 4
    for data in summary_data:
        ws_summary.cell(row=row, column=1).value = data['metric']
        ws_summary.cell(row=row, column=2).value = data['value']
        ws_summary.cell(row=row, column=3).value = data['percentage']

        rating_cell = ws_summary.cell(row=row, column=4)
        rating_cell.value = data['rating']

        # Check if this row is for anxiety
        if 'anxiety' in data['metric'].lower() or any(
                term in data['metric'] for term in ['חרדה', 'тревожности', 'القلق']):
            # For anxiety, reverse the color logic
            if trans('excellent') in data['rating']:
                rating_cell.fill = excellent_fill  # Low anxiety is excellent (green)
            elif trans('good') in data['rating']:
                rating_cell.fill = good_fill  # Medium anxiety is good (yellow)
            else:
                rating_cell.fill = poor_fill  # High anxiety is poor (red)
        else:
            # Normal color logic for other categories
            if trans('excellent') in data['rating']:
                rating_cell.fill = excellent_fill
            elif trans('good') in data['rating']:
                rating_cell.fill = good_fill
            else:
                rating_cell.fill = poor_fill

        ws_summary.cell(row=row, column=5).value = data['notes']

        # Apply borders
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).border = cell_border

        row += 1

    # 3. Weekly Goals Sheet
    ws_goals = wb.create_sheet(trans('weekly_goals'))

    # Goals title
    ws_goals.merge_cells('A1:I1')
    goals_title = ws_goals['A1']
    goals_title.value = f"{trans('weekly_goals')} & {trans('completion')}"
    goals_title.font = Font(bold=True, size=16)
    goals_title.alignment = header_alignment

    # Goals headers
    goal_headers = ['Goal'] + [day[:3] for day in days] + [trans('completion_rate')]
    for col, header in enumerate(goal_headers, 1):
        cell = ws_goals.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    # Get weekly goals
    weekly_goals = client.goals.filter_by(
        week_start=week_start,
        is_active=True
    ).all()

    row = 4
    for goal in weekly_goals:
        ws_goals.cell(row=row, column=1).value = goal.goal_text

        # Get completions for each day
        completions = goal.completions.filter(
            GoalCompletion.completion_date.between(week_start, week_end)
        ).all()

        completed_days = 0
        for day_idx in range(7):
            current_date = week_start + timedelta(days=day_idx)
            completion = next((c for c in completions if c.completion_date == current_date), None)

            cell = ws_goals.cell(row=row, column=day_idx + 2)
            if completion:
                if completion.completed:
                    cell.value = "✓"
                    cell.fill = excellent_fill
                    completed_days += 1
                else:
                    cell.value = "✗"
                    cell.fill = poor_fill
            else:
                cell.value = "-"
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            cell.alignment = Alignment(horizontal="center")
            cell.border = cell_border

        # Completion rate
        completion_rate = (completed_days / 7) * 100
        rate_cell = ws_goals.cell(row=row, column=9)
        rate_cell.value = f"{completed_days}/7 ({completion_rate:.0f}%)"
        if completion_rate >= 80:
            rate_cell.fill = excellent_fill
        elif completion_rate >= 50:
            rate_cell.fill = good_fill
        else:
            rate_cell.fill = poor_fill
        rate_cell.border = cell_border

        # Apply borders to goal text
        ws_goals.cell(row=row, column=1).border = cell_border

        row += 1

    # 4. Therapist Notes Sheet (only if therapist is provided)
    if therapist:
        ws_notes = wb.create_sheet(trans('therapist_notes'))

        # Notes title
        ws_notes.merge_cells('A1:D1')
        notes_title = ws_notes['A1']
        notes_title.value = f"{trans('therapist_notes')} & {trans('mission')}s"
        notes_title.font = Font(bold=True, size=16)
        notes_title.alignment = header_alignment

        # Notes headers
        note_headers = [trans('date'), trans('type'), trans('content'), trans('status')]
        for col, header in enumerate(note_headers, 1):
            cell = ws_notes.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Get therapist notes for the week
        week_start_datetime = datetime.combine(week_start, datetime.min.time())
        week_end_datetime = datetime.combine(week_end, datetime.max.time())

        notes = TherapistNote.query.filter(
            TherapistNote.client_id == client.id,
            TherapistNote.therapist_id == therapist.id,
            TherapistNote.created_at.between(week_start_datetime, week_end_datetime)
        ).order_by(TherapistNote.created_at).all()

        row = 4
        for note in notes:
            ws_notes.cell(row=row, column=1).value = note.created_at.strftime('%Y-%m-%d %H:%M')

            type_cell = ws_notes.cell(row=row, column=2)
            if note.is_mission:
                type_cell.value = trans('mission')
                type_cell.font = Font(bold=True, color="E91E63")
            else:
                type_cell.value = note.note_type.title()

            ws_notes.cell(row=row, column=3).value = note.content

            status_cell = ws_notes.cell(row=row, column=4)
            if note.is_mission:
                if note.mission_completed:
                    status_cell.value = trans('completed')
                    status_cell.fill = excellent_fill
                else:
                    status_cell.value = trans('pending')
                    status_cell.fill = good_fill
            else:
                status_cell.value = "-"

            # Apply borders
            for col in range(1, 5):
                ws_notes.cell(row=row, column=col).border = cell_border

            row += 1

    # Adjust column widths for all sheets
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass

            # Only set width if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    return wb


def create_weekly_report_excel_streaming(client, therapist, week_start, week_end, week_num, year, lang='en'):
    """Create Excel workbook for weekly report with streaming to reduce memory usage"""
    from io import BytesIO
    import tempfile
    import os

    # Use temporary file instead of memory for large workbooks
    with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx') as tmp:
        tmp_name = tmp.name

    try:
        # Create workbook normally (NOT write_only) to support formatting
        wb = openpyxl.Workbook()

        # Get translations
        trans = lambda key: translate_report_term(key, lang)
        days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])

        # 1. Daily Check-ins Sheet
        ws_checkins = wb.active
        ws_checkins.title = trans('daily_checkins')

        # Header styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Cell styles
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws_checkins.merge_cells('A1:J1')
        title_cell = ws_checkins['A1']
        title_cell.value = f"{trans('weekly_report_title')} - {trans('client')} {sanitize_input(client.client_name if client.client_name else client.client_serial)}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = header_alignment

        # Week info
        ws_checkins.merge_cells('A2:J2')
        week_cell = ws_checkins['A2']
        start_month = get_translated_month(week_start, lang)
        end_month = get_translated_month(week_end, lang)
        week_cell.value = f"{trans('week')} {week_num}, {year} ({start_month} {week_start.day} - {end_month} {week_end.day}, {year})"
        week_cell.font = Font(size=14)
        week_cell.alignment = header_alignment

        # Headers - dynamically based on categories
        all_categories = TrackingCategory.query.all()
        custom_categories = CustomCategory.query.filter_by(
            client_id=client.id,
            is_active=True
        ).all()

        headers = [trans('date'), trans('day'), trans('checkin_time')]

        # Add default category headers
        for category in all_categories:
            cat_name = translate_category_name(category.name, lang)
            headers.append(f"{cat_name} (1-5)")
            headers.append(f"{cat_name} {trans('notes')}")

        # Add custom category headers
        for custom_cat in custom_categories:
            headers.append(f"{custom_cat.name} (1-5)")
            headers.append(f"{custom_cat.name} {trans('notes')}")

        headers.append(trans('completion'))

        for col, header in enumerate(headers, 1):
            cell = ws_checkins.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Color fills for ratings
        excellent_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        good_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        poor_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        # Populate daily check-ins
        row = 5
        checkins_completed = 0
        category_values = {cat.id: [] for cat in all_categories}

        for i in range(7):
            current_date = week_start + timedelta(days=i)
            checkin = client.checkins.filter_by(checkin_date=current_date).first()

            ws_checkins.cell(row=row, column=1).value = current_date.strftime('%Y-%m-%d')
            ws_checkins.cell(row=row, column=2).value = days[i]

            if checkin:
                checkins_completed += 1
                ws_checkins.cell(row=row, column=3).value = checkin.checkin_time.strftime('%H:%M')

                # Get category responses
                col_idx = 4
                for category in all_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        category_id=category.id,
                        response_date=current_date
                    ).first()

                    if response:
                        # Value cell
                        value_cell = ws_checkins.cell(row=row, column=col_idx)
                        value_cell.value = response.value
                        category_values[category.id].append(response.value)

                        # Apply color coding
                        if 'anxiety' in category.name.lower():
                            if response.value <= 2:
                                value_cell.fill = excellent_fill
                            elif response.value == 3:
                                value_cell.fill = good_fill
                            else:
                                value_cell.fill = poor_fill
                        else:
                            if response.value >= 4:
                                value_cell.fill = excellent_fill
                            elif response.value == 3:
                                value_cell.fill = good_fill
                            else:
                                value_cell.fill = poor_fill

                        # Notes cell
                        ws_checkins.cell(row=row, column=col_idx + 1).value = response.notes or ''

                    col_idx += 2

                    # Handle custom category responses
                for custom_cat in custom_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        custom_category_id=custom_cat.id,  # Note: using custom_category_id field
                        response_date=current_date
                    ).first()

                    if response:
                        # Value cell
                        value_cell = ws_checkins.cell(row=row, column=col_idx)
                        value_cell.value = response.value

                        # Apply color coding based on reverse scoring
                        if custom_cat.reverse_scoring:
                            if response.value <= 2:
                                value_cell.fill = excellent_fill
                            elif response.value == 3:
                                value_cell.fill = good_fill
                            else:
                                value_cell.fill = poor_fill
                        else:
                            if response.value >= 4:
                                value_cell.fill = excellent_fill
                            elif response.value == 3:
                                value_cell.fill = good_fill
                            else:
                                value_cell.fill = poor_fill

                        # Notes cell
                        ws_checkins.cell(row=row, column=col_idx + 1).value = response.notes or ''

                    col_idx += 2




                ws_checkins.cell(row=row, column=len(headers)).value = "✓"
                ws_checkins.cell(row=row, column=len(headers)).fill = excellent_fill
            else:
                ws_checkins.cell(row=row, column=3).value = trans('no_checkin')
                ws_checkins.cell(row=row, column=3).font = Font(italic=True, color="999999")
                ws_checkins.cell(row=row, column=len(headers)).value = "✗"
                ws_checkins.cell(row=row, column=len(headers)).fill = poor_fill

            # Apply borders to all cells
            for col in range(1, len(headers) + 1):
                ws_checkins.cell(row=row, column=col).border = cell_border

            row += 1

        # 2. Weekly Summary Sheet
        ws_summary = wb.create_sheet(trans('weekly_summary'))

        # Summary title
        ws_summary.merge_cells('A1:E1')
        summary_title = ws_summary['A1']
        summary_title.value = trans('weekly_summary')
        summary_title.font = Font(bold=True, size=16)
        summary_title.alignment = header_alignment

        # Summary headers
        summary_headers = [
            trans('metric'),
            trans('value'),
            trans('percentage'),
            trans('rating'),
            trans('notes')
        ]





        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Calculate statistics
        summary_data = []

        # Check-in completion
        completion_rate = (checkins_completed / 7) * 100
        summary_data.append({
            'metric': trans('checkin_completion'),
            'value': f"{checkins_completed}/7 {trans('days')}",
            'percentage': f"{completion_rate:.1f}%",
            'rating': trans('excellent') if completion_rate >= 80 else trans('good') if completion_rate >= 60 else trans('needs_improvement'),
            'notes': ''
        })

        # Category averages
        for category in all_categories:
            if category_values[category.id]:
                avg_value = sum(category_values[category.id]) / len(category_values[category.id])
                cat_name = translate_category_name(category.name, lang)

                if 'anxiety' in category.name.lower():
                    rating = trans('excellent') if avg_value <= 2 else trans('good') if avg_value <= 3 else trans('needs_support')
                else:
                    rating = trans('excellent') if avg_value >= 4 else trans('good') if avg_value >= 3 else trans('needs_support')

                summary_data.append({
                    'metric': f"{trans('average_rating')} - {cat_name}",
                    'value': f"{avg_value:.2f}/5",
                    'percentage': f"{(avg_value / 5) * 100:.1f}%",
                    'rating': rating,
                    'notes': ''
                })

        # Write summary data
        row = 4
        for data in summary_data:
            ws_summary.cell(row=row, column=1).value = data['metric']
            ws_summary.cell(row=row, column=2).value = data['value']
            ws_summary.cell(row=row, column=3).value = data['percentage']

            rating_cell = ws_summary.cell(row=row, column=4)
            rating_cell.value = data['rating']

            # Apply color based on rating
            if 'anxiety' in data['metric'].lower() or any(term in data['metric'] for term in ['חרדה', 'тревожности', 'القلق']):
                if trans('excellent') in data['rating']:
                    rating_cell.fill = excellent_fill
                elif trans('good') in data['rating']:
                    rating_cell.fill = good_fill
                else:
                    rating_cell.fill = poor_fill
            else:
                if trans('excellent') in data['rating']:
                    rating_cell.fill = excellent_fill
                elif trans('good') in data['rating']:
                    rating_cell.fill = good_fill
                else:
                    rating_cell.fill = poor_fill

            ws_summary.cell(row=row, column=5).value = data['notes']

            # Apply borders
            for col in range(1, 6):
                ws_summary.cell(row=row, column=col).border = cell_border

            row += 1

        # 3. Weekly Goals Sheet
        ws_goals = wb.create_sheet(trans('weekly_goals'))

        # Goals title
        ws_goals.merge_cells('A1:I1')
        goals_title = ws_goals['A1']
        goals_title.value = f"{trans('weekly_goals')} & {trans('completion')}"
        goals_title.font = Font(bold=True, size=16)
        goals_title.alignment = header_alignment

        # Goals headers
        goal_headers = [trans('goal')] + [day[:3] for day in days] + [trans('completion_rate')]
        for col, header in enumerate(goal_headers, 1):
            cell = ws_goals.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Get weekly goals
        weekly_goals = client.goals.filter_by(
            week_start=week_start.date(),
            is_active=True
        ).all()

        row = 4
        for goal in weekly_goals:
            ws_goals.cell(row=row, column=1).value = goal.goal_text

            # Get completions for each day
            completions = goal.completions.filter(
                GoalCompletion.completion_date.between(week_start.date(), week_end.date())
            ).all()

            completed_days = 0
            for day_idx in range(7):
                current_date = week_start.date() + timedelta(days=day_idx)
                completion = next((c for c in completions if c.completion_date == current_date), None)

                cell = ws_goals.cell(row=row, column=day_idx + 2)
                if completion:
                    if completion.completed:
                        cell.value = "✓"
                        cell.fill = excellent_fill
                        completed_days += 1
                    else:
                        cell.value = "✗"
                        cell.fill = poor_fill
                else:
                    cell.value = "-"
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                cell.alignment = Alignment(horizontal="center")
                cell.border = cell_border

            # Completion rate
            completion_rate = (completed_days / 7) * 100
            rate_cell = ws_goals.cell(row=row, column=9)
            rate_cell.value = f"{completed_days}/7 ({completion_rate:.0f}%)"
            if completion_rate >= 80:
                rate_cell.fill = excellent_fill
            elif completion_rate >= 50:
                rate_cell.fill = good_fill
            else:
                rate_cell.fill = poor_fill
            rate_cell.border = cell_border

            # Apply borders to goal text
            ws_goals.cell(row=row, column=1).border = cell_border

            row += 1

        # 4. Therapist Notes Sheet (only if therapist is provided)
        if therapist:
            ws_notes = wb.create_sheet(trans('therapist_notes'))

            # Notes title
            ws_notes.merge_cells('A1:D1')
            notes_title = ws_notes['A1']
            notes_title.value = f"{trans('therapist_notes')} & {trans('mission')}s"
            notes_title.font = Font(bold=True, size=16)
            notes_title.alignment = header_alignment

            # Notes headers
            note_headers = [trans('date'), trans('type'), trans('content'), trans('status')]
            for col, header in enumerate(note_headers, 1):
                cell = ws_notes.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = cell_border

            # Get therapist notes for the week
            week_start_datetime = datetime.combine(week_start.date(), datetime.min.time())
            week_end_datetime = datetime.combine(week_end.date(), datetime.max.time())

            notes = TherapistNote.query.filter(
                TherapistNote.client_id == client.id,
                TherapistNote.therapist_id == therapist.id,
                TherapistNote.created_at.between(week_start_datetime, week_end_datetime)
            ).order_by(TherapistNote.created_at).all()

            row = 4
            for note in notes:
                ws_notes.cell(row=row, column=1).value = note.created_at.strftime('%Y-%m-%d %H:%M')

                type_cell = ws_notes.cell(row=row, column=2)
                if note.is_mission:
                    type_cell.value = trans('mission')
                    type_cell.font = Font(bold=True, color="E91E63")
                else:
                    type_cell.value = note.note_type.title()

                ws_notes.cell(row=row, column=3).value = note.content

                status_cell = ws_notes.cell(row=row, column=4)
                if note.is_mission:
                    if note.mission_completed:
                        status_cell.value = trans('completed')
                        status_cell.fill = excellent_fill
                    else:
                        status_cell.value = trans('pending')
                        status_cell.fill = good_fill
                else:
                    status_cell.value = "-"

                # Apply borders
                for col in range(1, 5):
                    ws_notes.cell(row=row, column=col).border = cell_border

                row += 1

        # Adjust column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = None

                for cell in column:
                    try:
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass

                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Save to temporary file
        wb.save(tmp_name)

        # Read back as BytesIO for returning
        with open(tmp_name, 'rb') as f:
            output = BytesIO(f.read())
        output.seek(0)

        return output

    finally:
        # Clean up temporary file
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)


def generate_report_html_content(client, therapist, week_start, week_end, week_num, year, lang, trans, days, is_rtl):
    """Generate HTML content for PDF report - shared between WeasyPrint and xhtml2pdf"""

    # Get all categories
    all_categories = list(TrackingCategory.query.all())
    custom_categories = CustomCategory.query.filter_by(
        client_id=client.id,
        is_active=True
    ).all()

    # Start HTML
    html_content = f"""
    <!DOCTYPE html>
    <html dir="{'rtl' if is_rtl else 'ltr'}">
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4 landscape;
                margin: 1cm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 10pt;
                direction: {'rtl' if is_rtl else 'ltr'};
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 4px;
                text-align: center;
            }}
            th {{
                background-color: #2C3E50;
                color: white;
            }}
            .good {{ background-color: #C8E6C9; }}
            .medium {{ background-color: #FFF9C4; }}
            .poor {{ background-color: #FFCDD2; }}
        </style>
    </head>
    <body>
        <h1>{trans('weekly_report_title')} - {trans('client')} {client.client_name}</h1>
        <p>{trans('week')} {week_num}, {year}</p>

        <h2>{trans('daily_checkins')}</h2>
        <table>
            <tr>
                <th>{trans('date')}</th>
                <th>{trans('day')}</th>
    """

    # Add category headers
    for cat in all_categories:
        cat_name = translate_category_name(cat.name, lang)
        html_content += f'<th>{cat_name}</th>'

    for custom_cat in custom_categories:
        html_content += f'<th>{custom_cat.name}</th>'

    html_content += '</tr>'

    # Add daily data
    checkins = client.checkins.filter(
        DailyCheckin.checkin_date.between(week_start, week_end)
    ).order_by(DailyCheckin.checkin_date).all()

    for i in range(7):
        current_date = week_start + timedelta(days=i)
        day_name = days[i]
        checkin = next((c for c in checkins if c.checkin_date == current_date), None)

        html_content += f"""
            <tr>
                <td>{current_date.strftime('%Y-%m-%d')}</td>
                <td>{day_name}</td>
        """

        if checkin:
            # Add category values
            for category in all_categories:
                response = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    category_id=category.id,
                    response_date=current_date
                ).first()

                if response:
                    value = response.value
                    css_class = 'good' if value >= 4 else 'medium' if value == 3 else 'poor'
                    html_content += f'<td class="{css_class}">{value}</td>'
                else:
                    html_content += '<td>-</td>'

            for custom_cat in custom_categories:
                response = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    custom_category_id=custom_cat.id,
                    response_date=current_date
                ).first()

                if response:
                    value = response.value
                    css_class = 'good' if value >= 4 else 'medium' if value == 3 else 'poor'
                    html_content += f'<td class="{css_class}">{value}</td>'
                else:
                    html_content += '<td>-</td>'
        else:
            col_count = len(all_categories) + len(custom_categories)
            html_content += f'<td colspan="{col_count}">{trans("no_checkin")}</td>'

        html_content += '</tr>'

    html_content += """
        </table>
    </body>
    </html>
    """

    return html_content


def test_pdf_libraries():
    """Test which PDF libraries are available"""
    results = {}

    # Test WeasyPrint
    try:
        from weasyprint import HTML
        results['weasyprint'] = "Available"
    except ImportError as e:
        results['weasyprint'] = f"Not available: {str(e)}"
    except Exception as e:
        results['weasyprint'] = f"Error: {str(e)}"

    # Test xhtml2pdf
    try:
        from xhtml2pdf import pisa
        results['xhtml2pdf'] = "Available"
    except ImportError as e:
        results['xhtml2pdf'] = f"Not available: {str(e)}"
    except Exception as e:
        results['xhtml2pdf'] = f"Error: {str(e)}"

    return results


def create_weekly_report_pdf(client, therapist, week_start, week_end, week_num, year, lang='en'):
    """Create PDF report with proper Unicode support via WeasyPrint"""
    # DEBUG: Check what types we're receiving
    if hasattr(week_start, 'date'):
        week_start = week_start.date()
    if hasattr(week_end, 'date'):
        week_end = week_end.date()
    print(f"DEBUG PDF Generation:")
    print(f"  week_start type: {type(week_start)}, value: {week_start}")
    print(f"  week_end type: {type(week_end)}, value: {week_end}")
    print(f"  client.id: {client.id}")
    # Log the attempt
    print(f"Creating PDF for client {client.id}, week {week_num}, year {year}, lang {lang}")

    try:
        from weasyprint import HTML
        from io import BytesIO

        # Get translations
        trans = lambda key: translate_report_term(key, lang)
        days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])

        # Determine text direction
        is_rtl = lang in ['he', 'ar']

        # Generate HTML content with proper fonts
        html_content = f"""
        <!DOCTYPE html>
        <html dir="{'rtl' if is_rtl else 'ltr'}">
        <head>
            <meta charset="UTF-8">
            <style>
                @import url('https://fonts.googleapis.com/css2?family=Noto+Sans:wght@400;700&family=Noto+Sans+Hebrew:wght@400;700&family=Noto+Sans+Arabic:wght@400;700&display=swap');

                @page {{
                    size: A4 landscape;
                    margin: 1cm;
                }}

                body {{
                    font-family: 'Noto Sans', 'Noto Sans Hebrew', 'Noto Sans Arabic', Arial, sans-serif;
                    font-size: 10pt;
                    direction: {'rtl' if is_rtl else 'ltr'};
                    text-align: {'right' if is_rtl else 'left'};
                }}

                h1 {{
                    text-align: center;
                    color: #2C3E50;
                    font-size: 18pt;
                    margin-bottom: 10px;
                }}

                .subtitle {{
                    text-align: center;
                    color: #34495E;
                    margin-bottom: 20px;
                }}

                h2 {{
                    color: #34495E;
                    font-size: 14pt;
                    margin-top: 20px;
                    margin-bottom: 10px;
                }}

                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                    font-size: 8pt;
                    direction: {'rtl' if is_rtl else 'ltr'};
                }}

                th, td {{
                    border: 1px solid #ddd;
                    padding: 4px;
                    text-align: center;
                }}

                th {{
                    background-color: #2C3E50;
                    color: white;
                    font-weight: bold;
                    padding: 6px 4px;
                }}

                td {{
                    height: 25px;
                }}

                .good {{
                    background-color: #C8E6C9;
                    font-weight: bold;
                }}

                .medium {{
                    background-color: #FFF9C4;
                }}

                .poor {{
                    background-color: #FFCDD2;
                }}

                .no-checkin {{
                    color: #999;
                    font-style: italic;
                }}

                .summary {{
                    margin-top: 30px;
                    padding: 15px;
                    background-color: #f8f9fa;
                    border-radius: 5px;
                }}

                .summary h2 {{
                    margin-top: 0;
                }}

                .summary-item {{
                    margin: 8px 0;
                    padding: 5px 0;
                }}

                .excellent-text {{ color: #2e7d32; font-weight: bold; }}
                .good-text {{ color: #f57c00; font-weight: bold; }}
                .poor-text {{ color: #c62828; font-weight: bold; }}
            </style>
        </head>
        <body>
            <h1>{trans('weekly_report_title')} - {trans('client')} {escape(client.client_name if client.client_name else client.client_serial)}</h1>
            <p class="subtitle">{trans('week')} {week_num}, {year} ({get_translated_month(week_start, lang)} {week_start.day} - {get_translated_month(week_end, lang)} {week_end.day}, {year})</p>

            <h2>{trans('daily_checkins')}</h2>
            <table>
                <tr>
                    <th style="width: 12%">{trans('date')}</th>
                    <th style="width: 12%">{trans('day')}</th>
        """

        # Get all categories
        all_categories = list(TrackingCategory.query.all())
        print(f"Found {len(all_categories)} tracking categories")

        # Add custom categories for this client
        custom_categories = CustomCategory.query.filter_by(
            client_id=client.id,
            is_active=True
        ).all()
        print(f"Found {len(custom_categories)} custom categories for client {client.id}")

        # Combine all categories
        all_category_data = []
        for cat in all_categories:
            all_category_data.append({
                'id': cat.id,
                'name': cat.name,
                'is_custom': False,
                'reverse_scoring': 'anxiety' in cat.name.lower()
            })

        for custom_cat in custom_categories:
            all_category_data.append({
                'id': f'custom_{custom_cat.id}',
                'name': custom_cat.name,
                'is_custom': True,
                'reverse_scoring': custom_cat.reverse_scoring
            })

        print(f"Total categories to display: {len(all_category_data)}")

        # Calculate column width for categories
        remaining_width = 76  # 100% - 24% (date + day columns)
        col_width = remaining_width // len(all_category_data) if all_category_data else remaining_width

        # Add category headers
        for cat_data in all_category_data:
            cat_name = translate_category_name(cat_data['name'], lang) if not cat_data['is_custom'] else cat_data[
                'name']
            # Shorten very long category names to fit
            if len(cat_name) > 15:
                cat_name = cat_name[:13] + '..'
            html_content += f'<th style="width: {col_width}%">{cat_name}</th>'

        html_content += "</tr>"

        # Get check-ins
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).order_by(DailyCheckin.checkin_date).all()
        print(f"Found {len(checkins)} checkins for the week")

        # Track statistics
        checkin_count = 0
        category_totals = {cat.id: [] for cat in all_categories}
        # FIXED: Also initialize totals for custom categories
        for custom_cat in custom_categories:
            category_totals[f'custom_{custom_cat.id}'] = []

        # Add daily data rows
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            day_name = days[i]
            # Shorten day names if needed
            if len(day_name) > 10:
                day_name = day_name[:3]

            html_content += f"""
                <tr>
                    <td>{current_date.strftime('%Y-%m-%d')}</td>
                    <td>{day_name}</td>
            """

            checkin = next((c for c in checkins if c.checkin_date == current_date), None)

            if checkin:
                checkin_count += 1

                # Get responses for each category
                for category in all_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        category_id=category.id,
                        response_date=current_date
                    ).first()

                    if response:
                        value = response.value
                        category_totals[category.id].append(value)

                        # Determine CSS class based on category type
                        if 'anxiety' in category.name.lower():
                            # Reverse colors for anxiety (low is good)
                            if value <= 2:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'
                        else:
                            # Normal colors (high is good)
                            if value >= 4:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'

                        html_content += f'<td class="{css_class}">{value}</td>'
                    else:
                        html_content += '<td>-</td>'

                for custom_cat in custom_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        custom_category_id=custom_cat.id,
                        response_date=current_date
                    ).first()

                    if response:
                        value = response.value
                        # Track custom category values
                        category_totals[f'custom_{custom_cat.id}'].append(value)

                        # Determine CSS class based on reverse scoring
                        if custom_cat.reverse_scoring:
                            # Reverse colors (low is good)
                            if value <= 2:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'
                        else:
                            # Normal colors (high is good)
                            if value >= 4:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'

                        html_content += f'<td class="{css_class}">{value}</td>'
                    else:
                        html_content += '<td>-</td>'
            else:
                # FIXED: No check-in this day - use correct colspan
                html_content += f'<td colspan="{len(all_category_data)}" class="no-checkin">{trans("no_checkin")}</td>'

            html_content += "</tr>"

        html_content += "</table>"

        # Add weekly summary section
        completion_rate = (checkin_count / 7) * 100

        html_content += f"""
            <div class="summary">
                <h2>{trans('weekly_summary')}</h2>
                <div class="summary-item">
                    <strong>{trans('checkin_completion')}:</strong> {checkin_count}/7 {trans('days')} ({completion_rate:.0f}%)
                </div>
        """

        # Add category averages
        for category in all_categories:
            if category_totals.get(category.id) and len(category_totals[category.id]) > 0:
                values = category_totals[category.id]
                avg_value = sum(values) / len(values)
                cat_name = translate_category_name(category.name, lang)

                # Determine rating based on category type
                if 'anxiety' in category.name.lower():
                    # Reverse logic for anxiety
                    if avg_value <= 2:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value <= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'
                else:
                    # Normal logic
                    if avg_value >= 4:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value >= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'

                html_content += f"""
                    <div class="summary-item">
                        <strong>{cat_name}:</strong> {avg_value:.1f}/5 - <span class="{rating_class}">{rating}</span>
                    </div>
                """

        # Add custom category averages
        for custom_cat in custom_categories:
            key = f'custom_{custom_cat.id}'
            if category_totals.get(key) and len(category_totals[key]) > 0:
                values = category_totals[key]
                avg_value = sum(values) / len(values)
                cat_name = custom_cat.name

                # Determine rating based on reverse scoring
                if custom_cat.reverse_scoring:
                    if avg_value <= 2:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value <= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'
                else:
                    if avg_value >= 4:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value >= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'

                html_content += f"""
                    <div class="summary-item">
                        <strong>{cat_name}:</strong> {avg_value:.1f}/5 - <span class="{rating_class}">{rating}</span>
                    </div>
                """

        html_content += """
            </div>
        </body>
        </html>
        """

        # Generate PDF with WeasyPrint
        print("Attempting to generate PDF with WeasyPrint...")
        pdf_buffer = BytesIO()
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)

        print("Successfully generated PDF with WeasyPrint")
        return pdf_buffer

    except ImportError as e:
        print(f"WeasyPrint not available: {e}")
    except Exception as e:
        print(f"Error generating PDF with WeasyPrint: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()

    # FALLBACK: xhtml2pdf implementation
    print("Falling back to xhtml2pdf...")
    try:
        from xhtml2pdf import pisa
        from io import BytesIO

        # Get translations
        trans = lambda key: translate_report_term(key, lang)
        days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])

        # Generate HTML content with proper styling
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4 landscape;
                    margin: 1cm;
                }}
                body {{
                    font-family: Arial, sans-serif;
                    font-size: 10pt;
                    direction: {'rtl' if lang in ['he', 'ar'] else 'ltr'};
                }}
                h1 {{
                    text-align: center;
                    color: #2C3E50;
                    font-size: 18pt;
                    margin-bottom: 10px;
                }}
                .subtitle {{
                    text-align: center;
                    color: #34495E;
                    margin-bottom: 20px;
                }}
                h2 {{
                    color: #34495E;
                    font-size: 14pt;
                    margin-top: 20px;
                    margin-bottom: 10px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                    font-size: 8pt;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 4px;
                    text-align: center;
                }}
                th {{
                    background-color: #2C3E50;
                    color: white;
                    font-weight: bold;
                    padding: 6px 4px;
                }}
                td {{
                    height: 25px;
                }}
                .good {{
                    background-color: #C8E6C9;
                    font-weight: bold;
                }}
                .medium {{
                    background-color: #FFF9C4;
                }}
                .poor {{
                    background-color: #FFCDD2;
                }}
                .no-checkin {{
                    color: #999;
                    font-style: italic;
                }}
                .summary {{
                    margin-top: 30px;
                    padding: 15px;
                    background-color: #f8f9fa;
                    border-radius: 5px;
                }}
                .summary h2 {{
                    margin-top: 0;
                }}
                .summary-item {{
                    margin: 8px 0;
                    padding: 5px 0;
                }}
                .summary-stats {{
                    display: table;
                    width: 100%;
                    margin-top: 15px;
                }}
                .stat-row {{
                    display: table-row;
                }}
                .stat-label {{
                    display: table-cell;
                    width: 60%;
                    padding: 5px;
                    font-weight: bold;
                }}
                .stat-value {{
                    display: table-cell;
                    width: 40%;
                    padding: 5px;
                }}
                .excellent-text {{ color: #2e7d32; font-weight: bold; }}
                .good-text {{ color: #f57c00; font-weight: bold; }}
                .poor-text {{ color: #c62828; font-weight: bold; }}
            </style>
        </head>
        <body>
            <h1>{trans('weekly_report_title')} - {trans('client')} {escape(client.client_name if client.client_name else client.client_serial)}</h1>
            <p class="subtitle">{trans('week')} {week_num}, {year} ({get_translated_month(week_start, lang)} {week_start.day} - {get_translated_month(week_end, lang)} {week_end.day}, {year})</p>

            <h2>{trans('daily_checkins')}</h2>
            <table>
                <tr>
                    <th style="width: 12%">{trans('date')}</th>
                    <th style="width: 12%">{trans('day')}</th>
        """

        # Get all categories
        all_categories = list(TrackingCategory.query.all())
        print(f"xhtml2pdf: Found {len(all_categories)} tracking categories")

        # Add custom categories for this client
        custom_categories = CustomCategory.query.filter_by(
            client_id=client.id,
            is_active=True
        ).all()
        print(f"xhtml2pdf: Found {len(custom_categories)} custom categories for client {client.id}")

        # Combine all categories
        all_category_data = []
        for cat in all_categories:
            all_category_data.append({
                'id': cat.id,
                'name': cat.name,
                'is_custom': False,
                'reverse_scoring': 'anxiety' in cat.name.lower()
            })

        for custom_cat in custom_categories:
            all_category_data.append({
                'id': f'custom_{custom_cat.id}',
                'name': custom_cat.name,
                'is_custom': True,
                'reverse_scoring': custom_cat.reverse_scoring
            })

        print(f"xhtml2pdf: Total categories to display: {len(all_category_data)}")

        # Calculate column width for categories
        remaining_width = 76  # 100% - 24% (date + day columns)
        col_width = remaining_width // len(all_category_data) if all_category_data else remaining_width

        # Add category headers
        for cat_data in all_category_data:
            cat_name = translate_category_name(cat_data['name'], lang) if not cat_data['is_custom'] else cat_data[
                'name']
            # Shorten very long category names to fit
            if len(cat_name) > 15:
                cat_name = cat_name[:13] + '..'
            html_content += f'<th style="width: {col_width}%">{escape(cat_name)}</th>'

        html_content += "</tr>"

        # Get check-ins
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).order_by(DailyCheckin.checkin_date).all()
        print(f"xhtml2pdf: Found {len(checkins)} checkins for the week")

        # Track statistics
        checkin_count = 0
        category_totals = {cat.id: [] for cat in all_categories}
        # FIXED: Also initialize totals for custom categories
        for custom_cat in custom_categories:
            category_totals[f'custom_{custom_cat.id}'] = []

        # Add daily data rows
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            day_name = days[i]
            # Shorten day names if needed
            if len(day_name) > 10:
                day_name = day_name[:3]

            html_content += f"""
                <tr>
                    <td>{current_date.strftime('%Y-%m-%d')}</td>
                    <td>{day_name}</td>
            """

            checkin = next((c for c in checkins if c.checkin_date == current_date), None)

            if checkin:
                checkin_count += 1

                # Get responses for each category
                for category in all_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        category_id=category.id,
                        response_date=current_date
                    ).first()

                    if response:
                        value = response.value
                        category_totals[category.id].append(value)

                        # Determine CSS class based on category type
                        if 'anxiety' in category.name.lower():
                            # Reverse colors for anxiety (low is good)
                            if value <= 2:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'
                        else:
                            # Normal colors (high is good)
                            if value >= 4:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'

                        html_content += f'<td class="{css_class}">{value}</td>'
                    else:
                        html_content += '<td>-</td>'

                for custom_cat in custom_categories:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        custom_category_id=custom_cat.id,  # Note: using custom_category_id field
                        response_date=current_date
                    ).first()

                    if response:
                        value = response.value
                        # Track custom category values
                        category_totals[f'custom_{custom_cat.id}'].append(value)

                        # Determine CSS class based on reverse scoring
                        if custom_cat.reverse_scoring:
                            # Reverse colors (low is good)
                            if value <= 2:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'
                        else:
                            # Normal colors (high is good)
                            if value >= 4:
                                css_class = 'good'
                            elif value == 3:
                                css_class = 'medium'
                            else:
                                css_class = 'poor'

                        html_content += f'<td class="{css_class}">{value}</td>'
                    else:
                        html_content += '<td>-</td>'
            else:
                # FIXED: No check-in this day - use correct colspan
                html_content += f'<td colspan="{len(all_category_data)}" class="no-checkin">{trans("no_checkin")}</td>'

            html_content += "</tr>"

        html_content += "</table>"

        # Add weekly summary section
        completion_rate = (checkin_count / 7) * 100

        html_content += f"""
            <div class="summary">
                <h2>{trans('weekly_summary')}</h2>
                <div class="summary-stats">
                    <div class="stat-row">
                        <div class="stat-label">{trans('checkin_completion')}:</div>
                        <div class="stat-value">{checkin_count}/7 {trans('days')} ({completion_rate:.0f}%)</div>
                    </div>
        """

        # Add category averages
        for category in all_categories:
            if category_totals.get(category.id) and len(category_totals[category.id]) > 0:
                values = category_totals[category.id]
                avg_value = sum(values) / len(values)
                cat_name = translate_category_name(category.name, lang)

                # Determine rating based on category type
                if 'anxiety' in category.name.lower():
                    # Reverse logic for anxiety
                    if avg_value <= 2:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value <= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'
                else:
                    # Normal logic
                    if avg_value >= 4:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value >= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'

                html_content += f"""
                    <div class="stat-row">
                        <div class="stat-label">{cat_name}:</div>
                        <div class="stat-value">{avg_value:.1f}/5 - <span class="{rating_class}">{rating}</span></div>
                    </div>
                """

        # Add custom category averages
        for custom_cat in custom_categories:
            key = f'custom_{custom_cat.id}'
            if category_totals.get(key) and len(category_totals[key]) > 0:
                values = category_totals[key]
                avg_value = sum(values) / len(values)
                cat_name = custom_cat.name

                # Determine rating based on reverse scoring
                if custom_cat.reverse_scoring:
                    if avg_value <= 2:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value <= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'
                else:
                    if avg_value >= 4:
                        rating = trans('excellent')
                        rating_class = 'excellent-text'
                    elif avg_value >= 3:
                        rating = trans('good')
                        rating_class = 'good-text'
                    else:
                        rating = trans('needs_support')
                        rating_class = 'poor-text'

                html_content += f"""
                    <div class="stat-row">
                        <div class="stat-label">{cat_name}:</div>
                        <div class="stat-value">{avg_value:.1f}/5 - <span class="{rating_class}">{rating}</span></div>
                    </div>
                """

        html_content += """
                </div>
            </div>
        </body>
        </html>
        """

        # Convert HTML to PDF
        print("Attempting to generate PDF with xhtml2pdf...")
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(
            html_content.encode('utf-8'),
            dest=pdf_buffer,
            encoding='utf-8'
        )

        if pisa_status.err:
            print(f"xhtml2pdf error: {pisa_status.err}")
            # Still return the buffer even if there was an error

        pdf_buffer.seek(0)
        print("Successfully generated PDF with xhtml2pdf")
        return pdf_buffer

    except ImportError as e:
        print(f"xhtml2pdf not available: {e}")
    except Exception as e:
        print(f"xhtml2pdf also failed: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()

    # Final fallback - return a simple error message PDF
    print("Both PDF libraries failed, creating fallback PDF")
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    if lang != 'en':
        elements.append(Paragraph(
            "PDF generation with Unicode support requires xhtml2pdf. Please install it or use Excel format for non-English reports.",
            styles['Normal']
        ))
    else:
        elements.append(Paragraph(
            "PDF generation failed. Please use Excel format instead.",
            styles['Normal']
        ))

    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer


# ============= REPORT GENERATION =============

@app.route('/api/reports/generate/<int:client_id>/<week>', methods=['GET'])
@require_auth(['therapist'])
@limiter.limit("100 per hour")  # Allow more report generation
def generate_report(client_id, week):
    """Generate comprehensive weekly Excel report with streaming"""
    try:
        therapist = request.current_user.therapist
        lang = get_language_from_header()

        logger.info('report_generation_attempt', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_id': client_id,
                'week': week,
                'report_type': 'excel',
                'language': lang,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            logger.warning('report_generation_failed', extra={
                'extra_data': {
                    'reason': 'client_not_found',
                    'client_id': client_id,
                    'therapist_id': therapist.id,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })
            return jsonify({'error': 'Client not found'}), 404

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Generate filename
        filename = f"therapy_report_{sanitize_input(client.client_serial)}_week_{week_num}_{year}.xlsx"

        # Log report generation start
        generation_start = time.time()

        # Stream the file generation
        # Create Excel workbook
        # Use streaming version for better memory efficiency
        output = create_weekly_report_excel_streaming(client, therapist, week_start, week_end, week_num, year, lang)

        # Log successful generation
        generation_time = time.time() - generation_start
        logger.info('report_generation_success', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_id': client_id,
                'client_serial': client.client_serial,
                'week': week,
                'report_type': 'excel',
                'filename': filename,
                'generation_time_ms': round(generation_time * 1000, 2),
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Return file directly
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error('report_generation_error', extra={
            'extra_data': {
                'error': str(e),
                'therapist_id': therapist.id if therapist else None,
                'client_id': client_id,
                'week': week,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/reports/generate-pdf/<int:client_id>/<week>', methods=['GET'])
@require_auth(['therapist'])
def generate_pdf_report(client_id, week):
    """Generate PDF weekly report for therapist"""
    try:
        therapist = request.current_user.therapist
        lang = get_language_from_header()

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400
        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # DEBUG: Add these lines
        print(f"DEBUG generate_pdf_report:")
        print(f"  week_start: {week_start} (type: {type(week_start)})")
        print(f"  week_end: {week_end} (type: {type(week_end)})")
        print(f"  These are datetime objects: {isinstance(week_start, datetime)}")

        # Create PDF
        pdf_buffer = create_weekly_report_pdf(client, therapist, week_start, week_end, week_num, year, lang)

        # Generate filename
        filename = f"therapy_report_{sanitize_input(client.client_serial)}_week_{week_num}_{year}.pdf"

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/generate-pdf/<week>', methods=['GET'])
@require_auth(['client'])
def client_generate_pdf(week):
    """Generate client's own weekly PDF report"""
    try:
        client = request.current_user.client
        lang = get_language_from_header()

        # Parse week
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Create PDF
        pdf_buffer = create_weekly_report_pdf(client, None, week_start, week_end, week_num, year, lang)

        # Generate filename
        filename = f"my_therapy_report_{sanitize_input(client.client_serial)}_week_{week_num}_{year}.pdf"

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/weekly-report-settings', methods=['GET'])
@require_auth(['therapist'])
def get_weekly_report_settings():
    """Get therapist's weekly report settings"""
    try:

        try:
            db.session.execute(text("""
                        ALTER TABLE reminders
                        ADD COLUMN IF NOT EXISTS day_of_week INTEGER DEFAULT 1
                    """))
            db.session.commit()
        except Exception:
            db.session.rollback()

        therapist = request.current_user.therapist




        therapist = request.current_user.therapist

        # Check if settings exist
        existing = db.session.execute(
            text("""
                SELECT reminder_time, reminder_email, day_of_week, local_reminder_time, reminder_language
                FROM reminders
                WHERE client_id = :therapist_id
                AND reminder_type = 'weekly_report'
                AND is_active = true
            """),
            {'therapist_id': therapist.id}
        ).first()

        if existing:
            return jsonify({
                'success': True,
                'settings': {
                    'time': existing[0].strftime('%H:%M') if existing[0] else '09:00',
                    'email': existing[1] or '',
                    'day_of_week': existing[2] if existing[2] is not None else 1,
                    'local_time': existing[3] or '09:00',
                    'language': existing[4] or 'en'
                }
            })

        return jsonify({
            'success': True,
            'settings': None
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/weekly-report-settings', methods=['POST'])
@require_auth(['therapist'])
def save_weekly_report_settings():
    """Save therapist's weekly report settings"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        time_str = data.get('time', '09:00')
        day_of_week = data.get('day_of_week', 1)
        email = data.get('email', '').strip()
        timezone_offset = data.get('timezone_offset', 0)
        language = data.get('language', 'en')
        user_timezone = data.get('timezone', 'Asia/Jerusalem')

        # Parse time
        hour, minute = map(int, time_str.split(':'))

        # Convert to UTC
        local_total_minutes = hour * 60 + minute
        utc_total_minutes = local_total_minutes - timezone_offset



        while utc_total_minutes < 0:
            utc_total_minutes += 24 * 60
        while utc_total_minutes >= 24 * 60:
            utc_total_minutes -= 24 * 60

        utc_hour = utc_total_minutes // 60
        utc_minute = utc_total_minutes % 60

        import pytz
        from datetime import datetime

        # Create UTC datetime
        utc_now = datetime.utcnow()
        utc_dt = utc_now.replace(hour=utc_hour, minute=utc_minute, second=0, microsecond=0)

        # Convert to Jerusalem timezone
        jerusalem_tz = pytz.timezone('Asia/Jerusalem')
        utc_tz = pytz.timezone('UTC')
        utc_dt = utc_tz.localize(utc_dt)
        jerusalem_dt = utc_dt.astimezone(jerusalem_tz)

        # This is the actual Jerusalem time
        jerusalem_time_str = f"{jerusalem_dt.hour:02d}:{jerusalem_dt.minute:02d}"



        import datetime
        time_obj = datetime.time(utc_hour, utc_minute)

        # Check if settings exist
        existing = Reminder.query.filter_by(
            client_id=therapist.id,  # Using therapist.id as client_id for storage
            reminder_type='weekly_report'
        ).first()

        if existing:
            existing.reminder_time = time_obj
            existing.local_reminder_time = time_str
            existing.reminder_email = email if email else None
            existing.day_of_week = day_of_week
            existing.reminder_language = language
            existing.is_active = True
        else:
            reminder = Reminder(
                client_id=therapist.id,  # Using therapist.id as client_id for storage
                reminder_type='weekly_report',
                reminder_time=time_obj,
                local_reminder_time=time_str,
                reminder_email=email if email else None,
                day_of_week=day_of_week,
                reminder_language=language,
                is_active=True
            )
            db.session.add(reminder)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Weekly report settings saved successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/test-weekly-report', methods=['POST'])
@require_auth(['therapist'])
def test_weekly_report():
    """Send a test weekly report email"""
    try:
        therapist = request.current_user.therapist

        # Get settings
        settings = Reminder.query.filter_by(
            client_id=therapist.id,
            reminder_type='weekly_report'
        ).first()

        email = therapist.user.email
        if settings and settings.reminder_email:
            email = settings.reminder_email

        # Get ALL active clients
        active_clients = therapist.clients.filter_by(is_active=True).all()
        if not active_clients:
            return jsonify({'error': 'No active clients found'}), 400

        # Get current week
        import datetime
        today = datetime.date.today()
        week_start = today - datetime.timedelta(days=today.weekday())
        week_end = week_start + datetime.timedelta(days=6)
        year = today.year
        week_num = today.isocalendar()[1]

        # Create email
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_USERNAME']
        msg['To'] = email

        # Language
        lang = settings.reminder_language if settings else 'en'

        # Translated subjects
        subjects = {
            'en': f"Weekly Therapy Report - Week {week_num}, {year}",
            'he': f"דוח טיפולי שבועי - שבוע {week_num}, {year}",
            'ru': f"Еженедельный терапевтический отчет - Неделя {week_num}, {year}",
            'ar': f"التقرير العلاجي الأسبوعي - الأسبوع {week_num}, {year}"
        }
        msg['Subject'] = subjects.get(lang, subjects['en'])

        # Translated body
        bodies = {
            'en': "Weekly checkin time. Reminder to see how your clients are doing.",
            'he': "זמן סיכום שבועי. תזכורת לבדוק איך מסתדרים המטופלים שלך.",
            'ru': "Время еженедельной проверки. Напоминание проверить, как дела у ваших клиентов.",
            'ar': "وقت الفحص الأسبوعي. تذكير لمعرفة كيف حال عملائك."
        }
        body = bodies.get(lang, bodies['en'])
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Generate and attach PDF for EACH client
        from io import BytesIO
        attachment_count = 0

        for client in active_clients:
            try:
                pdf_buffer = create_weekly_report_pdf(
                    client, therapist, week_start, week_end, week_num, year, lang
                )

                # Attach PDF
                pdf_attachment = MIMEBase('application', 'pdf')
                pdf_attachment.set_payload(pdf_buffer.read())
                encoders.encode_base64(pdf_attachment)
                safe_name = client.client_name.replace(' ', '_').replace('/', '_').replace('\\','_') if client.client_name else client.client_serial
                pdf_attachment.add_header(
                    'Content-Disposition',
                    f'attachment; filename=report_{safe_name}_week_{week_num}_{year}.pdf'
                )
                msg.attach(pdf_attachment)
                attachment_count += 1

            except Exception as e:
                print(f"Failed to generate PDF for client {client.client_serial}: {e}")

        if attachment_count == 0:
            return jsonify({'error': 'Failed to generate any PDF reports'}), 500

        # Send email
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.starttls()
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        server.send_message(msg)
        server.quit()

        return jsonify({
            'success': True,
            'message': f'Test report sent successfully with {attachment_count} client reports'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/test-weekly-report-batch', methods=['POST'])
@require_auth(['therapist'])
def test_weekly_report_batch():
    """Send weekly report in batches for therapists with many clients"""
    try:
        therapist = request.current_user.therapist
        data = request.json
        batch_size = data.get('batch_size', 10)

        # Validate batch size
        if batch_size < 1 or batch_size > 50:
            batch_size = 10

        # Check if Celery is available
        if not celery:
            return jsonify({'error': 'Background task system not available'}), 503

        # Check client count
        client_count = therapist.clients.filter_by(is_active=True).count()

        # Use batch task if more than 15 clients, otherwise use regular task
        if client_count > 15:
            from celery_app import send_weekly_report_batch_task
            task = send_weekly_report_batch_task.delay(therapist.id, batch_size)
        else:
            from celery_app import send_weekly_report_task
            task = send_weekly_report_task.delay(therapist.id)

        return jsonify({
            'success': True,
            'task_id': task.id,
            'message': f'Report generation started for {client_count} clients.',
            'batch_mode': client_count > 15
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500






# UPDATED EMAIL REPORT FUNCTION
@app.route('/api/therapist/email-report', methods=['POST'])
@require_auth(['therapist'])
def email_therapy_report():
    """Send therapy report via email with translations"""
    try:
        therapist = request.current_user.therapist
        lang = get_language_from_header()
        data = request.json

        client_id = data.get('client_id')
        week = data.get('week')
        recipient_email = data.get('recipient_email')

        # Validate input
        if not client_id:
            return jsonify({'error': 'Client ID is required'}), 400

        if not week:
            return jsonify({'error': 'Week is required'}), 400

        # Ensure client_id is an integer
        try:
            client_id = int(client_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid client ID'}), 400

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week
        # Validate week format
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        try:
            year, week_num = week.split('-W')
            year = int(year)
            week_num = int(week_num)
        except (ValueError, AttributeError):
            return jsonify({'error': 'Invalid week format'}), 400

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get check-ins for summary
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).order_by(DailyCheckin.checkin_date).all()

        checkins_completed = len(checkins)

        # Build email content with translations
        trans = lambda key: translate_report_term(key, lang)

        # Calculate summary statistics from category responses
        summary_stats = []
        for category in TrackingCategory.query.all():
            responses = []
            for checkin in checkins:
                response = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    category_id=category.id,
                    response_date=checkin.checkin_date
                ).first()
                if response:
                    responses.append(response.value)

            if responses:
                avg_value = sum(responses) / len(responses)
                cat_name = translate_category_name(category.name, lang)
                summary_stats.append(f"- {cat_name}: {avg_value:.1f}/5")

        # Generate language-specific email content
        if lang == 'he':
            email_content = f"""
שלום,

מצורף דוח הטיפול השבועי עבור מטופל {sanitize_input(client.client_serial)}.

תקופת הדוח: {week_start.day} {get_translated_month(week_start, lang)} - {week_end.day} {get_translated_month(week_end, lang)}, {year} (שבוע {week_num})

סיכום:
- צ'ק-אינים שהושלמו: {checkins_completed}/7 ימים
{chr(10).join(summary_stats)}

הקובץ המצורף כולל:
- צ'ק-אינים יומיים מפורטים עם דירוגים צבעוניים
- סטטיסטיקות סיכום שבועיות
- מעקב השלמת יעדים
- הערות ומשימות מטפל

אנא עיין/י בדוח המצורף וצור/י קשר אם יש שאלות.

בברכה,
{escape(therapist.name)}
{escape(therapist.organization or '')}
        """
        elif lang == 'ru':
            email_content = f"""
Здравствуйте,

Прилагается еженедельный терапевтический отчет для клиента {sanitize_input(client.client_serial)}.

Период отчета: {week_start.day} {get_translated_month(week_start, lang)} - {week_end.day} {get_translated_month(week_end, lang)} {year} (Неделя {week_num})

Сводка:
- Выполнено отметок: {checkins_completed}/7 дней
{chr(10).join(summary_stats)}

Прилагаемый файл Excel содержит:
- Подробные ежедневные отметки с цветовой кодировкой
- Еженедельная сводная статистика
- Отслеживание выполнения целей
- Заметки и задания терапевта

Пожалуйста, просмотрите прилагаемый отчет и свяжитесь со мной, если у вас есть вопросы.

С наилучшими пожеланиями,
{escape(therapist.name)}
{escape(therapist.organization or '')}
        """
        elif lang == 'ar':
            email_content = f"""
مرحباً،

يرجى الاطلاع على التقرير العلاجي الأسبوعي المرفق للعميل {sanitize_input(client.client_serial)}.

:فترة التقرير: {week_start.day} {get_translated_month(week_start, lang)} - {week_end.day} {get_translated_month(week_end, lang)} {year} (الأسبوع {week_num})

الملخص:
- تسجيلات الحضور المكتملة: {checkins_completed}/7 أيام
{chr(10).join(summary_stats)}

يحتوي ملف Excel المرفق على:
- تسجيلات الحضور اليومية المفصلة مع التقييمات الملونة
- إحصائيات الملخص الأسبوعي
- تتبع إنجاز الأهداف
- ملاحظات ومهام المعالج

يرجى مراجعة التقرير المرفق والاتصال بي إذا كان لديك أي أسئلة.

مع أطيب التحيات،
{therapist.name}
{therapist.organization or ''}
        """
        else:  # Default to English
            email_content = f"""
Dear Colleague,

Please find attached the weekly therapy report for client {sanitize_input(client.client_serial)}.

Report Period: {week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')} (Week {week_num}, {year})

Summary:
- Check-ins completed: {checkins_completed}/7 days
{chr(10).join(summary_stats)}

The attached Excel file contains:
- Detailed daily check-ins with color-coded ratings
- Weekly summary statistics
- Goal completion tracking
- Therapist notes and missions

Please review the attached report and contact me if you have any questions.

Best regards,
{therapist.name}
{therapist.organization or ''}
        """

        # If no email configuration, return content for manual sending
        if not app.config.get('MAIL_USERNAME'):
            return jsonify({
                'success': True,
                'email_content': email_content.strip(),
                'recipient': recipient_email or therapist.user.email or 'Your email',
                'subject': f"Weekly Therapy Report - Client {sanitize_input(client.client_serial)} - Week {week_num}, {year}",
                'note': 'Email configuration not set up. Please copy this content and attach the downloaded Excel file to send manually.'
            })

        # If email is configured, send it
        try:
            # Create the Excel workbook using the shared function
            excel_buffer = create_weekly_report_excel_streaming(client, therapist, week_start, week_end, week_num, year, lang)





            # Create email
            msg = MIMEMultipart()
            msg['From'] = app.config['MAIL_USERNAME']
            msg['To'] = recipient_email or therapist.user.email
            msg['Subject'] = f"Weekly Therapy Report - Client {sanitize_input(client.client_serial)} - Week {week_num}, {year}"

            # Email body
            msg.attach(MIMEText(email_content, 'plain'))

            # Attach Excel file
            excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_attachment.set_payload(excel_buffer.read())
            encoders.encode_base64(excel_attachment)
            excel_attachment.add_header(
                'Content-Disposition',
                f'attachment; filename=therapy_report_{sanitize_input(client.client_serial)}_week_{week_num}_{year}.xlsx'
            )
            msg.attach(excel_attachment)

            # Create PDF attachment
            pdf_buffer = create_weekly_report_pdf(client, therapist, week_start, week_end, week_num, year, lang)
            pdf_attachment = MIMEBase('application', 'pdf')
            pdf_attachment.set_payload(pdf_buffer.read())
            encoders.encode_base64(pdf_attachment)
            safe_name = client.client_name.replace(' ', '_').replace('/', '_').replace('\\','_') if client.client_name else client.client_serial
            pdf_attachment.add_header(
                'Content-Disposition',
                f'attachment; filename=report_{safe_name}_week_{week_num}_{year}.pdf'
            )
            msg.attach(pdf_attachment)

            # Send email
            server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
            server.starttls()
            server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
            server.send_message(msg)
            server.quit()

            return jsonify({
                'success': True,
                'message': f'Report sent successfully to {recipient_email or therapist.user.email}'
            })

        except Exception as e:
            # Return the email content even if sending fails
            return jsonify({
                'success': True,
                'email_content': email_content.strip(),
                'recipient': recipient_email or therapist.user.email or 'Your email',
                'subject': f"Weekly Therapy Report - Client {sanitize_input(client.client_serial)} - Week {week_num}, {year}",
                'error': f'Failed to send email: {str(e)}',
                'note': 'Email could not be sent automatically. Please copy this content and send it manually.'
            })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= HEALTH CHECK =============

@app.route('/api/health', methods=['GET'])
@cache_response(timeout=60)
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat()
    })


@app.route('/api/health/detailed', methods=['GET'])
@require_auth(['therapist'])  # Or create a special monitoring role
def detailed_health_check():
    """Detailed health check for monitoring"""
    health_status = {
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'services': {}
    }

    # Check database
    try:
        db.session.execute(text('SELECT 1'))
        health_status['services']['database'] = {'status': 'healthy'}
    except Exception as e:
        health_status['services']['database'] = {'status': 'unhealthy', 'error': str(e)}
        health_status['status'] = 'unhealthy'

    # Check Redis
    try:
        if redis_client:
            redis_client.ping()
            health_status['services']['redis'] = {'status': 'healthy'}
        else:
            health_status['services']['redis'] = {'status': 'unavailable'}
    except Exception as e:
        health_status['services']['redis'] = {'status': 'unhealthy', 'error': str(e)}

    # Check Celery
    if celery:
        try:
            inspector = celery.control.inspect()
            stats = inspector.stats()
            if stats:
                health_status['services']['celery'] = {
                    'status': 'healthy',
                    'workers': len(stats)
                }
            else:
                health_status['services']['celery'] = {'status': 'no_workers'}
        except Exception as e:
            health_status['services']['celery'] = {'status': 'unhealthy', 'error': str(e)}
    else:
        health_status['services']['celery'] = {'status': 'not_configured'}

    # Check email
    email_configured = bool(app.config.get('MAIL_USERNAME'))
    health_status['services']['email'] = {
        'status': 'configured' if email_configured else 'not_configured'
    }

    # Feature availability based on service health
    health_status['features'] = {
        'reminders': health_status['services']['celery']['status'] == 'healthy',
        'email_reports': email_configured,
        'search': search_manager.es is not None
    }

    return jsonify(health_status), 200 if health_status['status'] == 'healthy' else 503

@app.route('/api/test-pdf-libs', methods=['GET'])
def test_pdf_libs():
    """Test PDF library availability"""
    results = test_pdf_libraries()
    return jsonify(results)



# ============= ERROR HANDLERS =============

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    logger.warning('not_found', extra={
        'extra_data': {
            'path': request.path,
            'method': request.method,
            'request_id': getattr(g, 'request_id', 'unknown')
        },
        'request_id': getattr(g, 'request_id', 'unknown')
    })
    return jsonify({
        'error': 'Resource not found',
        'request_id': getattr(g, 'request_id', 'unknown')
    }), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    db.session.rollback()
    logger.error('internal_server_error', extra={
        'extra_data': {
            'error': str(error),
            'request_id': getattr(g, 'request_id', 'unknown')
        },
        'request_id': getattr(g, 'request_id', 'unknown')
    }, exc_info=True)
    return jsonify({
        'error': 'Internal server error',
        'request_id': getattr(g, 'request_id', 'unknown')
    }), 500


@app.errorhandler(Exception)
def handle_exception(error):
    """Handle unexpected exceptions"""
    db.session.rollback()

    logger.error('unhandled_exception', extra={
        'extra_data': {
            'error_type': type(error).__name__,
            'error_message': str(error),
            'path': request.path,
            'method': request.method,
            'request_id': getattr(g, 'request_id', 'unknown')
        },
        'request_id': getattr(g, 'request_id', 'unknown')
    }, exc_info=True)

    return jsonify({
        'error': 'An unexpected error occurred',
        'request_id': getattr(g, 'request_id', 'unknown')
    }), 500


# ============= STATIC FILE SERVING =============


@app.route('/favicon.svg')
def favicon():
    """Serve favicon"""
    favicon_path = os.path.join(BASE_DIR, 'favicon.svg')
    if os.path.exists(favicon_path):
        return send_file(favicon_path, mimetype='image/svg+xml')
    else:
        return '', 204


# ============= INITIALIZATION =============
def ensure_client_names():
    """Ensure all clients have names - use email as default"""
    try:
        # Check if client_name column exists
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        columns = [col['name'] for col in inspector.get_columns('clients')]

        if 'client_name' not in columns:
            # Add the column if it doesn't exist
            db.session.execute(text('ALTER TABLE clients ADD COLUMN client_name VARCHAR(255)'))
            db.session.commit()
            print("Added client_name column to clients table")

        # Update all clients without names
        clients_without_names = Client.query.filter(
            or_(Client.client_name == None, Client.client_name == '')
        ).all()

        updated_count = 0
        for client in clients_without_names:
            user = User.query.get(client.user_id)
            if user:
                client.client_name = user.email
                updated_count += 1
                print(f"Updated client {sanitize_input(client.client_serial)} with name: {user.email}")
            else:
                # Fallback to serial if user not found
                client.client_name = client.client_serial
                updated_count += 1
                print(f"Updated client {sanitize_input(client.client_serial)} with name: {sanitize_input(client.client_serial)}")

        if updated_count > 0:
            db.session.commit()
            print(f"Updated {updated_count} clients with names")

    except Exception as e:
        print(f"Error ensuring client names: {e}")
        db.session.rollback()









# Flag to ensure single initialization
_initialized = False


def initialize_database():
    """Initialize database with default data"""
    global _initialized
    if _initialized:
        return

    _initialized = True

    # Use database-based locking to prevent race conditions
    lock_acquired = False

    try:
        # Try to create a lock record in the database
        with db.session.begin_nested():
            # Create tables if they don't exist
            db.create_all()

            try:
                # Check if column exists
                result = db.session.execute("""
                                SELECT column_name
                                FROM information_schema.columns
                                WHERE table_name='reminders' AND column_name='day_of_week'
                            """).fetchone()

                if not result:
                    # Add the column
                    db.session.execute("""
                                    ALTER TABLE reminders
                                    ADD COLUMN day_of_week INTEGER DEFAULT 1
                                """)
                    db.session.commit()
                    print("Added day_of_week column to reminders table")
            except Exception as e:
                print(f"Note: Could not check/add day_of_week column: {e}")
                # This is okay - it might already exist


            # Check if already initialized
            from sqlalchemy import text
            existing_lock = db.session.execute(
                text("SELECT COUNT(*) FROM tracking_categories")
            ).scalar()

            if existing_lock >= 8:
                print("Database already initialized by another process")
                # Still ensure client names even if categories exist
                ensure_client_names()
                return

            # Use advisory lock for PostgreSQL
            db.session.execute("SELECT pg_advisory_lock(12345)")
            lock_acquired = True

            # Double-check after acquiring lock
            category_count = TrackingCategory.query.count()
            if category_count >= 8:
                print("Database already initialized by another process")
                # Still ensure client names even if categories exist
                ensure_client_names()
                return

            # Always ensure all default categories exist
            ensure_default_categories()

            # Automatically fix any existing clients
            fix_existing_clients()

            # Ensure all clients have names
            ensure_client_names()

            print("Database initialized with all default tracking categories")
            print("All existing clients have been updated with missing categories")
            print("All existing clients have been updated with names")

            # Commit the transaction
            db.session.commit()

    except Exception as e:
        print(f"Database initialization error: {e}")
        db.session.rollback()
        _initialized = False
    finally:
        if lock_acquired:
            try:
                db.session.execute("SELECT pg_advisory_unlock(12345)")
                db.session.commit()
            except:
                pass





def migrate_existing_data_if_needed():
    """Automatically encrypt existing data on first run"""
    try:
        # Check if migration has been done
        migration_done = db.session.execute(
            text("SELECT COUNT(*) FROM daily_checkins WHERE emotional_notes_encrypted IS NOT NULL")
        ).scalar() > 0

        if migration_done:
            return  # Already migrated

        # Get all checkins with unencrypted notes
        checkins = DailyCheckin.query.filter(
            or_(
                DailyCheckin.emotional_notes != None,
                DailyCheckin.medication_notes != None,
                DailyCheckin.activity_notes != None
            )
        ).all()

        migrated = 0
        for checkin in checkins:
            try:
                # Encrypt existing notes
                if checkin.emotional_notes:
                    checkin.emotional_notes_encrypted = encrypt_field(checkin.emotional_notes)
                if checkin.medication_notes:
                    checkin.medication_notes_encrypted = encrypt_field(checkin.medication_notes)
                if checkin.activity_notes:
                    checkin.activity_notes_encrypted = encrypt_field(checkin.activity_notes)
                migrated += 1
            except Exception as e:
                logger.error(f"Failed to encrypt checkin {checkin.id}: {e}")

        if migrated > 0:
            db.session.commit()
            logger.info(f"Successfully encrypted {migrated} existing checkins")

    except Exception as e:
        logger.error(f"Migration failed: {e}")
        db.session.rollback()







# Initialize on first request
def initialize_app_data():
    """Initialize application data - run once at startup"""
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)

        if 'tracking_categories' in inspector.get_table_names():
            with app.app_context():
                category_count = TrackingCategory.query.count()
                if category_count < 8:
                    print(f"Found only {category_count} categories, ensuring all defaults exist...")
                    ensure_default_categories()
                    fix_existing_clients()
                    print("Database initialization completed")
    except Exception as e:
        print(f"Error during initialization: {e}")


# Don't initialize on import for production
# Let init_db.py handle it during deployment
# Initialize based on environment
if not os.environ.get('PRODUCTION'):
    # Development mode
    with app.app_context():
        initialize_database()
else:
    # Production mode - ensure all initialization runs
    with app.app_context():
        try:
            from sqlalchemy import inspect

            inspector = inspect(db.engine)

            if 'users' in inspector.get_table_names():
                # Tables exist, run full initialization
                print("Running production startup initialization...")

                # Create any missing tables
                db.create_all()

                # Initialize database with categories
                initialize_database()

                # Ensure all clients have names
                ensure_client_names()

                # Run migrations if needed
                migrate_existing_data_if_needed()

                print("Production startup initialization complete")
            else:
                print("Database tables don't exist yet in production")

        except Exception as e:
            print(f"Error during production initialization: {e}")
            import traceback

            traceback.print_exc()

# ============= MAIN ENTRY POINT =============




@app.route('/api/therapist/create-client', methods=['POST'])
@require_auth(['therapist'])
def create_client():
    """Create new client with ALL categories guaranteed"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        # Create user account for client
        email = data.get('email')
        client_name = data.get('client_name', '').strip()

        # If no client name provided, use email
        if not client_name:
            client_name = email
        # Generate a secure password that meets requirements
        if not data.get('password'):
            # Generate components
            uppercase = ''.join(random.choice(string.ascii_uppercase) for _ in range(3))
            lowercase = ''.join(random.choice(string.ascii_lowercase) for _ in range(4))
            digits = ''.join(random.choice(string.digits) for _ in range(3))
            special = ''.join(random.choice('@$!%*?&') for _ in range(2))

            # Combine and shuffle
            password_chars = list(uppercase + lowercase + digits + special)
            random.shuffle(password_chars)
            password = ''.join(password_chars)
        else:
            password = data.get('password')

        logger.info('create_client_attempt', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_email': email,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        if User.query.filter_by(email=email).first():
            logger.warning('create_client_failed', extra={
                'extra_data': {
                    'reason': 'email_already_exists',
                    'email': email,
                    'therapist_id': therapist.id,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })
            return jsonify({'error': 'Email already registered'}), 400

        # Create user
        user = User(
            email=email,
            password_hash=bcrypt.generate_password_hash(password).decode('utf-8'),
            role='client'
        )
        db.session.add(user)
        db.session.flush()

        # Create client
        client = Client(
            user_id=user.id,
            client_serial=generate_client_serial(),
            client_name=client_name,
            therapist_id=therapist.id,
            start_date=date.today()
        )
        db.session.add(client)
        db.session.flush()

        # ALWAYS ensure we have all 8 categories before adding them
        ensure_default_categories()  # Just ensure they exist, don't use returned objects

        # Query categories fresh in current session
        all_categories = TrackingCategory.query.all()

        # Verify we have exactly 8 categories
        if len(all_categories) != 8:
            logger.warning('category_count_mismatch', extra={
                'extra_data': {
                    'expected': 8,
                    'found': len(all_categories),
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })
            # Try to fix it
            ensure_default_categories()
            all_categories = TrackingCategory.query.all()

        # Add ALL tracking categories to this client
        categories_added = 0
        for category in all_categories:
            plan = ClientTrackingPlan(
                client_id=client.id,
                category_id=category.id,
                is_active=True
            )
            db.session.add(plan)
            categories_added += 1

        logger.info('categories_assigned', extra={
            'extra_data': {
                'client_id': client.id,
                'client_serial': client.client_serial,
                'categories_added': categories_added,
                'request_id': g.request_id
            },
            'request_id': g.request_id
        })

        # Verify the count
        if categories_added != 8:
            logger.warning('incorrect_category_count', extra={
                'extra_data': {
                    'expected': 8,
                    'actual': categories_added,
                    'client_serial': client.client_serial,
                    'request_id': g.request_id
                },
                'request_id': g.request_id
            })

        # Add initial goals if provided
        goals = data.get('initial_goals', [])
        week_start = date.today() - timedelta(days=date.today().weekday())
        goals_added = 0
        for goal_text in goals:
            if goal_text.strip():
                goal = WeeklyGoal(
                    client_id=client.id,
                    therapist_id=therapist.id,
                    goal_text=sanitize_input(goal_text),
                    week_start=week_start
                )
                db.session.add(goal)
                goals_added += 1

        # Add welcome note
        welcome_note = TherapistNote(
            client_id=client.id,
            therapist_id=therapist.id,
            note_type='welcome',
            content=f"Welcome to therapy! Your journey begins today. Your temporary password is: {password}"
        )
        db.session.add(welcome_note)

        # Create custom categories if provided
        custom_categories = data.get('custom_categories', [])
        custom_categories_created = 0

        for custom_cat_data in custom_categories[:4]:  # Limit to 4 custom categories
            if custom_cat_data.get('name'):
                custom_category = CustomCategory(
                    therapist_id=therapist.id,
                    client_id=client.id,
                    name=custom_cat_data['name'],
                    description=custom_cat_data.get('description', ''),
                    reverse_scoring=custom_cat_data.get('reverse_scoring', False)
                )
                db.session.add(custom_category)
                custom_categories_created += 1

                # Create custom categories if provided
        custom_categories = data.get('custom_categories', [])
        custom_categories_created = 0

        for custom_cat_data in custom_categories[:4]:  # Limit to 4 custom categories
                    if custom_cat_data.get('name'):
                        custom_category = CustomCategory(
                            therapist_id=therapist.id,
                            client_id=client.id,
                            name=custom_cat_data['name'],
                            description=custom_cat_data.get('description', ''),
                            reverse_scoring=custom_cat_data.get('reverse_scoring', False)
                        )
                        db.session.add(custom_category)
                        custom_categories_created += 1

        if search_manager.es:
            checkin_stats = {
                'last_checkin': None,
                'checkin_count': 0,
                'avg_emotional': 0,
                'avg_medication': 0
            }
            search_manager.index_client(client, checkin_stats)




        db.session.commit()

                # Final verification
        client_categories = ClientTrackingPlan.query.filter_by(client_id=client.id).count()

        logger.info('create_client_success', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_id': client.id,
                'client_serial': client.client_serial,
                'categories_assigned': client_categories,
                'goals_added': goals_added,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        return jsonify({
            'success': True,
            'client': {
                 'id': client.id,
                'serial': client.client_serial,
                'client_name': client.client_name,
                'email': email,
                'temporary_password': password,
                'categories_assigned': client_categories,
                'expected_categories': 8,
                'all_categories_assigned': client_categories == 8
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error('create_client_error', extra={
            'extra_data': {
                'error': str(e),
                'therapist_id': therapist.id,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/add-goal', methods=['POST'])
@require_auth(['therapist'])
def add_weekly_goal():
    """Add weekly goal for client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        goal_text = data.get('goal_text')
        week_start_str = data.get('week_start')

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Parse week start
        if week_start_str:
            week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()
        else:
            # Default to current week
            week_start = date.today() - timedelta(days=date.today().weekday())

        # Create goal
        goal = WeeklyGoal(
            client_id=client_id,
            therapist_id=therapist.id,
            goal_text=sanitize_input(goal_text),
            week_start=week_start
        )
        db.session.add(goal)
        db.session.commit()

        return jsonify({
            'success': True,
            'goal': {
                'id': goal.id,
                'text': goal.goal_text,
                'week_start': goal.week_start.isoformat()
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/therapist/delete-client/<int:client_id>', methods=['DELETE'])
@require_auth(['therapist'])
def delete_client(client_id):
    """Delete a client and all associated data"""
    try:

        try:
            db.session.execute(text("""
                        ALTER TABLE reminders
                        ADD COLUMN IF NOT EXISTS day_of_week INTEGER DEFAULT 1
                    """))
            db.session.commit()
        except Exception:
            # Column might already exist, which is fine
            db.session.rollback()




        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Get user to delete
        user = client.user

        # Log the deletion for audit
        logger.info('client_deletion', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_id': client.id,
                'client_serial': client.client_serial,
                'client_email': user.email if user else 'unknown',
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Audit log for compliance
        log_audit(
            action='DELETE_CLIENT',
            resource_type='client',
            resource_id=client_id,
            details={
                'client_serial': client.client_serial,
                'therapist_id': therapist.id
            },
            phi_accessed=True
        )

        # Delete all related data (cascade should handle most)
        # But explicitly delete some for safety

        # Delete category responses
        CategoryResponse.query.filter_by(client_id=client_id).delete()

        # Delete goal completions through goals
        for goal in client.goals:
            GoalCompletion.query.filter_by(goal_id=goal.id).delete()

        # Delete custom categories
        CustomCategory.query.filter_by(client_id=client_id).delete()

        # Delete therapist notes
        TherapistNote.query.filter_by(client_id=client_id).delete()

        # Delete consent records
        ConsentRecord.query.filter_by(client_id=client_id).delete()

        # Delete the client (cascade will handle checkins, goals, reminders, tracking plans)

        # Fix for day_of_week column - ensure it exists before any reminder operations
        try:
            db.session.execute(text("""
                        ALTER TABLE reminders
                        ADD COLUMN IF NOT EXISTS day_of_week INTEGER DEFAULT 1
                    """))
            db.session.commit()
        except Exception:
            # Column might already exist or syntax might not support IF NOT EXISTS
            try:
                db.session.execute(text("""
                            ALTER TABLE reminders
                            ADD COLUMN day_of_week INTEGER DEFAULT 1
                        """))
                db.session.commit()
            except Exception:
                # Column already exists, which is fine
                db.session.rollback()

        # Now safe to delete reminders
        Reminder.query.filter_by(client_id=client_id).delete()


        db.session.delete(client)

        # Delete the user account if it exists
        if user:
            # Delete password resets
            PasswordReset.query.filter_by(user_id=user.id).delete()

            # Delete session tokens
            SessionToken.query.filter_by(user_id=user.id).delete()

            # Delete audit logs (optional - you might want to keep these)
            # AuditLog.query.filter_by(user_id=user.id).delete()

            # Delete the user
            db.session.delete(user)

        # Commit all deletions
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Client and all associated data deleted successfully'
        })

    except Exception as e:
        db.session.rollback()
        logger.error('delete_client_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': client_id,
                'therapist_id': therapist.id,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500








@app.route('/api/therapist/add-note', methods=['POST'])
@require_auth(['therapist'])
def add_therapist_note():
    """Add note or mission for client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        content = data.get('content')
        is_mission = data.get('is_mission', False)
        note_type = data.get('note_type', 'general')

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Create note
        note = TherapistNote(
            client_id=client_id,
            therapist_id=therapist.id,
            content=sanitize_input(content),
            is_mission=is_mission,
            note_type=note_type
        )
        db.session.add(note)
        db.session.commit()

        return jsonify({
            'success': True,
            'note': {
                'id': note.id,
                'content': note.content,
                'is_mission': note.is_mission,
                'created_at': note.created_at.isoformat()
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500












@app.route('/api/therapist/create-custom-category', methods=['POST'])
@require_auth(['therapist'])
def create_custom_category():
    """Create custom category for a specific client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        name = data.get('name')
        description = data.get('description', '')
        reverse_scoring = data.get('reverse_scoring', False)
        scale_min = data.get('scale_min', 1)
        scale_max = data.get('scale_max', 5)

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Check if custom category with same name exists for this client
        existing = CustomCategory.query.filter_by(
            client_id=client_id,
            name=name,
            is_active=True
        ).first()

        if existing:
            return jsonify({'error': 'Category with this name already exists for this client'}), 400

        # Create custom category
        custom_category = CustomCategory(
            therapist_id=therapist.id,
            client_id=client_id,
            name=name,
            description=description,
            scale_min=scale_min,
            scale_max=scale_max,
            reverse_scoring=reverse_scoring
        )
        db.session.add(custom_category)
        db.session.flush()

        # Add to client's tracking plan
        plan = ClientTrackingPlan(
            client_id=client_id,
            category_id=None,  # NULL for custom categories
            is_active=True
        )
        db.session.add(plan)
        db.session.commit()

        logger.info('custom_category_created', extra={
            'extra_data': {
                'therapist_id': therapist.id,
                'client_id': client_id,
                'category_name': name,
                'custom_category_id': custom_category.id,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        return jsonify({
            'success': True,
            'custom_category': {
                'id': custom_category.id,
                'name': custom_category.name,
                'description': custom_category.description,
                'reverse_scoring': custom_category.reverse_scoring
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error('create_custom_category_error', extra={
            'extra_data': {
                'error': str(e),
                'therapist_id': therapist.id,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500



@app.route('/api/therapist/client/<int:client_id>/add-custom-categories', methods=['POST'])
@require_auth(['therapist'])
def add_custom_categories_to_existing_client(client_id):
    """Add custom categories to an existing client"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Count existing custom categories
        existing_count = CustomCategory.query.filter_by(
            client_id=client_id,
            is_active=True
        ).count()

        if existing_count >= 4:
            return jsonify({'error': 'Client already has maximum custom categories'}), 400

        # Add new custom categories
        custom_categories = data.get('custom_categories', [])
        added_categories = []

        for custom_cat_data in custom_categories[:4 - existing_count]:
            if custom_cat_data.get('name'):
                # Check for duplicate name
                existing = CustomCategory.query.filter_by(
                    client_id=client_id,
                    name=custom_cat_data['name'],
                    is_active=True
                ).first()

                if existing:
                    continue

                custom_category = CustomCategory(
                    therapist_id=therapist.id,
                    client_id=client_id,
                    name=custom_cat_data['name'],
                    description=custom_cat_data.get('description', ''),
                    reverse_scoring=custom_cat_data.get('reverse_scoring', False)
                )
                db.session.add(custom_category)
                added_categories.append(custom_category)

        db.session.commit()

        return jsonify({
            'success': True,
            'added_count': len(added_categories),
            'categories': [{
                'id': cat.id,
                'name': cat.name,
                'description': cat.description
            } for cat in added_categories]
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500




# ============= CLIENT ENDPOINTS =============

@app.route('/api/client/dashboard', methods=['GET'])
@require_auth(['client'])
@cached_endpoint('client_dashboard', ttl=300, key_func=lambda: f"{request.current_user.id}:{get_language_from_header()}")
def client_dashboard():
    """Get client dashboard data with translated category names"""
    try:
        client = request.current_user.client
        lang = get_language_from_header()

        categories_cache_key = f"categories:{client.id}:{lang}"  # NEW LINE
        tracking_categories = cache.get('client_categories', categories_cache_key)

        # Get today's check-in status
        today_checkin = client.checkins.filter_by(checkin_date=date.today()).first()

        # Get active tracking categories with translations
        # Get active tracking categories with translations
        if not tracking_categories:  # NEW LINE
            # Build tracking categories (rest of existing code continues...)  # NEW COMMENT
            tracking_categories = []  # INDENTED

            # Batch load all today's responses for efficiency           # NEW SECTION START
            today_responses = {}
            if client.tracking_plans.count() > 0:
                responses = CategoryResponse.query.filter(
                    CategoryResponse.client_id == client.id,
                    CategoryResponse.response_date == date.today()
                ).all()

                for resp in responses:
                    if resp.category_id:
                        today_responses[resp.category_id] = resp
                    elif resp.custom_category_id:
                        today_responses[f'custom_{resp.custom_category_id}'] = resp
                        # NEW SECTION END












        # First, get all default categories
            for plan in client.tracking_plans.filter_by(is_active=True):
                if plan.category_id:  # Default category
                    # Get today's response
                    today_response = today_responses.get(plan.category_id)

                    # Translate category name and description
                    translated_name = translate_category_name(plan.category.name, lang)

                    tracking_categories.append({
                        'id': plan.category_id,
                        'name': translated_name,
                        'original_name': plan.category.name,
                        'description': translate_category_name(plan.category.name + '_desc', lang),
                        'today_value': today_response.value if today_response else None,
                        'is_custom': False
                    })

            # Then, get all custom categories for this client
            custom_categories = CustomCategory.query.filter_by(
                client_id=client.id,
                is_active=True
            ).all()

            for custom_cat in custom_categories:
                # Get today's response
                try:
                    # Get today's response
                    today_response = today_responses.get(f'custom_{custom_cat.id}')

                    tracking_categories.append({
                        'id': f'custom_{custom_cat.id}',
                        'name': custom_cat.name,
                        'original_name': custom_cat.name,
                        'description': custom_cat.description or '',
                        'today_value': today_response.value if today_response else None,
                        'is_custom': True,
                        'reverse_scoring': custom_cat.reverse_scoring
                    })
                except Exception as e:
                    logger.error(f"Error loading custom category {custom_cat.id}: {e}")
                    continue

            cache.set('client_categories', categories_cache_key, tracking_categories, ttl=1800)

        else:  # <-- ONLY THIS SECTION IS IN THE ELSE
            # Update today's values from fresh data for cached categories
            responses = CategoryResponse.query.filter(
                CategoryResponse.client_id == client.id,
                CategoryResponse.response_date == date.today()
            ).all()

            today_values = {}
            for resp in responses:
                if resp.category_id:
                    today_values[resp.category_id] = resp.value
                elif resp.custom_category_id:
                    today_values[f'custom_{resp.custom_category_id}'] = resp.value

            # Update cached categories with today's values
            for cat in tracking_categories:
                cat_id = cat['id']
                cat['today_value'] = today_values.get(cat_id)


            # Get this week's check-ins count
        week_start = date.today() - timedelta(days=date.today().weekday())
        week_end = week_start + timedelta(days=6)
        week_checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).count()

        # Calculate current streak
        current_streak = 0
        check_date = date.today()

        while True:
            checkin = client.checkins.filter_by(checkin_date=check_date).first()
            if checkin:
                current_streak += 1
                check_date -= timedelta(days=1)
            else:
                break



        weekly_goals = []
        for goal in client.goals.filter_by(
                week_start=week_start,
                is_active=True
        ):
            # Get today's completion
            today_completion = goal.completions.filter_by(
                completion_date=date.today()
            ).first()

            weekly_goals.append({
                'id': goal.id,
                'text': goal.goal_text,
                'today_completed': today_completion.completed if today_completion else None
            })

        # Get reminders
        reminders = []
        for reminder in client.reminders.filter_by(is_active=True):
            reminders.append({
                'type': reminder.reminder_type,
                'time': reminder.reminder_time.strftime('%H:%M')
            })

        return jsonify({
            'success': True,
            'client': {
                'serial': client.client_serial,  # Use client_serial, not serial
                'client_serial': client.client_serial,  # Also provide as client_serial for compatibility
                'client_name': client.client_name if client.client_name else client.client_serial,
                'start_date': client.start_date.isoformat()
            },
            'today': {
                'has_checkin': today_checkin is not None,
                'date': date.today().isoformat()
            },
            'week_checkins_count': week_checkins,
            'current_streak': current_streak,
            'tracking_categories': tracking_categories,
            'weekly_goals': weekly_goals,
            'reminders': reminders
        })




    except Exception as e:
        # Log the actual error
        logger.error('client_dashboard_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': client.id if 'client' in locals() else None,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/consent', methods=['GET'])
@require_auth(['client'])
def get_consent_status():
    """Get client's consent status"""
    try:
        client = request.current_user.client

        consent_types = ['data_sharing', 'treatment', 'communication']
        consent_status = {}

        for consent_type in consent_types:
            record = ConsentRecord.query.filter_by(
                client_id=client.id,
                consent_type=consent_type,
                withdrawal_date=None
            ).order_by(ConsentRecord.consent_date.desc()).first()

            consent_status[consent_type] = {
                'consented': record.consented if record else False,
                'date': record.consent_date.isoformat() if record else None,
                'version': record.consent_version if record else None
            }

        return jsonify({
            'success': True,
            'consents': consent_status
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/consent', methods=['POST'])
@require_auth(['client'])
def update_consent():
    """Update client consent"""
    try:
        client = request.current_user.client
        data = request.json

        consent_type = data.get('consent_type')
        consented = data.get('consented')

        if consent_type not in ['data_sharing', 'treatment', 'communication']:
            return jsonify({'error': 'Invalid consent type'}), 400

        # Create new consent record
        consent = ConsentRecord(
            client_id=client.id,
            consent_type=consent_type,
            consent_version='1.0',  # Update this when consent forms change
            consented=consented,
            ip_address=request.remote_addr
        )
        db.session.add(consent)
        db.session.commit()

        # Audit log
        log_audit(
            action='UPDATE_CONSENT',
            resource_type='consent',
            resource_id=consent.id,
            details={
                'consent_type': consent_type,
                'consented': consented
            }
        )

        return jsonify({
            'success': True,
            'message': 'Consent updated successfully'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500







@app.route('/api/client/checkin', methods=['POST'])
@require_auth(['client'])
def submit_checkin():
    """Submit daily check-in"""
    try:
        client = request.current_user.client
        data = request.json

        checkin_date_str = data.get('date', date.today().isoformat())
        # Validate date string format first
        if not isinstance(checkin_date_str, str) or len(checkin_date_str) != 10:
            return jsonify({'error': 'Invalid date format'}), 400
        checkin_date = checkin_date_str

        logger.info('checkin_data_received', extra={
            'extra_data': {
                'client_id': client.id,
                'date': data.get('date'),
                'category_responses': data.get('category_responses'),
                'has_responses': bool(data.get('category_responses')),
                'response_count': len(data.get('category_responses', {})),
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Validate date format
        try:
            checkin_date = datetime.strptime(checkin_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        # Log the parsed date and validation checks
        logger.info('checkin_validation_details', extra={
            'extra_data': {
                'client_id': client.id,
                'checkin_date': checkin_date.isoformat(),
                'today_date': date.today().isoformat(),
                'is_future': checkin_date > date.today(),
                'client_start_date': client.start_date.isoformat(),
                'is_before_start': checkin_date < client.start_date,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Check if future date
        # Check if future date
        if checkin_date > date.today():
            logger.warning('future_checkin_attempt', extra={
                'extra_data': {
                    'client_id': client.id,
                    'attempted_date': checkin_date.isoformat(),
                    'current_date': date.today().isoformat(),
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })
            return jsonify({'error': 'Cannot submit check-in for future dates'}), 400

        # Check if date is too old (more than 1 year)
        if checkin_date < date.today() - timedelta(days=365):
            return jsonify({'error': 'Cannot submit check-in for dates more than 1 year ago'}), 400

        logger.info('checkin_attempt', extra={
            'extra_data': {
                'client_id': client.id,
                'client_serial': client.client_serial,
                'checkin_date': checkin_date.isoformat(),
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Check if check-in exists
        existing = client.checkins.filter_by(checkin_date=checkin_date).first()
        is_update = existing is not None

        if existing:
            # Update existing check-in
            existing.checkin_time = datetime.now().time()

            # Clear old category responses for this date
            CategoryResponse.query.filter_by(
                client_id=client.id,
                response_date=checkin_date
            ).delete()
        else:
            # Create new check-in
            checkin = DailyCheckin(
                client_id=client.id,
                checkin_date=checkin_date,
                checkin_time=datetime.now().time()
            )
            db.session.add(checkin)
            db.session.flush()
            existing = checkin

        category_notes = data.get('category_notes', {})
        for cat_id, notes in category_notes.items():
            if notes and len(notes) > 500:
                return jsonify({'error': f'Notes too long for category {cat_id}. Maximum 500 characters.'}), 400

        for cat_id, notes in category_notes.items():
            if notes:
                category_notes[cat_id] = sanitize_input(notes)[:500]  # Also enforce length

            # Validate category responses
        category_responses = data.get('category_responses', {})
        validated_responses = {}
        client_category_ids = set()
        for plan in client.tracking_plans.filter_by(is_active=True):
            client_category_ids.add(plan.category_id)

        for cat_id, value in category_responses.items():
            try:
                value_int = int(value)
                if value_int < 0 or value_int > 5:
                    return jsonify({'error': f'Invalid value for category {cat_id}. Must be between 0 and 5.'}), 400

                # Keep custom category IDs as strings
                if isinstance(cat_id, str) and cat_id.startswith('custom_'):
                    validated_responses[cat_id] = value_int
                else:
                    cat_id_int = int(cat_id)
                    if cat_id_int in client_category_ids:
                        validated_responses[cat_id_int] = value_int
            except (ValueError, TypeError):
                return jsonify({'error': f'Invalid value for category {cat_id}. Must be a number.'}), 400

        # Use validated responses
        category_responses = validated_responses

        # Process category responses
        responses_logged = []
        validation_errors = []

        for cat_id, value in category_responses.items():
            try:
                # Check if it's a custom category
                if isinstance(cat_id, str) and cat_id.startswith('custom_'):
                    # Custom category
                    custom_cat_id = int(cat_id.replace('custom_', ''))
                    custom_category = CustomCategory.query.get(custom_cat_id)

                    if not custom_category or custom_category.client_id != client.id:
                        continue

                    # Use the helper method for custom category
                    response = CategoryResponse.create_for_custom_category(
                        client_id=client.id,
                        custom_category_id=custom_cat_id,
                        response_date=checkin_date,
                        value=value,
                        notes=category_notes.get(cat_id, '')
                    )
                    response.save(commit=False)

                    responses_logged.append({
                        'category': custom_category.name,
                        'value': value,
                        'has_notes': bool(category_notes.get(cat_id, '')),
                        'is_custom': True
                    })
                else:
                    # Default category
                    category = TrackingCategory.query.get(int(cat_id))
                    if not category:
                        continue

                    # Use the helper method for standard category
                    response = CategoryResponse.create_for_category(
                        client_id=client.id,
                        category_id=int(cat_id),
                        response_date=checkin_date,
                        value=value,
                        notes=sanitize_input(category_notes.get(str(cat_id), '')[:500])
                    )
                    response.save(commit=False)

                    responses_logged.append({
                        'category': category.name,
                        'value': value,
                        'has_notes': bool(category_notes.get(str(cat_id), '')),
                        'is_custom': False
                    })

                    if 'emotion' in category.name.lower():
                        existing.emotional_value = value
                        existing.emotional_notes = category_notes.get(str(cat_id), '')
                    elif 'medication' in category.name.lower():
                        existing.medication_value = value
                        existing.medication_notes = category_notes.get(str(cat_id), '')
                    elif 'physical activity' in category.name.lower() or 'activity' in category.name.lower():
                        existing.activity_value = value
                        existing.activity_notes = category_notes.get(str(cat_id), '')

            except ValueError as e:
                validation_errors.append(f"Category {cat_id}: {str(e)}")
                logger.error(f"Invalid category response for {cat_id}: {e}")

        if validation_errors:  # Level 1 (OUTSIDE the for loop, same as for)
            db.session.rollback()
            return jsonify({
                'error': 'Invalid category data',
                'details': validation_errors
            }), 400



        # Save goal completions
        goal_completions = data.get('goal_completions', {})
        goals_updated = 0
        for goal_id, completed in goal_completions.items():
            # Check if completion exists
            existing_completion = GoalCompletion.query.filter_by(
                goal_id=int(goal_id),
                completion_date=checkin_date
            ).first()

            if existing_completion:
                existing_completion.completed = completed
            else:
                completion = GoalCompletion(
                    goal_id=int(goal_id),
                    completion_date=checkin_date,
                    completed=completed
                )
                db.session.add(completion)
            goals_updated += 1

        cache.delete('client_dashboard', f"{request.current_user.id}:{get_language_from_header()}")
        cache.invalidate_pattern(f"client_categories:{client.id}:*")


        db.session.commit()

        log_audit(
            action='CREATE_CHECKIN' if not is_update else 'UPDATE_CHECKIN',
            resource_type='checkin',
            resource_id=existing.id,
            details={
                'client_serial': client.client_serial,
                'checkin_date': str(checkin_date),
                'categories_tracked': len(responses_logged)
            },
            phi_accessed=True
        )

        # Send completion email if this is today's checkin
        if checkin_date == date.today() and not is_update:
            try:
                # Get user's language preference
                lang = get_language_from_header()

                # Generate positive message
                positive_messages = {
                    'en': [
                        "Great job completing your daily check-in! Your consistency is admirable.",
                        "Well done! Every check-in brings you closer to your wellness goals.",
                        "Fantastic! Your dedication to tracking your progress is inspiring.",
                        "Excellent work today! Your therapist will be pleased with your commitment.",
                        "Amazing! You're building healthy habits one day at a time."
                    ],
                    'he': [
                        "כל הכבוד על השלמת הצ'ק-אין היומי! העקביות שלך ראויה להערכה.",
                        "עבודה טובה! כל צ'ק-אין מקרב אותך ליעדי הבריאות שלך.",
                        "מצוין! המסירות שלך למעקב אחר ההתקדמות שלך מעוררת השראה.",
                        "עבודה מעולה היום! המטפל שלך ישמח לראות את המחויבות שלך.",
                        "מדהים! אתה בונה הרגלים בריאים יום אחר יום."
                    ],
                    'ru': [
                        "Отличная работа по заполнению ежедневной отметки! Ваша последовательность достойна восхищения.",
                        "Молодец! Каждая отметка приближает вас к вашим целям благополучия.",
                        "Фантастика! Ваша преданность отслеживанию прогресса вдохновляет.",
                        "Отличная работа сегодня! Ваш терапевт будет доволен вашей приверженностью.",
                        "Потрясающе! Вы формируете здоровые привычки день за днем."
                    ],
                    'ar': [
                        "عمل رائع في إكمال تسجيل الحضور اليومي! ثباتك يستحق الإعجاب.",
                        "أحسنت! كل تسجيل حضور يقربك من أهداف عافيتك.",
                        "رائع! تفانيك في تتبع تقدمك ملهم.",
                        "عمل ممتاز اليوم! سيكون معالجك سعيدًا بالتزامك.",
                        "مذهل! أنت تبني عادات صحية يومًا بعد يوم."
                    ]
                }

                import random
                messages = positive_messages.get(lang, positive_messages['en'])
                selected_message = random.choice(messages)

                # Email subject by language
                subjects = {
                    'en': "Well done on today's check-in! 🌟",
                    'he': "כל הכבוד על הצ'ק-אין של היום! 🌟",
                    'ru': "Молодец, вы выполнили сегодняшнюю отметку! 🌟",
                    'ar': "أحسنت في تسجيل حضور اليوم! 🌟"
                }

                subject = subjects.get(lang, subjects['en'])

                # Create email with the positive message
                email_queue = EmailQueue(
                    to_email=request.current_user.email,
                    subject=subject,
                    body=selected_message,
                    html_body=f"""
                            <html>
                            <body style="font-family: Arial, sans-serif; direction: {'rtl' if lang in ['he', 'ar'] else 'ltr'};">
                                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                                    <h2 style="color: #4CAF50; text-align: center;">{subject}</h2>
                                    <p style="font-size: 18px; line-height: 1.6; color: #333;">
                                        {selected_message}
                                    </p>
                                    <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                                    <p style="color: #666; font-size: 14px; text-align: center;">
                                        {'המסע הטיפולי שלך' if lang == 'he' else
                    'Ваш терапевтический путь' if lang == 'ru' else
                    'رحلتك العلاجية' if lang == 'ar' else
                    'Your Therapeutic Journey'}
                                    </p>
                                </div>
                            </body>
                            </html>
                            """,
                    status='pending'
                )
                db.session.add(email_queue)
                db.session.commit()

                # Trigger email processing
                if celery:
                    from celery_app import process_email_queue_task
                    process_email_queue_task.delay()

            except Exception as e:
                logger.error(f"Failed to queue completion email: {e}")

        logger.info('checkin_success', extra={
            'extra_data': {
                'client_id': client.id,
                'client_serial': client.client_serial,
                'checkin_date': checkin_date.isoformat(),
                'is_update': is_update,
                'categories_tracked': len(responses_logged),
                'goals_updated': goals_updated,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        return jsonify({
            'success': True,
            'message': 'Check-in saved successfully'
        })

    except Exception as e:
        db.session.rollback()
        logger.error('checkin_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': client.id,
                'checkin_date': checkin_date.isoformat() if checkin_date else None,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/progress', methods=['GET'])
@require_auth(['client'])
def get_client_progress():
    """Get client's progress data with pagination for large datasets"""
    try:
        client = request.current_user.client

        # Get date range and pagination
        end_date = date.today()
        days = int(request.args.get('days', 30))
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        # Validate parameters
        page = max(1, page)
        per_page = min(max(1, per_page), 100)
        days = min(max(1, days), 365)  # Max 1 year

        start_date = end_date - timedelta(days=days)

        # Log request
        logger.info('get_progress_request', extra={
            'extra_data': {
                'client_id': client.id,
                'days': days,
                'page': page,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Get paginated check-ins
        checkins_query = client.checkins.filter(
            DailyCheckin.checkin_date.between(start_date, end_date)
        ).order_by(DailyCheckin.checkin_date.desc())

        pagination = checkins_query.paginate(page=page, per_page=per_page, error_out=False)
        checkins = pagination.items if pagination else []

        checkin_data = []
        for checkin in checkins:
            # Ensure all values are present and valid
            checkin_data.append({
                'date': checkin.checkin_date.isoformat(),
                'time': checkin.created_at.strftime('%H:%M') if checkin.created_at else '00:00',
                'emotional': checkin.emotional_value if checkin.emotional_value is not None else 0,
                'medication': checkin.medication_value if checkin.medication_value is not None else 0,
                'activity': checkin.activity_value if checkin.activity_value is not None else 0
            })

        # Get category responses (also paginated if needed)
        category_data = {}
        try:
            for plan in client.tracking_plans.filter_by(is_active=True):
                responses = CategoryResponse.query.filter(
                    CategoryResponse.client_id == client.id,
                    CategoryResponse.category_id == plan.category_id,
                    CategoryResponse.response_date.between(start_date, end_date)
                ).order_by(CategoryResponse.response_date.desc()).limit(per_page).all()

                category_data[plan.category.name] = [
                    {
                        'date': resp.response_date.isoformat(),
                        'value': resp.value if resp.value is not None else 0
                    } for resp in responses
                ]
        except Exception as cat_error:
            # Log but don't fail the whole request
            logger.warning('get_progress_category_error', extra={
                'extra_data': {
                    'error': str(cat_error),
                    'client_id': client.id,
                    'request_id': g.request_id
                }
            })
            category_data = {}

        # Log success
        logger.info('get_progress_success', extra={
            'extra_data': {
                'client_id': client.id,
                'checkin_count': len(checkin_data),
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        return jsonify({
            'success': True,
            'progress': {
                'checkins': checkin_data,
                'categories': category_data
            },
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': pagination.total if pagination else 0,
                'pages': pagination.pages if pagination else 1,
                'has_prev': pagination.has_prev if pagination else False,
                'has_next': pagination.has_next if pagination else False
            }
        })

    except Exception as e:
        logger.error('get_progress_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': getattr(request.current_user, 'client', {}).get('id', None) if hasattr(request,
                                                                                                    'current_user') else None,
                'request_id': g.request_id if hasattr(g, 'request_id') else None
            },
            'request_id': g.request_id if hasattr(g, 'request_id') else None,
            'user_id': request.current_user.id if hasattr(request, 'current_user') else None
        }, exc_info=True)

        # Return valid empty structure with 200 status to prevent frontend errors
        return jsonify({
            'success': False,
            'error': 'Unable to load progress data',
            'progress': {
                'checkins': [],
                'categories': {}
            },
            'pagination': {
                'page': 1,
                'per_page': 50,
                'total': 0,
                'pages': 1,
                'has_prev': False,
                'has_next': False
            }
        }), 200


@app.route('/api/client/change-password', methods=['POST'])
@require_auth(['client'])
def change_client_password():
    """Allow client to change their own password"""
    try:
        data = request.json
        current_password = data.get('current_password')
        new_password = data.get('new_password')

        # Validate input
        if not all([current_password, new_password]):
            return jsonify({'error': 'Current password and new password are required'}), 400

        if len(new_password) < 8:
            return jsonify({'error': 'New password must be at least 8 characters long'}), 400

        # Verify current password
        user = request.current_user
        if not bcrypt.check_password_hash(user.password_hash, current_password):
            return jsonify({'error': 'Current password is incorrect'}), 401

        # Update password
        user.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Password changed successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/missions', methods=['GET'])
@require_auth(['client'])
def get_client_missions():
    """Get client's missions from therapist"""
    try:
        client = request.current_user.client

        missions = TherapistNote.query.filter_by(
            client_id=client.id,
            is_mission=True
        ).order_by(TherapistNote.created_at.desc()).all()

        mission_data = []
        for mission in missions:
            mission_data.append({
                'id': mission.id,
                'content': mission.content,
                'completed': mission.mission_completed,
                'created_at': mission.created_at.isoformat(),
                'completed_at': mission.completed_at.isoformat() if mission.completed_at else None
            })

        return jsonify({
            'success': True,
            'missions': mission_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/complete-mission/<int:mission_id>', methods=['POST'])
@require_auth(['client'])
def complete_mission(mission_id):
    """Mark mission as completed"""
    try:
        client = request.current_user.client

        mission = TherapistNote.query.filter_by(
            id=mission_id,
            client_id=client.id,
            is_mission=True
        ).first()

        if not mission:
            return jsonify({'error': 'Mission not found'}), 404

        mission.mission_completed = True
        mission.completed_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Mission completed!'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/save-goals', methods=['POST'])
@require_auth(['client'])
def save_goals():
    """Save goals for a specific week"""
    try:
        client = request.current_user.client
        data = request.json

        week_start_str = data.get('week_start')
        goals = data.get('goals', '').strip()

        if not week_start_str:
            return jsonify({'error': 'Week start date is required'}), 400

        if not goals:
            return jsonify({'error': 'Goals are required'}), 400

        # Validate word count
        word_count = len(goals.split())
        if word_count > 500:
            return jsonify({'error': 'Goals must be 500 words or less'}), 400

        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()

        # Check if goals exist for this week
        existing = TherapistNote.query.filter_by(
            client_id=client.id,
            note_type='weekly_goals',
            created_at=week_start
        ).first()

        if existing:
            existing.content = goals
            existing.updated_at = datetime.utcnow()
        else:
            note = TherapistNote(
                client_id=client.id,
                therapist_id=client.therapist_id,
                note_type='weekly_goals',
                content=goals,
                is_mission=False,
                created_at=week_start  # Use week_start as created_at for easy retrieval
            )
            db.session.add(note)

        db.session.commit()

        return jsonify({'success': True, 'message': 'Goals saved successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/get-goals', methods=['GET'])
@require_auth(['client'])
def get_goals():
    """Get goals for a specific week"""
    try:
        client = request.current_user.client
        week_start_str = request.args.get('week_start')

        if not week_start_str:
            return jsonify({'error': 'Week start date is required'}), 400

        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()

        goals = TherapistNote.query.filter_by(
            client_id=client.id,
            note_type='weekly_goals',
            created_at=week_start
        ).first()

        return jsonify({
            'goals': goals.content if goals else None
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/save-brief-goals', methods=['POST'])
@require_auth(['client'])
def save_brief_goals():
    """Save brief goals for the week"""
    try:
        client = request.current_user.client
        data = request.json

        week_start_str = data.get('week_start')
        brief_goals = data.get('brief_goals', '').strip()

        if not week_start_str:
            return jsonify({'error': 'Week start date is required'}), 400

        if not brief_goals:
            return jsonify({'error': 'Brief goals are required'}), 400

        # Validate word count
        word_count = len(brief_goals.split())
        if word_count > 20:
            return jsonify({'error': 'Brief goals must be 20 words or less'}), 400

        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()

        # Check if brief goals exist for this week
        existing = TherapistNote.query.filter_by(
            client_id=client.id,
            note_type='brief_goals',
            created_at=week_start
        ).first()

        if existing:
            existing.content = sanitize_input(brief_goals)
            existing.updated_at = datetime.utcnow()
        else:
            note = TherapistNote(
                client_id=client.id,
                therapist_id=client.therapist_id,
                note_type='brief_goals',
                content=brief_goals,
                is_mission=False,
                created_at=week_start
            )
            db.session.add(note)

        db.session.commit()

        return jsonify({'success': True, 'message': 'Brief goals saved successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/get-brief-goals', methods=['GET'])
@require_auth(['client'])
def get_brief_goals():
    """Get brief goals for a specific week"""
    try:
        client = request.current_user.client
        week_start_str = request.args.get('week_start')

        if not week_start_str:
            return jsonify({'error': 'Week start date is required'}), 400

        week_start = datetime.strptime(week_start_str, '%Y-%m-%d').date()

        brief_goals = TherapistNote.query.filter_by(
            client_id=client.id,
            note_type='brief_goals',
            created_at=week_start
        ).first()

        return jsonify({
            'brief_goals': brief_goals.content if brief_goals else None
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/log-error', methods=['POST'])
@require_auth(['client'])  # Use your auth decorator
def log_client_error():
    try:
        data = request.json
        errors = data.get('errors', [])

        for error in errors:
            logger.error('client_frontend_error', extra={
                'extra_data': {
                    'client_id': request.current_user.client.id,
                    'error': error,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })

        return jsonify({'success': True})
    except Exception as e:
        logger.error('log_client_error_failed', extra={
            'extra_data': {
                'error': str(e),
                'request_id': g.request_id
            },
            'request_id': g.request_id
        })
        return jsonify({'success': False}), 200



# ============= TRACKING CATEGORY MANAGEMENT =============

@app.route('/api/categories', methods=['GET'])
@require_auth(['therapist', 'client'])
@cache_response(timeout=3600)
def get_tracking_categories():
    """Get all tracking categories with translations"""
    try:
        lang = get_language_from_header()
        categories = TrackingCategory.query.all()

        category_data = []
        for category in categories:
            translated_name = translate_category_name(category.name, lang)
            category_data.append({
                'id': category.id,
                'name': translated_name,
                'original_name': category.name,  # Keep original for reference
                'description': category.description,
                'is_default': category.is_default,
                'scale_min': category.scale_min,
                'scale_max': category.scale_max
            })

        return jsonify({
            'success': True,
            'categories': category_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# NEW ENDPOINT - Added after /api/categories
@app.route('/api/categories/translated', methods=['GET'])
@require_auth(['therapist', 'client'])
@cache_response(timeout=3600)
def get_translated_categories():
    """Get all tracking categories with translations for current language"""
    try:
        lang = get_language_from_header()
        categories = TrackingCategory.query.all()

        category_data = []
        for category in categories:
            translated_name = translate_category_name(category.name, lang)
            # Also translate the description
            desc_key = category.name + '_desc'
            translated_desc = CATEGORY_TRANSLATIONS.get(lang, {}).get(
                desc_key,
                category.description
            )

            category_data.append({
                'id': category.id,
                'name': translated_name,
                'original_name': category.name,
                'description': translated_desc,
                'is_default': category.is_default,
                'scale_min': category.scale_min,
                'scale_max': category.scale_max
            })

        return jsonify({
            'success': True,
            'categories': category_data,
            'language': lang
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/api/therapist/client/<int:client_id>/custom-categories', methods=['GET'])
@require_auth(['therapist'])
def get_client_custom_categories(client_id):
    """Get all custom categories for a specific client"""
    try:
        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Get custom categories
        custom_categories = CustomCategory.query.filter_by(
            client_id=client_id,
            is_active=True
        ).all()

        category_data = []
        for cat in custom_categories:
            category_data.append({
                'id': cat.id,
                'name': cat.name,
                'description': cat.description,
                'reverse_scoring': cat.reverse_scoring,
                'scale_min': cat.scale_min,
                'scale_max': cat.scale_max,
                'created_at': cat.created_at.isoformat()
            })

        return jsonify({
            'success': True,
            'custom_categories': category_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500





@app.route('/api/therapist/update-tracking-plan', methods=['POST'])
@require_auth(['therapist'])
def update_client_tracking_plan():
    """Update client's tracking plan"""
    try:
        therapist = request.current_user.therapist
        data = request.json

        client_id = data.get('client_id')
        category_updates = data.get('categories', {})

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Update tracking plans
        for category_id, is_active in category_updates.items():
            plan = ClientTrackingPlan.query.filter_by(
                client_id=client_id,
                category_id=int(category_id)
            ).first()

            if plan:
                plan.is_active = is_active
            else:
                # Create new plan if doesn't exist
                plan = ClientTrackingPlan(
                    client_id=client_id,
                    category_id=int(category_id),
                    is_active=is_active
                )
                db.session.add(plan)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Tracking plan updated successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============= REMINDER ENDPOINTS =============

@app.route('/api/client/reminders', methods=['GET'])
@require_auth(['client'])
def get_reminders():
    """Get client's reminders"""
    try:
        client = request.current_user.client
        reminders = client.reminders.filter_by(is_active=True).all()

        reminder_data = []
        for reminder in reminders:
            # IMPORTANT: Always use local_reminder_time if available
            if reminder.local_reminder_time:
                display_time = reminder.local_reminder_time
            else:
                # If no local time stored, we need to convert UTC back to local
                # This is a fallback - ideally local_reminder_time should always be set
                display_time = reminder.reminder_time.strftime('%H:%M')

            reminder_data.append({
                'id': reminder.id,
                'type': reminder.reminder_type,
                'time': display_time,
                'utc_time': reminder.reminder_time.strftime('%H:%M'),  # Add this for debugging
                'email': reminder.reminder_email,
                'is_active': reminder.is_active,
                'last_sent': reminder.last_sent.isoformat() if reminder.last_sent else None
            })

        return jsonify({
            'success': True,
            'reminders': reminder_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/update-reminder', methods=['GET'])
@require_auth(['client'])
def update_reminder_get():
    """Handle GET requests to update-reminder endpoint"""
    return jsonify({
        'error': 'This endpoint only accepts POST requests',
        'method': 'GET not allowed'
    }), 405


@app.route('/api/client/update-reminder', methods=['POST'])
@require_auth(['client'])
def update_reminder():
    """Update or create reminder"""
    try:
        client = request.current_user.client
        data = request.json

        # Use structured logger for visibility
        logger.info('update_reminder_request', extra={
            'extra_data': {
                'client_id': client.id,
                'client_serial': client.client_serial,
                'request_data': data,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        reminder_type = data.get('type')
        reminder_time = data.get('time')
        reminder_email = data.get('email')
        is_active = data.get('is_active', True)
        timezone_offset = data.get('timezone_offset', 0)# in minutes
        reminder_language = data.get('language', 'en')

        # Parse time
        hour, minute = map(int, reminder_time.split(':'))

        local_time_str = f"{hour:02d}:{minute:02d}"

        # Log the timezone conversion details
        if abs(timezone_offset) > 840:  # Max UTC+14 or UTC-14
            return jsonify({'error': 'Invalid timezone offset'}), 400

        # Log without exposing exact timezone
        logger.info('timezone_conversion_start', extra={
            'extra_data': {
                'local_time': f"{hour:02d}:{minute:02d}",
                'timezone_offset': 'provided',  # Don't log exact offset
                'client_serial': client.client_serial,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # FIX: The hour/minute we receive is in the user's local time
        # We need to convert it to UTC for storage

        # JavaScript's getTimezoneOffset() returns minutes to SUBTRACT from local time to get UTC
        # For PDT (UTC-7), it returns 420
        # For Jerusalem (UTC+3), it returns -180
        # So to convert local time to UTC, we need to SUBTRACT the offset

        # Create a UTC datetime for today with the local time values
        # This ensures we're working with UTC regardless of server timezone
        logger.info('timezone_conversion_input', extra={
            'extra_data': {
                'local_time': f"{hour:02d}:{minute:02d}",
                'timezone_offset': timezone_offset,
                'offset_sign': 'positive (west of UTC)' if timezone_offset > 0 else 'negative (east of UTC)',
                'client_serial': client.client_serial,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # The correct formula is: UTC = LOCAL - offset_minutes
        # For PDT (offset=420): 7:00 AM - 420 minutes = 0:00 AM UTC (previous day, so +24 hours = midnight UTC)
        # For Jerusalem (offset=-180): 9:00 AM - (-180 minutes) = 9:00 AM + 180 minutes = 12:00 PM UTC

        # Convert local time to total minutes
        local_total_minutes = hour * 60 + minute

        # Subtract the offset to get UTC
        utc_total_minutes = local_total_minutes - timezone_offset

        # Handle day wraparound
        while utc_total_minutes < 0:
            utc_total_minutes += 24 * 60
        while utc_total_minutes >= 24 * 60:
            utc_total_minutes -= 24 * 60

        # Convert back to hours and minutes
        utc_hour = utc_total_minutes // 60
        utc_minute = utc_total_minutes % 60

        # Create the time object - use datetime.time to avoid any conflicts
        import datetime
        time_obj = datetime.time(utc_hour, utc_minute)

        # Log the conversion result for debugging - FIX THE CALCULATION STRING
        logger.info('timezone_conversion_output', extra={
            'extra_data': {
                'input_local': f"{hour:02d}:{minute:02d}",
                'input_offset': timezone_offset,
                'output_utc': f"{utc_hour:02d}:{utc_minute:02d}",
                'calculation': f"{local_total_minutes} - {timezone_offset} = {utc_total_minutes} minutes = {utc_hour:02d}:{utc_minute:02d} UTC",
                'client_serial': client.client_serial,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Log the conversion result - FIX THE CALCULATION STRING
        logger.info('timezone_conversion_result', extra={
            'extra_data': {
                'input_local_time': f"{hour:02d}:{minute:02d}",
                'output_utc_time': str(time_obj),
                'utc_hour': time_obj.hour,
                'utc_minute': time_obj.minute,
                'timezone_offset_minutes': timezone_offset,
                'expected_offset_sign': 'negative' if timezone_offset < 0 else 'positive',
                'calculation': f"({hour}*60+{minute})-{timezone_offset} = {utc_hour}:{utc_minute} UTC",
                'client_serial': client.client_serial,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Check if reminder exists
        reminder = client.reminders.filter_by(reminder_type=reminder_type).first()

        if reminder:
            old_time = reminder.reminder_time
            reminder.reminder_time = time_obj
            reminder.local_reminder_time = local_time_str
            reminder.is_active = is_active
            reminder.reminder_email = reminder_email
            reminder.reminder_language = reminder_language

            logger.info('reminder_updated', extra={
                'extra_data': {
                    'client_serial': client.client_serial,
                    'old_time': str(old_time),
                    'new_time': str(time_obj),
                    'email': reminder_email or 'using account email',
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })
        else:
            reminder = Reminder(
                client_id=client.id,
                reminder_type=reminder_type,
                reminder_time=time_obj,
                local_reminder_time=local_time_str,
                reminder_email=reminder_email,
                reminder_language=reminder_language,
                is_active=is_active
            )
            db.session.add(reminder)

            logger.info('reminder_created', extra={
                'extra_data': {
                    'client_serial': client.client_serial,
                    'time': str(time_obj),
                    'email': reminder_email or 'using account email',
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Reminder updated successfully',
            'debug': {
                'local_time': f"{hour:02d}:{minute:02d}",
                'utc_time': str(time_obj),
                'timezone_offset': timezone_offset
            }
        })

    except Exception as e:
        db.session.rollback()
        logger.error('update_reminder_error', extra={
            'extra_data': {
                'error': str(e),
                'client_id': client.id if client else None,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/debug/reminder-times', methods=['GET'])
@require_auth(['therapist'])
def debug_reminder_times():
    """Debug endpoint to check all reminder times and timezone handling"""
    try:
        from datetime import datetime

        reminders = db.session.query(
            Reminder.id,
            Reminder.reminder_type,
            Reminder.reminder_time,
            Reminder.is_active,
            Reminder.last_sent,
            Client.client_serial,
            User.email
        ).join(
            Client, Reminder.client_id == Client.id
        ).join(
            User, Client.user_id == User.id
        ).filter(
            Reminder.reminder_type == 'daily_checkin',
            Reminder.is_active == True
        ).all()

        current_utc = datetime.utcnow()
        current_utc_hour = current_utc.hour

        reminder_data = []
        hour_distribution = {}

        for r in reminders:
            reminder_hour = r.reminder_time.hour if r.reminder_time else None

            # Count hour distribution
            if reminder_hour is not None:
                hour_distribution[reminder_hour] = hour_distribution.get(reminder_hour, 0) + 1

            # Check if should send this hour
            should_send_now = (
                    r.is_active and
                    reminder_hour == current_utc_hour
            )

            # Skip test emails in count
            is_test_email = (
                    r.email.endswith('example.com') or
                    r.email.endswith('test.test')
            )

            reminder_data.append({
                'id': r.id,
                'client_serial': r.client_serial,
                'email': r.email,
                'is_test_email': is_test_email,
                'reminder_time': r.reminder_time.strftime('%H:%M') if r.reminder_time else None,
                'reminder_hour': reminder_hour,
                'is_active': r.is_active,
                'last_sent': r.last_sent.isoformat() if r.last_sent else None,
                'should_send_now': should_send_now
            })

        # Sort hour distribution
        hour_dist_sorted = dict(sorted(hour_distribution.items()))

        # Count reminders that should send this hour (excluding test emails)
        should_send_count = sum(1 for r in reminder_data
                                if r['should_send_now'] and not r['is_test_email'])

        return jsonify({
            'current_utc_time': current_utc.isoformat(),
            'current_utc_hour': current_utc_hour,
            'total_active_reminders': len(reminder_data),
            'should_send_this_hour': should_send_count,
            'hour_distribution': hour_dist_sorted,
            'reminders': reminder_data[:20]  # First 20 for review
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/debug/reminders', methods=['GET'])
@require_auth(['therapist', 'client'])  # Allow both roles for debugging
def debug_reminders():
    """Debug endpoint to check reminder status"""
    try:
        from datetime import datetime, timedelta

        # Get all reminders with client info
        reminders = db.session.query(
            Reminder.id,
            Reminder.reminder_type,
            Reminder.reminder_time,
            Reminder.is_active,
            Reminder.last_sent,
            Client.client_serial,
            User.email
        ).join(
            Client, Reminder.client_id == Client.id
        ).join(
            User, Client.user_id == User.id
        ).all()

        current_hour = datetime.now().hour
        current_time = datetime.now()
        reminder_data = []

        for r in reminders:
            reminder_hour = r.reminder_time.hour if r.reminder_time else None
            should_send_this_hour = (
                    r.is_active and
                    r.reminder_type == 'daily_checkin' and
                    reminder_hour == current_hour
            )

            time_since_sent = None
            sent_recently = False
            if r.last_sent:
                delta = datetime.utcnow() - r.last_sent
                hours = delta.total_seconds() / 3600
                minutes = delta.total_seconds() / 60

                if minutes < 60:
                    time_since_sent = f"{int(minutes)} minutes ago"
                    sent_recently = minutes < 5  # Sent in last 5 minutes
                elif hours < 24:
                    time_since_sent = f"{int(hours)} hours ago"
                else:
                    time_since_sent = f"{int(hours / 24)} days ago"

            reminder_data.append({
                'id': r.id,
                'client_serial': r.client_serial,
                'email': r.email,
                'type': r.reminder_type,
                'time': r.reminder_time.strftime('%H:%M') if r.reminder_time else None,
                'is_active': r.is_active,
                'last_sent': r.last_sent.isoformat() if r.last_sent else None,
                'time_since_sent': time_since_sent,
                'sent_recently': sent_recently,
                'should_send_this_hour': should_send_this_hour,
                'reminder_hour': reminder_hour,
                'current_hour': current_hour
            })

        # Get reminders sent in last hour
        one_hour_ago = datetime.utcnow() - timedelta(hours=1)
        recent_sends = [r for r in reminder_data if r['last_sent'] and
                        datetime.fromisoformat(r['last_sent'].replace('Z', '+00:00')) > one_hour_ago]

        # Check Celery status
        celery_status = "Unknown"
        active_tasks = 0

        if celery:
            try:
                inspector = celery.control.inspect()
                stats = inspector.stats()
                if stats:
                    celery_status = "Connected"
                    active = inspector.active()
                    if active:
                        active_tasks = sum(len(tasks) for tasks in active.values())
                else:
                    celery_status = "No workers found"
            except:
                celery_status = "Connection failed"
        else:
            celery_status = "Not configured"

        return jsonify({
            'success': True,
            'current_time': current_time.isoformat(),
            'current_hour': current_hour,
            'celery_status': celery_status,
            'active_tasks': active_tasks,
            'total_reminders': len(reminder_data),
            'active_reminders': sum(1 for r in reminder_data if r['is_active']),
            'sent_in_last_hour': len(recent_sends),
            'should_send_this_hour': sum(1 for r in reminder_data if r['should_send_this_hour']),
            'reminders': reminder_data,
            'recent_sends': recent_sends,
            'email_configured': bool(app.config.get('MAIL_USERNAME')),
            'smtp_server': app.config.get('MAIL_SERVER', 'Not configured')
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/debug/server-time', methods=['GET'])
def debug_server_time():
    """Debug endpoint to check server time and timezone"""
    import time
    from datetime import datetime

    return jsonify({
        'server_local_time': datetime.now().isoformat(),
        'server_utc_time': datetime.utcnow().isoformat(),
        'timezone_name': time.tzname,
        'timezone_offset_seconds': time.timezone,
        'dst_offset_seconds': time.altzone,
        'is_dst': time.daylight,
        'python_thinks_offset_minutes': int(time.timezone / -60),  # Python's offset convention
        'datetime_now': str(datetime.now()),
        'datetime_utcnow': str(datetime.utcnow()),
        'difference_minutes': int((datetime.now() - datetime.utcnow()).total_seconds() / 60)
    })


@app.route('/api/debug/reminder-timezone-test', methods=['GET'])
@require_auth(['therapist', 'client'])
def debug_reminder_timezone():
    """Debug timezone handling for reminders"""
    try:
        from datetime import datetime

        current_utc = datetime.utcnow()
        current_local = datetime.now()

        # Get all reminders
        reminders = Reminder.query.filter_by(
            is_active=True,
            reminder_type='daily_checkin'
        ).all()

        reminder_data = []
        for r in reminders:
            reminder_data.append({
                'client_serial': r.client.client_serial,
                'stored_time': r.reminder_time.strftime('%H:%M'),
                'stored_hour': r.reminder_time.hour,
                'email_to_use': r.reminder_email if r.reminder_email else r.client.user.email,
                'matches_current_utc_hour': r.reminder_time.hour == current_utc.hour
            })

        return jsonify({
            'current_utc_time': current_utc.strftime('%Y-%m-%d %H:%M:%S'),
            'current_utc_hour': current_utc.hour,
            'current_local_time': current_local.strftime('%Y-%m-%d %H:%M:%S'),
            'local_offset_minutes': int((current_local - current_utc).total_seconds() / 60),
            'reminders': reminder_data,
            'reminders_for_current_hour': [r for r in reminder_data if r['matches_current_utc_hour']]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500





@app.route('/api/test/send-reminder-now', methods=['POST'])
@require_auth(['therapist', 'client'])
def test_send_reminder_now():
    """Force send a reminder email right now for testing"""
    try:
        # Get current user's email
        email = request.current_user.email

        # Send test reminder
        subject = "TEST: Daily Check-in Reminder"
        body = f"""
This is a TEST reminder email sent at {datetime.now().isoformat()}.

If you're receiving this, the email system is working correctly!

Your actual reminders will come at your scheduled time.

Best regards,
Therapeutic Companion Team
        """

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #e91e63;">TEST Email</h2>
            <p>This is a TEST reminder email sent at {datetime.now().isoformat()}.</p>
            <p>If you're receiving this, the email system is working correctly!</p>
            <p>Your actual reminders will come at your scheduled time.</p>
            <hr>
            <p style="color: #666;">Best regards,<br>Therapeutic Companion Team</p>
        </body>
        </html>
        """

        # Try to send directly
        success = send_email(email, subject, body, html_body)

        # Also try through Celery if available
        celery_task_id = None
        if celery:
            try:
                from celery_app import send_reminder_test
                task = send_reminder_test.delay(email)
                celery_task_id = task.id
            except:
                pass

        return jsonify({
            'success': True,
            'direct_send': success,
            'celery_task_id': celery_task_id,
            'email': email,
            'timestamp': datetime.now().isoformat(),
            'email_configured': bool(app.config.get('MAIL_USERNAME'))
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ============= PROFILE MANAGEMENT =============

@app.route('/api/user/profile', methods=['GET'])
@require_auth(['therapist', 'client'])
def get_user_profile():
    """Get user profile information"""
    try:
        user = request.current_user

        profile_data = {
            'email': user.email,
            'role': user.role,
            'created_at': user.created_at.isoformat(),
            'last_login': user.last_login.isoformat() if user.last_login else None
        }

        if user.role == 'therapist' and user.therapist:
            profile_data['therapist'] = {
                'name': user.therapist.name,
                'license_number': user.therapist.license_number,
                'organization': user.therapist.organization,
                'specializations': user.therapist.specializations or []
            }
        elif user.role == 'client' and user.client:
            profile_data['client'] = {
                'serial': user.client.client_serial,
                'start_date': user.client.start_date.isoformat(),
                'therapist_name': user.client.therapist.name if user.client.therapist else None
            }

        return jsonify({
            'success': True,
            'profile': profile_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/user/update-profile', methods=['POST'])
@require_auth(['therapist', 'client'])
def update_user_profile():
    """Update user profile"""
    try:
        user = request.current_user
        data = request.json

        # Update email if provided
        new_email = data.get('email')
        if new_email and new_email != user.email:
            # Check if email already exists
            if User.query.filter_by(email=new_email).first():
                return jsonify({'error': 'Email already in use'}), 400
            user.email = new_email

        # Update role-specific data
        if user.role == 'therapist' and user.therapist:
            if 'name' in data:
                user.therapist.name = data['name']
            if 'organization' in data:
                user.therapist.organization = data['organization']
            if 'specializations' in data:
                user.therapist.specializations = data['specializations']

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Profile updated successfully'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============= ANALYTICS ENDPOINTS =============


# ============= EXPORT ENDPOINTS =============

@app.route('/api/export/client-data/<int:client_id>', methods=['GET'])
@require_auth(['therapist'])
def export_client_data(client_id):
    """Export all client data as JSON"""
    try:
        therapist = request.current_user.therapist

        # Verify client belongs to therapist
        client = Client.query.filter_by(
            id=client_id,
            therapist_id=therapist.id
        ).first()

        if not client:
            return jsonify({'error': 'Client not found'}), 404

        # Compile all client data
        export_data = {
            'client': {
                 'serial': client.client_serial,
        'client_name': client.client_name if client.client_name else client.client_serial,
        'start_date': client.start_date.isoformat(),
        'is_active': client.is_active,
        'created_at': client.created_at.isoformat()
            },
            'checkins': [],
            'category_responses': [],
            'goals': [],
            'notes': [],
            'reminders': []
        }

        # Get all check-ins
        for checkin in client.checkins.order_by(DailyCheckin.checkin_date):
            export_data['checkins'].append({
                'date': checkin.checkin_date.isoformat(),
                'time': checkin.checkin_time.strftime('%H:%M'),
                'emotional': checkin.emotional_value,
                'emotional_notes': checkin.emotional_notes,
                'medication': checkin.medication_value,
                'medication_notes': checkin.medication_notes,
                'activity': checkin.activity_value,
                'activity_notes': checkin.activity_notes
            })

        # Get all category responses
        responses = CategoryResponse.query.filter_by(client_id=client_id).order_by(CategoryResponse.response_date).all()
        for response in responses:
            export_data['category_responses'].append({
                'date': response.response_date.isoformat(),
                'category': response.category.name,
                'value': response.value,
                'notes': response.notes
            })

        # Get all goals
        for goal in client.goals.order_by(WeeklyGoal.week_start):
            goal_data = {
                'text': goal.goal_text,
                'week_start': goal.week_start.isoformat(),
                'is_active': goal.is_active,
                'completions': []
            }
            for completion in goal.completions.order_by(GoalCompletion.completion_date):
                goal_data['completions'].append({
                    'date': completion.completion_date.isoformat(),
                    'completed': completion.completed,
                    'notes': completion.notes
                })
            export_data['goals'].append(goal_data)

        # Get all therapist notes
        notes = TherapistNote.query.filter_by(
            client_id=client_id,
            therapist_id=therapist.id
        ).order_by(TherapistNote.created_at).all()

        for note in notes:
            export_data['notes'].append({
                'type': note.note_type,
                'content': note.content,
                'is_mission': note.is_mission,
                'completed': note.mission_completed,
                'created_at': note.created_at.isoformat(),
                'completed_at': note.completed_at.isoformat() if note.completed_at else None
            })

        # Get reminders
        for reminder in client.reminders.all():
            export_data['reminders'].append({
                'type': reminder.reminder_type,
                'time': reminder.reminder_time.strftime('%H:%M'),
                'is_active': reminder.is_active
            })

        return jsonify(export_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= ADMIN ENDPOINTS (if needed) =============


@app.route('/api/debug/categories', methods=['GET'])
@require_auth(['therapist'])
def debug_categories():
    """Debug endpoint to check category translations"""
    categories = TrackingCategory.query.all()
    lang = get_language_from_header()

    results = []
    for cat in categories:
        translation_exists = cat.name in CATEGORY_TRANSLATIONS.get(lang, {})
        translated_name = translate_category_name(cat.name, lang)

        results.append({
            'db_name': cat.name,
            'translation_exists': translation_exists,
            'translated_to': translated_name,
            'is_same': cat.name == translated_name
        })

    return jsonify({
        'language': lang,
        'categories': results
    })


# ============= CLIENT REPORT ENDPOINTS =============


@app.route('/api/client/email-report', methods=['POST'])
@require_auth(['client'])
def client_email_report():
    """Prepare email report for client to send to therapist with translations"""
    try:
        client = request.current_user.client
        lang = get_language_from_header()
        data = request.json
        week = data.get('week')

        if not week:
            return jsonify({'error': 'Week is required'}), 400

        # Parse week
        # Validate week format
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get therapist info
        therapist = client.therapist
        therapist_email = therapist.user.email if therapist and therapist.user else "therapist@example.com"
        therapist_name = therapist.name if therapist else "Therapist"

        # Get check-ins for the week
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).order_by(DailyCheckin.checkin_date).all()

        # Get translations
        trans = lambda key: translate_report_term(key, lang)
        days = DAYS_TRANSLATIONS.get(lang, DAYS_TRANSLATIONS['en'])

        # Build translated email content
        if lang == 'he':
            subject = f"דוח טיפולי שבועי - {sanitize_input(client.client_serial)} - שבוע {week_num}, {year}"
            content = f"""מטפל יקר,

הנה דוח ההתקדמות השבועי שלי עבור {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}.

מטופל: {sanitize_input(client.client_serial)}
שבוע: {week_num}, {year}
צ׳ק-אינים שהושלמו: {len(checkins)}/7

סיכום צ׳ק-אין יומי:
"""
        elif lang == 'ru':
            subject = f"Еженедельный терапевтический отчет - {sanitize_input(client.client_serial)} - Неделя {week_num}, {year}"
            content = f"""Уважаемый терапевт,

Вот мой еженедельный отчет о прогрессе за {week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}.

КЛИЕНТ: {sanitize_input(client.client_serial)}
НЕДЕЛЯ: {week_num}, {year}
ВЫПОЛНЕНО ОТМЕТОК: {len(checkins)}/7

ЕЖЕДНЕВНАЯ СВОДКА ОТМЕТОК:
"""
        elif lang == 'ar':
            subject = f"التقرير العلاجي الأسبوعي - {sanitize_input(client.client_serial)} - الأسبوع {week_num}, {year}"
            content = f"""المعالج العزيز،

هذا هو تقرير التقدم الأسبوعي الخاص بي لـ {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}.

العميل: {sanitize_input(client.client_serial)}
الأسبوع: {week_num}, {year}
تسجيلات الحضور المكتملة: {len(checkins)}/7

ملخص تسجيل الحضور اليومي:
"""
        else:  # English
            subject = f"Weekly Therapy Report - {sanitize_input(client.client_serial)} - Week {week_num}, {year}"
            content = f"""Dear {therapist_name},

Here is my weekly progress report for {get_translated_month(week_start, lang)} {week_start.day} - {get_translated_month(week_end, lang)} {week_end.day}, {year}.

CLIENT: {sanitize_input(client.client_serial)}
WEEK: {week_num}, {year}
CHECK-INS COMPLETED: {len(checkins)}/7

DAILY CHECK-IN SUMMARY:
"""

        # Add daily details with proper translations
        for i in range(7):
            current_date = week_start + timedelta(days=i)
            checkin = next((c for c in checkins if c.checkin_date == current_date), None)

            content += f"\n{days[i]} ({current_date.strftime('%m/%d')}):\n"

            if checkin:
                content += f"  ✓ {trans('checkin_time')}: {checkin.checkin_time.strftime('%H:%M')}\n"

                # Get translated category responses
                category_responses = CategoryResponse.query.filter_by(
                    client_id=client.id,
                    response_date=current_date
                ).all()

                for response in category_responses:
                    cat_name = translate_category_name(response.category.name, lang)
                    content += f"  - {cat_name}: {response.value}/5"
                    if response.notes:
                        content += f" ({trans('notes')}: {escape(response.notes)})"
                    content += "\n"
            else:
                content += f"  ✗ {trans('no_checkin')}\n"

        # Add summary with translations
        if checkins:
            content += f"\n{trans('weekly_summary').upper()}:\n"

            total_checkins = len(checkins)

            # Calculate averages from category responses
            for category in TrackingCategory.query.all():
                responses = []
                for checkin in checkins:
                    response = CategoryResponse.query.filter_by(
                        client_id=client.id,
                        category_id=category.id,
                        response_date=checkin.checkin_date
                    ).first()
                    if response:
                        responses.append(response.value)

                if responses:
                    avg_value = sum(responses) / len(responses)
                    cat_name = translate_category_name(category.name, lang)
                    content += f"- {cat_name}: {avg_value:.1f}/5\n"

            content += f"- {trans('completion_rate')}: {total_checkins}/7 {trans('days')} ({(total_checkins / 7) * 100:.0f}%)\n"

        # Add goals if any
        weekly_goals = client.goals.filter_by(week_start=week_start.date()).all()
        if weekly_goals:
            content += f"\n{trans('weekly_goals').upper()}:\n"
            for goal in weekly_goals:
                completions = goal.completions.filter(
                    GoalCompletion.completion_date.between(week_start.date(), week_end.date())
                ).all()
                completed_days = sum(1 for c in completions if c.completed)
                content += f"- {escape(goal.goal_text)}: {trans('completed')} {completed_days}/7 {trans('days')}\n"

        # Add footer with proper translation
        if lang == 'he':
            content += f"\nהדוח נוצר בתאריך: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\nבברכה,\nמטופל {sanitize_input(client.client_serial)}"
        elif lang == 'ru':
            content += f"\nОтчет создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\nС уважением,\nКлиент {sanitize_input(client.client_serial)}"
        elif lang == 'ar':
            content += f"\nتم إنشاء التقرير في: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\nمع أطيب التحيات،\nالعميل {sanitize_input(client.client_serial)}"
        else:
            content += f"\nReport generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\nBest regards,\nClient {sanitize_input(client.client_serial)}"

        return jsonify({
            'success': True,
            'recipient': therapist_email,
            'subject': subject,
            'email_content': content,
            'note': 'Copy this email content to send to your therapist'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/week-checkins/<week>', methods=['GET'])
@require_auth(['client'])
def get_client_week_checkins(week):
    """Get client's check-ins for a specific week"""
    try:
        client = request.current_user.client

        # Parse week
        # Validate week format
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week dates
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)

        # Get check-ins
        checkins = client.checkins.filter(
            DailyCheckin.checkin_date.between(week_start, week_end)
        ).all()

        # Format response
        checkin_data = {}
        for checkin in checkins:
            checkin_data[checkin.checkin_date.isoformat()] = {
                'time': checkin.checkin_time.strftime('%H:%M'),
                'emotional': checkin.emotional_value,
                'emotional_notes': checkin.emotional_notes,
                'medication': checkin.medication_value,
                'medication_notes': checkin.medication_notes,
                'activity': checkin.activity_value,
                'activity_notes': checkin.activity_notes
            }

        return jsonify({
            'success': True,
            'week_start': week_start.date().isoformat(),
            'week_end': week_end.date().isoformat(),
            'checkins': checkin_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/goals/<week>', methods=['GET'])
@require_auth(['client'])
def get_client_week_goals(week):
    """Get client's goals for a specific week"""
    try:
        client = request.current_user.client

        # Parse week
        import re
        week_pattern = re.compile(r'^\d{4}-W\d{2}$')
        if not week_pattern.match(week):
            return jsonify({'error': 'Invalid week format. Use YYYY-Wnn'}), 400

        # Parse week
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Validate ranges
        if year < 2020 or year > 2030:
            return jsonify({'error': 'Invalid year'}), 400
        if week_num < 1 or week_num > 53:
            return jsonify({'error': 'Invalid week number'}), 400

        # Calculate week start
        jan1 = datetime(year, 1, 1)
        days_to_monday = (7 - jan1.weekday()) % 7
        if days_to_monday == 0:
            days_to_monday = 7
        first_monday = jan1 + timedelta(days=days_to_monday - 7)
        week_start = first_monday + timedelta(weeks=week_num - 1)

        # Get goals for the week
        goals = client.goals.filter_by(
            week_start=week_start.date(),
            is_active=True
        ).all()

        # Format response
        goals_data = []
        for goal in goals:
            # Get completions for the week
            completions = {}
            for i in range(7):
                day = week_start.date() + timedelta(days=i)
                completion = goal.completions.filter_by(completion_date=day).first()
                completions[day.isoformat()] = completion.completed if completion else None

            goals_data.append({
                'id': goal.id,
                'text': goal.goal_text,
                'week_start': goal.week_start.isoformat(),
                'completions': completions
            })

        return jsonify({
            'success': True,
            'goals': goals_data
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/data-retention', methods=['POST'])
@require_auth(['therapist'])  # Should be admin role
def apply_data_retention():
    """Apply data retention policy - GDPR/HIPAA compliance"""
    try:
        # Only allow specific admin therapists
        if request.current_user.email not in app.config.get('ADMIN_EMAILS', []):
            return jsonify({'error': 'Unauthorized'}), 403

        retention_days = 2555  # 7 years for HIPAA
        cutoff_date = date.today() - timedelta(days=retention_days)

        # Archive old check-ins
        old_checkins = DailyCheckin.query.filter(
            DailyCheckin.checkin_date < cutoff_date
        ).all()

        archived_count = 0
        for checkin in old_checkins:
            # Archive to cold storage (implement based on your needs)
            archive_data = {
                'client_id': checkin.client_id,
                'date': checkin.checkin_date.isoformat(),
                'data': {
                    'emotional': checkin.emotional_value,
                    'medication': checkin.medication_value,
                    'activity': checkin.activity_value
                }
            }
            # TODO: Send to archive storage

            # Delete from active database
            db.session.delete(checkin)
            archived_count += 1

        db.session.commit()

        # Audit log
        log_audit(
            action='DATA_RETENTION_APPLIED',
            details={
                'cutoff_date': cutoff_date.isoformat(),
                'archived_count': archived_count
            }
        )

        return jsonify({
            'success': True,
            'archived': archived_count
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/request-deletion', methods=['POST'])
@require_auth(['client'])
def request_data_deletion():
    """GDPR right to deletion"""
    try:
        client = request.current_user.client

        # Create deletion request (implement approval workflow)
        deletion_request = {
            'client_id': client.id,
            'requested_at': datetime.utcnow(),
            'status': 'pending'
        }

        # Notify therapist
        # TODO: Send notification

        if client.therapist and client.therapist.user:
            therapist_email = client.therapist.user.email
            subject = f"Data Deletion Request - Client {client.client_serial}"

            body = f"""Dear {client.therapist.name},

        Your client {client.client_name or client.client_serial} has requested deletion of their data.

        Client ID: {client.client_serial}
        Request Date: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}

        This is a GDPR data deletion request. Please review and process according to your data retention policies.

        To approve or discuss this request, please contact the client directly.

        Best regards,
        Therapeutic Companion System"""

            html_body = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif;">
                        <h2>Data Deletion Request</h2>
                        <p>Dear {client.therapist.name},</p>
                        <p>Your client <strong>{client.client_name or client.client_serial}</strong> has requested deletion of their data.</p>
                        <div style="background-color: #f5f5f5; padding: 15px; margin: 20px 0;">
                            <p><strong>Client ID:</strong> {client.client_serial}</p>
                            <p><strong>Request Date:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}</p>
                        </div>
                        <p style="color: #d9534f; font-weight: bold;">This is a GDPR data deletion request.</p>
                        <p>Please review and process according to your data retention policies.</p>
                        <p>To approve or discuss this request, please contact the client directly.</p>
                        <hr>
                        <p>Best regards,<br>Therapeutic Companion System</p>
                    </body>
                    </html>
                    """

            # Send the email
            send_email(therapist_email, subject, body, html_body)


        # Audit log
        log_audit(
            action='DATA_DELETION_REQUESTED',
            resource_type='client',
            resource_id=client.id
        )

        return jsonify({
            'success': True,
            'message': 'Deletion request submitted. Your therapist will be notified.'
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500





# The create_weekly_report_excel function continues from Part 1
# It's already complete in Part 1, starting at line 2324


# ============= REPORT GENERATION =============


# The email_therapy_report function is already complete in Part 1
# It's the UPDATED version starting at line 2773


# ============= ERROR HANDLERS =============

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({'error': 'Resource not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    db.session.rollback()
    return jsonify({'error': 'Internal server error'}), 500


@app.errorhandler(Exception)
def handle_exception(error):
    """Handle unexpected exceptions"""
    db.session.rollback()
    app.logger.error(f"Unexpected error: {str(error)}")
    return jsonify({'error': 'An unexpected error occurred'}), 500


@app.route('/api/unsubscribe/<token>', methods=['GET', 'POST'])
def unsubscribe(token):
    """Handle email unsubscribe requests"""
    try:
        # Decode the token
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        email_type = payload.get('email_type', 'all')

        if not user_id:
            return jsonify({'error': 'Invalid unsubscribe link'}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        if request.method == 'POST':
            # Process unsubscribe
            if email_type == 'reminders' and user.client:
                # Disable reminders
                reminders = user.client.reminders.filter_by(is_active=True).all()
                for reminder in reminders:
                    reminder.is_active = False
                db.session.commit()

                return jsonify({
                    'success': True,
                    'message': 'You have been unsubscribed from reminder emails'
                })
            else:
                # Mark email as invalid for all emails
                user.email_valid = False
                db.session.commit()

                return jsonify({
                    'success': True,
                    'message': 'You have been unsubscribed from all emails'
                })

        # GET request - show unsubscribe page
        return f"""
        <html>
        <head>
            <title>Unsubscribe - Therapeutic Companion</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; text-align: center; }}
                .container {{ max-width: 500px; margin: 0 auto; }}
                .btn {{ background: #dc3545; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; }}
                .btn:hover {{ background: #c82333; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Unsubscribe from Emails</h2>
                <p>Are you sure you want to unsubscribe from {escape(email_type)} emails?</p>
                <p>Email: {escape(user.email)}</p>
                <form method="POST">
                    <button type="submit" class="btn">Confirm Unsubscribe</button>
                </form>
                <p style="margin-top: 20px;"><a href="/">Return to site</a></p>
            </div>
        </body>
        </html>
        """

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Unsubscribe link has expired'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def generate_unsubscribe_token(user_id, email_type='all'):
    """Generate unsubscribe token"""
    payload = {
        'user_id': user_id,
        'email_type': email_type,
        'exp': datetime.utcnow() + timedelta(days=30)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


# Import and configure Celery
celery = None
try:
    from celery_app import celery, send_reminder_test, send_email_task, process_email_queue_task

    # Log Celery availability
    if celery:
        logger.info("Celery successfully imported and configured")

        # Test Redis connection
        try:
            celery.backend.get('test')
            logger.info("Redis connection successful")
        except Exception as e:
            logger.warning(f"Redis connection test failed: {e}")

except ImportError as e:
    logger.warning(f"Celery not available: {e}")
    celery = None
except Exception as e:
    logger.error(f"Error initializing Celery: {e}")
    celery = None


    # Add Celery context to Flask
    class FlaskCelery(Celery):
        def __init__(self, *args, **kwargs):
            super(FlaskCelery, self).__init__(*args, **kwargs)
            self.patch_task()

        def patch_task(self):
            TaskBase = self.Task
            _celery = self

            class ContextTask(TaskBase):
                abstract = True

                def __call__(self, *args, **kwargs):
                    with app.app_context():
                        return TaskBase.__call__(self, *args, **kwargs)

            self.Task = ContextTask


    # Update celery configuration with Flask app config
    celery.conf.update(app.config)

except ImportError:
    logger.warning("Celery not available, background tasks disabled")
    celery = None


# Helper function for sending reminder emails
def send_single_reminder_email_sync(client):
    """Synchronous version of reminder email sending"""
    try:
        base_url = os.environ.get('APP_BASE_URL', 'https://therapy-companion.onrender.com')

        # Get the reminder to check for custom email
        reminder = client.reminders.filter_by(
            reminder_type='daily_checkin',
            is_active=True
        ).first()

        # Use reminder email if set and not empty, otherwise use user email
        email_to_use = client.user.email  # Default to user email
        if reminder and reminder.reminder_email and reminder.reminder_email.strip():
            email_to_use = reminder.reminder_email.strip()

        app.logger.info(f"Sending reminder to {email_to_use} for client {sanitize_input(client.client_serial)}")

        subject = "Daily Check-in Reminder - Therapeutic Companion"

        client_name = client.client_name if client.client_name else client.client_serial

        body = f"""Hello,

        This is your daily reminder to complete your therapy check-in.

        Your therapist is tracking your progress, and your daily input is valuable for your treatment.

        Click here to log in and complete today's check-in:
        {base_url}/login.html

        Client ID: {sanitize_input(client_name)}

If you've already completed today's check-in, please disregard this message.

Best regards,
Your Therapy Team"""

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #2c3e50;">Daily Check-in Reminder</h2>
                <p>Hello,</p>
                <p>This is your daily reminder to complete your therapy check-in.</p>
                <p>Your therapist is tracking your progress, and your daily input is valuable for your treatment.</p>
                <div style="text-align: center; margin: 40px 0;">
                    <a href="{base_url}/login.html"
                       style="background-color: #4CAF50; color: white; padding: 15px 40px;
                              text-decoration: none; border-radius: 5px; display: inline-block;
                              font-weight: bold; font-size: 16px;">
                        Complete Today's Check-in
                    </a>
                </div>
                <p style="color: #666; font-size: 14px;">Client ID: {sanitize_input(client_name)}</p>
                <p style="color: #666; font-size: 14px;">
                    If you've already completed today's check-in, please disregard this message.
                </p>
                <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                <p style="color: #999; font-size: 12px;">
                    Best regards,<br>
                    Your Therapy Team
                </p>
            </div>
        </body>
        </html>
        """

        return send_email(email_to_use, subject, body, html_body)

    except Exception as e:
        logger.error(f"Error sending reminder email: {str(e)}")
        return False


@app.route('/api/admin/celery-status', methods=['GET'])
@require_auth(['therapist'])
def celery_status():
    """Check Celery worker status"""
    if not celery:
        return jsonify({'status': 'disabled', 'message': 'Celery not configured'})

    try:
        # Try to inspect active workers
        inspector = celery.control.inspect()
        stats = inspector.stats()

        if stats:
            worker_count = len(stats)
            return jsonify({
                'status': 'active',
                'workers': worker_count,
                'stats': stats
            })
        else:
            return jsonify({
                'status': 'offline',
                'message': 'No active workers found'
            })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        })


@app.route('/api/admin/test-celery', methods=['POST'])
@require_auth(['therapist'])
def test_celery():
    """Test Celery by sending a test email"""
    if not celery:
        return jsonify({'error': 'Celery not configured'}), 503

    try:
        email = request.json.get('email', request.current_user.email)

        # Queue the test task
        from celery_app import send_reminder_test
        task = send_reminder_test.delay(email)

        return jsonify({
            'success': True,
            'task_id': task.id,
            'message': f'Test email queued for {email}'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/queue-stats', methods=['GET'])
@require_auth(['therapist'])
def queue_stats():
    """Get queue statistics"""
    if not celery:
        return jsonify({'error': 'Celery not configured'}), 503

    try:
        inspector = celery.control.inspect()

        # Get various stats
        active_tasks = inspector.active()
        scheduled_tasks = inspector.scheduled()
        reserved_tasks = inspector.reserved()

        # Count tasks
        total_active = sum(len(tasks) for tasks in (active_tasks or {}).values())
        total_scheduled = sum(len(tasks) for tasks in (scheduled_tasks or {}).values())
        total_reserved = sum(len(tasks) for tasks in (reserved_tasks or {}).values())

        return jsonify({
            'active': total_active,
            'scheduled': total_scheduled,
            'reserved': total_reserved,
            'workers': list((active_tasks or {}).keys())
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/api/admin/db-pool-status', methods=['GET'])
@require_auth(['therapist'])
def db_pool_status():
    """Monitor database connection pool health"""
    try:
        # Get pool statistics
        pool = db.engine.pool

        return jsonify({
            'size': pool.size(),
            'checked_in_connections': pool.checkedin(),
            'overflow': pool.overflow(),
            'total': pool.size() + pool.overflow(),
            'status': 'healthy' if pool.checkedin() > 0 else 'warning'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============= STATIC FILE SERVING =============

@app.route('/reset-password.html')
def reset_password_page():
    """Serve the password reset page"""
    try:
        file_path = os.path.join(BASE_DIR, 'reset-password.html')
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            app.logger.error(f"reset-password.html not found at {file_path}")
            return "reset-password.html not found", 404
    except Exception as e:
        app.logger.error(f"Error serving reset-password.html: {e}")
        return f"Error: {str(e)}", 500


# ============= CLIENT REMINDER ENDPOINTS =============

@app.route('/api/client/test-reminder', methods=['POST'])
@require_auth(['client'])
def test_client_reminder():
    """Test reminder for client"""
    try:
        # Add debug logging at the start
        logger.info('test_reminder_start', extra={
            'extra_data': {
                'celery_available': celery is not None,
                'user_id': request.current_user.id,
                'user_email': request.current_user.email,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        if not celery:
            logger.error('test_reminder_no_celery', extra={
                'extra_data': {
                    'error': 'Celery not configured',
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })
            return jsonify({'error': 'Celery not configured'}), 503

        client = request.current_user.client

        # Handle both JSON and non-JSON requests
        email = None
        if request.content_type == 'application/json':
            try:
                json_data = request.get_json(force=True, silent=True)
                if json_data:
                    email = json_data.get('email')
            except:
                pass

        # Use the client's reminder email if configured, otherwise use account email
        if not email:
            reminder = client.reminders.filter_by(
                reminder_type='daily_checkin',
                is_active=True
            ).first()

            if reminder and reminder.reminder_email:
                email = reminder.reminder_email
            else:
                email = request.current_user.email

        logger.info('test_reminder_email_determined', extra={
            'extra_data': {
                'email': email,
                'has_reminder': reminder is not None if 'reminder' in locals() else False,
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        })

        # Queue the test task
        try:
            from celery_app import send_reminder_test
            logger.info('test_reminder_importing_celery', extra={
                'extra_data': {
                    'import_successful': True,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })

            task = send_reminder_test.delay(email, client_id=client.id)

            logger.info('test_reminder_task_queued', extra={
                'extra_data': {
                    'task_id': task.id,
                    'email': email,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            })

        except ImportError as ie:
            logger.error('test_reminder_import_error', extra={
                'extra_data': {
                    'error': str(ie),
                    'error_type': 'ImportError',
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            }, exc_info=True)
            return jsonify({'error': f'Celery import error: {str(ie)}'}), 500

        except Exception as ce:
            logger.error('test_reminder_celery_error', extra={
                'extra_data': {
                    'error': str(ce),
                    'error_type': type(ce).__name__,
                    'request_id': g.request_id
                },
                'request_id': g.request_id,
                'user_id': request.current_user.id
            }, exc_info=True)
            return jsonify({'error': f'Celery error: {str(ce)}'}), 500

        return jsonify({
            'success': True,
            'task_id': task.id,
            'message': f'Test email queued for {email}'
        })

    except Exception as e:
        logger.error('test_reminder_error', extra={
            'extra_data': {
                'error': str(e),
                'error_type': type(e).__name__,
                'traceback': traceback.format_exc(),
                'request_id': g.request_id
            },
            'request_id': g.request_id,
            'user_id': request.current_user.id
        }, exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/client/reminder-queue-status', methods=['GET'])
@require_auth(['client'])
def client_queue_status():
    """Get queue status for client"""
    if not celery:
        return jsonify({'error': 'Celery not configured'}), 503

    try:
        inspector = celery.control.inspect()

        # Get various stats
        active_tasks = inspector.active()
        scheduled_tasks = inspector.scheduled()
        reserved_tasks = inspector.reserved()

        # Count tasks
        total_active = sum(len(tasks) for tasks in (active_tasks or {}).values())
        total_scheduled = sum(len(tasks) for tasks in (scheduled_tasks or {}).values())
        total_reserved = sum(len(tasks) for tasks in (reserved_tasks or {}).values())

        return jsonify({
            'active': total_active,
            'scheduled': total_scheduled,
            'reserved': total_reserved,
            'workers': list((active_tasks or {}).keys()) if active_tasks else []
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500





# ============= INITIALIZATION =============

# Flag to ensure single initialization
_initialized = False





# Don't initialize on import for production
# Let init_db.py handle it during deployment
if not os.environ.get('PRODUCTION'):
    with app.app_context():
        initialize_database()

# ============= MAIN ENTRY POINT =============

if __name__ == '__main__':
    # Ensure database is created when running directly
    with app.app_context():
        db.create_all()
        initialize_database()
        migrate_existing_data_if_needed()
        ensure_client_names()  # Ensure all clients have names

    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=not os.environ.get('PRODUCTION'))
